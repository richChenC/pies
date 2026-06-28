"""
数据导出器
支持导出Excel、CSV等格式
"""
import pandas as pd
from pathlib import Path
from typing import List, Dict
import logging

from .models import ProbeRecord, ProbeStatistics

logger = logging.getLogger(__name__)


class DataExporter:
    """数据导出器"""

    @staticmethod
    def _format_datetime(value) -> str:
        return value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
    
    def export_records_to_excel(self, records: List[ProbeRecord], output_path: str):
        """导出记录到Excel"""
        try:
            data = []
            for record in records:
                # 使用原始值显示，而不是转换后的值
                probe_type_display = (
                    record.probe_type_raw
                    if hasattr(record, 'probe_type_raw') and record.probe_type_raw
                    else str(getattr(record.probe_type, 'value', record.probe_type) or '')
                )
                
                data.append({
                    '大修': record.outage if hasattr(record, 'outage') else '',  # 原始数据（如D223）
                    '蒸汽发生器编号': record.sg_id if hasattr(record, 'sg_id') else '',  # 原始数据
                    '数据组': record.data_group,  # 原始数据
                    '操作员': record.operator,  # 原始数据
                    '探头类型': probe_type_display,  # 原始数据（如BOBBIN）
                    '探头编码': record.probe_sn,  # 原始数据
                    '探头型号': record.model,  # 原始数据
                    '管道数量': record.tube_number,  # 原始数据
                    '开始时间': self._format_datetime(record.start_time),  # 原始数据
                    '结束时间': self._format_datetime(record.end_time),  # 原始数据
                    '使用时长(分钟)': round(record.duration_minutes, 2),
                    '警告信息': '; '.join(record.warnings) if record.warnings else ''  # 警告信息
                })
            
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='探头记录')
                
                worksheet = writer.sheets['探头记录']
                
                # 设置列宽
                column_widths = {
                    'A': 9, 'B': 17, 'C': 14, 'D': 11,
                    'E': 13, 'F': 15, 'G': 17, 'H': 11, 'I': 19, 'J': 19, 'K': 15, 'L': 20
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置数据对齐
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"成功导出 {len(records)} 条记录到 {output_path}")
            
        except Exception as e:
            logger.error(f"导出Excel时出错: {str(e)}")
            raise
    
    def export_statistics_to_excel(self, statistics: Dict[str, ProbeStatistics], output_path: str):
        """导出统计信息到Excel"""
        try:
            data = []
            for stat in statistics.values():
                data.append({
                    '探头编码': stat.probe_sn,
                    '探头类型': stat.probe_type,
                    '探头型号': stat.model,
                    '总使用次数': stat.total_uses,
                    '总使用时长(小时)': round(stat.lifetime_hours, 2),
                    '总使用时长(分钟)': round(stat.total_duration_minutes, 2),
                    '首次使用时间': stat.first_use_time.strftime('%Y-%m-%d %H:%M:%S') if stat.first_use_time else '',
                    '最后使用时间': stat.last_use_time.strftime('%Y-%m-%d %H:%M:%S') if stat.last_use_time else '',
                    '使用段数': stat.continuous_use_count,
                    '最长连续使用时长(小时)': round(stat.longest_continuous_duration_hours, 2),
                    '最长连续使用时长(分钟)': round(stat.longest_continuous_duration_minutes, 2),
                    '管道数量': stat.unique_tube_count,
                    '是否完全连续使用': '是' if stat.is_continuous_use else '否'
                })
            
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='探头统计')
                
                worksheet = writer.sheets['探头统计']
                
                # 设置列宽
                column_widths = {
                    'A': 15, 'B': 11, 'C': 13, 'D': 15, 'E': 17,
                    'F': 19, 'G': 19, 'H': 13, 'I': 19, 'J': 19,
                    'K': 13, 'L': 15
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置数据对齐
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"成功导出 {len(statistics)} 条统计信息到 {output_path}")
            
        except Exception as e:
            logger.error(f"导出统计信息时出错: {str(e)}")
            raise
    
    def export_to_csv(self, records: List[ProbeRecord], output_path: str):
        """导出记录到CSV"""
        try:
            data = []
            for record in records:
                # 使用原始值显示，而不是转换后的值
                probe_type_display = (
                    record.probe_type_raw
                    if hasattr(record, 'probe_type_raw') and record.probe_type_raw
                    else str(getattr(record.probe_type, 'value', record.probe_type) or '')
                )
                
                data.append({
                    '大修': record.outage if hasattr(record, 'outage') else '',  # 原始数据（如D223）
                    '蒸汽发生器编号': record.sg_id if hasattr(record, 'sg_id') else '',  # 原始数据
                    '数据组': record.data_group,  # 原始数据
                    '操作员': record.operator,  # 原始数据
                    '探头类型': probe_type_display,  # 原始数据（如BOBBIN）
                    '探头编码': record.probe_sn,  # 原始数据
                    '探头型号': record.model,  # 原始数据
                    '管道数量': record.tube_number,  # 原始数据
                    '开始时间': self._format_datetime(record.start_time),  # 原始数据
                    '结束时间': self._format_datetime(record.end_time),  # 原始数据
                    '使用时长(分钟)': round(record.duration_minutes, 2)
                })
            
            df = pd.DataFrame(data)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            logger.info(f"成功导出 {len(records)} 条记录到 {output_path}")
            
        except Exception as e:
            logger.error(f"导出CSV时出错: {str(e)}")
            raise
