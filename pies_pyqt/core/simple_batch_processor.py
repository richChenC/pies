"""
简化版批量处理器
避免pandas依赖，使用内置库处理Excel导出
"""
import os
import re
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import logging
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict
import csv

from .sum_parser import SumFileParser, SumRecord

logger = logging.getLogger(__name__)

@dataclass
class EnhancedRecord:
    """增强版记录，包含更多从文件名和路径提取的信息"""
    # 基础SUM记录信息
    sum_record: SumRecord = field(default_factory=SumRecord)
    
    # 从文件夹结构提取的信息
    steam_generator: str = ""       # 蒸汽发生器编号 (SG1, SG2, SG3)
    data_group_id: str = ""         # 数据组编号 (00001-00999)
    folder_path: str = ""           # 完整文件夹路径
    
    # 从文件名提取的信息
    analyst_code: str = ""          # 分析员代码 (DBY, KYY, WCL, ZZB等)
    priority_code: str = ""         # 优先级代码 (PRI, TER, SEC等)
    report_type: str = ""           # 报告类型 (RES, ACQ等)
    
    # RPT文件相关信息
    rpt_files: List[str] = field(default_factory=list)  # 关联的RPT文件列表
    rpt_content: Dict[str, str] = field(default_factory=dict)  # RPT文件内容摘要

class SimpleBatchProcessor:
    """简化版批量处理器"""
    
    def __init__(self):
        self.sum_parser = SumFileParser()
        self.supported_extensions = {'.sum', '.xml'}
        self.rpt_extensions = {'.rpt'}
        
        # 文件名模式匹配
        self.analyst_pattern = re.compile(r'-([A-Z]{3})-')  # 匹配分析员代码
        self.priority_pattern = re.compile(r'-(PRI|TER|SEC)(?:\.|$)')  # 匹配优先级
        self.sg_pattern = re.compile(r'SG(\d+)')  # 匹配蒸汽发生器编号
        self.group_pattern = re.compile(r'(\d{5,6})$')  # 匹配数据组编号
    
    def process_root_directory(self, root_path: Path, progress_callback=None) -> List[EnhancedRecord]:
        """
        处理根目录，递归遍历所有子文件夹
        
        Args:
            root_path: 根目录路径
            progress_callback: 进度回调函数
            
        Returns:
            增强记录列表
        """
        if not root_path.exists() or not root_path.is_dir():
            logger.error(f"根目录不存在或不是有效目录: {root_path}")
            return []
        
        all_records = []
        
        # 查找所有符合模式的数据文件夹
        data_folders = self._find_data_folders(root_path)
        total_folders = len(data_folders)
        
        logger.info(f"找到 {total_folders} 个数据文件夹")
        
        for i, folder_path in enumerate(data_folders):
            if progress_callback:
                progress_callback(i, total_folders, f"处理文件夹: {folder_path.name}")
            
            try:
                folder_records = self._process_data_folder(folder_path)
                all_records.extend(folder_records)
                logger.info(f"文件夹 {folder_path.name} 处理完成，获得 {len(folder_records)} 条记录")
            except Exception as e:
                logger.error(f"处理文件夹 {folder_path} 时出错: {e}")
                continue
        
        logger.info(f"批量处理完成，总共获得 {len(all_records)} 条记录")
        return all_records
    
    def _find_data_folders(self, root_path: Path) -> List[Path]:
        """
        查找所有符合模式的数据文件夹
        模式: SG[1-9]XXXXX[00001-99999]
        """
        data_folders = []
        
        # 递归查找所有文件夹
        for item in root_path.rglob('*'):
            if item.is_dir():
                folder_name = item.name
                
                # 检查是否符合数据文件夹命名模式
                if self._is_data_folder(folder_name):
                    # 检查文件夹中是否包含.sum文件
                    if self._has_sum_files(item):
                        data_folders.append(item)
        
        # 按文件夹名称排序
        data_folders.sort(key=lambda x: x.name)
        return data_folders
    
    def _is_data_folder(self, folder_name: str) -> bool:
        """
        检查文件夹名称是否符合数据文件夹模式
        例如: SG1CPSICAL00001, SG2CPSICAL00002等
        """
        # 基本模式：以SG开头，包含数字
        if not folder_name.startswith('SG'):
            return False
        
        # 检查是否包含蒸汽发生器编号
        sg_match = self.sg_pattern.search(folder_name)
        if not sg_match:
            return False
        
        # 检查是否以数字结尾（数据组编号）
        if not re.search(r'\d{4,6}$', folder_name):
            return False
        
        return True
    
    def _has_sum_files(self, folder_path: Path) -> bool:
        """检查文件夹中是否包含.sum文件"""
        for file_path in folder_path.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in self.supported_extensions:
                return True
        return False
    
    def _process_data_folder(self, folder_path: Path) -> List[EnhancedRecord]:
        """
        处理单个数据文件夹
        
        Args:
            folder_path: 数据文件夹路径
            
        Returns:
            该文件夹的增强记录列表
        """
        records = []
        folder_name = folder_path.name
        
        # 从文件夹名称提取信息
        steam_generator = self._extract_steam_generator(folder_name)
        data_group_id = self._extract_data_group_id(folder_name)
        
        # 查找所有.sum文件
        sum_files = list(folder_path.glob('*.sum')) + list(folder_path.glob('*.SUM'))
        
        # 查找所有.rpt文件
        rpt_files = list(folder_path.glob('*.rpt')) + list(folder_path.glob('*.RPT'))
        
        logger.info(f"文件夹 {folder_name}: 找到 {len(sum_files)} 个SUM文件, {len(rpt_files)} 个RPT文件")
        
        for sum_file in sum_files:
            try:
                # 解析SUM文件
                sum_record = self.sum_parser.parse_file(sum_file)
                if not sum_record:
                    continue
                
                # 创建增强记录
                enhanced_record = EnhancedRecord()
                enhanced_record.sum_record = sum_record
                enhanced_record.steam_generator = steam_generator
                enhanced_record.data_group_id = data_group_id
                enhanced_record.folder_path = str(folder_path)
                
                # 从文件名提取分析员和优先级信息
                self._extract_file_info(sum_file.name, enhanced_record)
                
                # 查找相关的RPT文件
                related_rpt_files = self._find_related_rpt_files(sum_file, rpt_files)
                enhanced_record.rpt_files = [str(f) for f in related_rpt_files]
                
                # 从RPT文件名中提取分析员和优先级信息
                self._extract_info_from_rpt_files(enhanced_record)
                
                # 解析RPT文件内容摘要
                for rpt_file in related_rpt_files:
                    rpt_summary = self._parse_rpt_file_summary(rpt_file)
                    if rpt_summary:
                        enhanced_record.rpt_content[rpt_file.name] = rpt_summary
                
                records.append(enhanced_record)
                
            except Exception as e:
                logger.error(f"处理SUM文件 {sum_file} 时出错: {e}")
                continue
        
        return records
    
    def _extract_steam_generator(self, folder_name: str) -> str:
        """从文件夹名称提取蒸汽发生器编号"""
        match = self.sg_pattern.search(folder_name)
        return f"SG{match.group(1)}" if match else ""
    
    def _extract_data_group_id(self, folder_name: str) -> str:
        """从文件夹名称提取数据组编号"""
        match = re.search(r'(\d{4,6})$', folder_name)
        return match.group(1) if match else ""
    
    def _extract_file_info(self, filename: str, record: EnhancedRecord):
        """从文件名提取分析员和优先级信息"""
        # 提取分析员代码
        analyst_match = self.analyst_pattern.search(filename)
        if analyst_match:
            record.analyst_code = analyst_match.group(1)
        
        # 提取优先级代码
        priority_match = self.priority_pattern.search(filename)
        if priority_match:
            record.priority_code = priority_match.group(1)
        
        # 提取报告类型
        if '-RES' in filename:
            record.report_type = 'RES'
        elif '-ACQ' in filename:
            record.report_type = 'ACQ'
        elif '-TER' in filename:
            record.report_type = 'TER'
    
    def _extract_info_from_rpt_files(self, record: EnhancedRecord):
        """从关联的RPT文件中提取分析员和优先级信息"""
        for rpt_file_path in record.rpt_files:
            rpt_filename = Path(rpt_file_path).name
            
            # 提取分析员代码
            if not record.analyst_code:
                analyst_match = self.analyst_pattern.search(rpt_filename)
                if analyst_match:
                    record.analyst_code = analyst_match.group(1)
            
            # 提取优先级代码
            if not record.priority_code:
                priority_match = self.priority_pattern.search(rpt_filename)
                if priority_match:
                    record.priority_code = priority_match.group(1)
            
            # 提取报告类型
            if not record.report_type:
                if '-RES' in rpt_filename:
                    record.report_type = 'RES'
                elif '-ACQ' in rpt_filename:
                    record.report_type = 'ACQ'
                elif '-TER' in rpt_filename:
                    record.report_type = 'TER'
    
    def _find_related_rpt_files(self, sum_file: Path, rpt_files: List[Path]) -> List[Path]:
        """查找与SUM文件相关的RPT文件"""
        related_files = []
        
        # 获取SUM文件所在的文件夹名称，用于匹配RPT文件
        sum_folder = sum_file.parent.name
        
        for rpt_file in rpt_files:
            rpt_basename = rpt_file.stem
            
            # 方法1: 检查RPT文件名是否包含文件夹名称中的数据组编号
            if sum_folder in rpt_basename:
                related_files.append(rpt_file)
                continue
            
            # 方法2: 检查是否有共同的数据组编号
            sum_group = re.search(r'(\d{5,6})', sum_folder)
            rpt_group = re.search(r'(\d{5,6})', rpt_basename)
            
            if sum_group and rpt_group and sum_group.group(1) == rpt_group.group(1):
                related_files.append(rpt_file)
                continue
        
        return related_files
    
    def _files_are_related(self, sum_name: str, rpt_name: str) -> bool:
        """判断SUM文件和RPT文件是否相关"""
        # 方法1: 检查是否有共同的数据组编号
        sum_group = re.search(r'(\d{5,6})', sum_name)
        rpt_group = re.search(r'(\d{5,6})', rpt_name)
        
        if sum_group and rpt_group and sum_group.group(1) == rpt_group.group(1):
            return True
        
        # 方法2: 检查是否有共同的探头编号或其他标识
        # 这里可以根据实际文件命名规则进行调整
        
        return False
    
    def _parse_rpt_file_summary(self, rpt_file: Path) -> Optional[str]:
        """解析RPT文件，提取关键信息摘要"""
        try:
            with open(rpt_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # 提取关键信息（这里需要根据实际RPT文件格式调整）
            summary_lines = []
            
            # 查找缺陷数量
            defect_match = re.search(r'缺陷数量[：:]\s*(\d+)', content)
            if defect_match:
                summary_lines.append(f"缺陷数量: {defect_match.group(1)}")
            
            # 查找检测管道数量
            tube_match = re.search(r'检测管道[：:]\s*(\d+)', content)
            if tube_match:
                summary_lines.append(f"检测管道: {tube_match.group(1)}")
            
            # 查找其他关键信息
            # 这里可以根据实际RPT文件内容添加更多解析规则
            
            return '; '.join(summary_lines) if summary_lines else None
            
        except Exception as e:
            logger.warning(f"解析RPT文件 {rpt_file} 时出错: {e}")
            return None
    
    def generate_excel_report(self, records: List[EnhancedRecord], output_path: Path) -> bool:
        """
        生成中文 Excel 报告。
        
        Args:
            records: 增强记录列表
            output_path: 输出文件路径
            
        Returns:
            是否成功生成报告
        """
        try:
            # 尝试导入openpyxl，如果失败则使用xlsxwriter
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                use_openpyxl = True
            except ImportError:
                try:
                    import xlsxwriter
                    use_openpyxl = False
                except ImportError:
                    logger.error("需要安装 openpyxl 或 xlsxwriter 来生成Excel文件")
                    return False
            
            # 转换为Excel格式的数据
            data_rows = []
            
            headers = [
                '大修编号', '蒸汽发生器编号', '数据组', '操作员', '探头类型',
                '探头编码', '探头型号', '管道数量', '开始时间', '结束时间'
            ]
            
            for i, record in enumerate(records, 1):
                sum_rec = record.sum_record
                
                row = [
                    sum_rec.outage or '',
                    record.steam_generator or sum_rec.sg_id or '',
                    f"{record.steam_generator}-{record.data_group_id}" if record.steam_generator and record.data_group_id else sum_rec.sg_id or '',
                    sum_rec.operator_name or sum_rec.operator_id or '',
                    sum_rec.probe_type or '',
                    sum_rec.probe_sn or '',
                    sum_rec.model or '',
                    sum_rec.tube_count or 0,
                    sum_rec.start_time.strftime('%Y-%m-%d %H:%M:%S') if sum_rec.start_time else '',
                    sum_rec.end_time.strftime('%Y-%m-%d %H:%M:%S') if sum_rec.end_time else ''
                ]
                
                data_rows.append(row)
            
            if use_openpyxl:
                # 使用openpyxl创建Excel文件
                wb = Workbook()
                ws = wb.active
                ws.title = "探头数据"
                
                # 写入标题行
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # 写入数据行
                for row_idx, row_data in enumerate(data_rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 调整列宽
                column_widths = {
                    'A': 12,
                    'B': 14,
                    'C': 18,
                    'D': 14,
                    'E': 14,
                    'F': 16,
                    'G': 22,
                    'H': 12,
                    'I': 20,
                    'J': 20
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # 保存文件
                wb.save(output_path)
                
            else:
                # 使用xlsxwriter创建Excel文件
                workbook = xlsxwriter.Workbook(str(output_path))
                worksheet = workbook.add_worksheet('探头数据')
                
                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'align': 'center',
                    'bg_color': '#CCCCCC',
                    'border': 1
                })
                
                cell_format = workbook.add_format({
                    'border': 1,
                    'align': 'left'
                })
                
                # 写入标题行
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)
                
                # 写入数据行
                for row_idx, row_data in enumerate(data_rows):
                    for col_idx, value in enumerate(row_data):
                        worksheet.write(row_idx + 1, col_idx, value, cell_format)
                
                # 设置列宽
                column_widths = [12, 14, 18, 14, 14, 16, 22, 12, 20, 20]
                for col, width in enumerate(column_widths):
                    worksheet.set_column(col, col, width)
                
                workbook.close()
            
            logger.info(f"Excel报告已生成: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成Excel报告时出错: {e}")
            return False
    
    def _generate_statistics(self, records: List[EnhancedRecord]) -> List[Dict]:
        """生成统计信息"""
        if not records:
            return []
        
        stats = []
        
        # 按蒸汽发生器统计
        sg_stats = defaultdict(int)
        analyst_stats = defaultdict(int)
        probe_type_stats = defaultdict(int)
        
        for record in records:
            sg_stats[record.steam_generator] += 1
            analyst_stats[record.analyst_code] += 1
            probe_type_stats[record.sum_record.probe_type] += 1
        
        # 蒸汽发生器统计
        for sg, count in sg_stats.items():
            stats.append({
                '统计类型': '蒸汽发生器',
                '项目': sg,
                '数量': count,
                '百分比': f"{count/len(records)*100:.1f}%"
            })
        
        # 分析员统计
        for analyst, count in analyst_stats.items():
            if analyst:  # 只统计有分析员代码的记录
                stats.append({
                    '统计类型': '分析员',
                    '项目': analyst,
                    '数量': count,
                    '百分比': f"{count/len(records)*100:.1f}%"
                })
        
        # 探头类型统计
        for probe_type, count in probe_type_stats.items():
            if probe_type:  # 只统计有探头类型的记录
                stats.append({
                    '统计类型': '探头类型',
                    '项目': probe_type,
                    '数量': count,
                    '百分比': f"{count/len(records)*100:.1f}%"
                })
        
        return stats

# 使用示例
if __name__ == "__main__":
    processor = SimpleBatchProcessor()
    
    # 处理根目录
    root_dir = Path("./test_data")  # 替换为实际的根目录路径
    
    def progress_callback(current, total, message):
        print(f"进度: {current}/{total} - {message}")
    
    records = processor.process_root_directory(root_dir, progress_callback)
    
    if records:
        # 生成CSV报告
        output_file = Path("./enhanced_report.csv")
        success = processor.generate_csv_report(records, output_file)
        
        if success:
            print(f"报告已生成: {output_file}")
        else:
            print("报告生成失败")
    else:
        print("未找到任何记录")
