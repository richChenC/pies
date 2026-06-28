"""
PyQt5 主窗口 - 1:1 还原旧版所有功能、布局、配色、交互
"""
from __future__ import annotations

import json
import logging
import os
import re
import threading
import ast
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei UI', 'Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
matplotlib.rcParams['axes.unicode_minus'] = False

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from PyQt5.QtCore import (
    QObject, QRunnable, QSize, Qt, QThreadPool, QTimer, pyqtSignal, pyqtSlot,
)
from PyQt5.QtGui import QColor, QCursor, QFont, QIcon
from PyQt5.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QDialog,
    QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMainWindow,
    QListWidget, QListWidgetItem, QMenu, QMessageBox, QProgressBar, QPushButton, QScrollArea,
    QSizePolicy, QSplitter, QStatusBar, QTableWidget,
    QTableWidgetItem, QTextEdit, QVBoxLayout, QWidget, QGridLayout,
    QScrollBar,
)

from pies_pyqt.core.extractor import SummaryFileExtractor
from pies_pyqt.core.analyzer import ProbeAnalyzer
from pies_pyqt.core.exporter import DataExporter
from pies_pyqt.core.visualizer import DataVisualizer
from pies_pyqt.core.models import ProbeStatistics, ProbeRecord, normalize_model_name
from pies_pyqt.core.sum_parser import SumFileParser
from pies_pyqt.core.batch_lifetime_analyzer import BatchLifetimeAnalyzer
from pies_pyqt.constants import (
    APP_TITLE, DEFAULT_WINDOW_SIZE, MIN_WINDOW_SIZE, LOGO_PATH,
    HISTORY_STORE_PATH, SUMMARY_TABLE_HEADERS,
    LEGACY_HISTORY_STORE_PATH, CHART_GROUPS, SOFTWARE_NOTES, SAFETY_REMINDERS,
)

logger = logging.getLogger(__name__)


def _normalize_filter_model_name(model: str) -> str:
    normalized = normalize_model_name(model)
    return normalized or "UNKNOWN"

VISUALIZATION_EXCLUDED_MODELS = {"ZRPF-FH/2C", "ZRPC-FH/2C"}
TABLE_HEADER_FILTERABLE_COLUMNS = (
    '大修', '蒸汽发生器编号', '数据组', '操作员', '探头类型', '探头编码', '探头型号',
    '管道数量', '累计管道数量',
)

# ── 配色 ──────────────────────────────────────────────────────
PRIMARY   = "#3F78C5"
PRIMARY_H = "#4D87D5"
PRIMARY_P = "#3568AC"
ACCENT    = "#00B42A"
ACCENT_H  = "#23C343"
WARNING_C = "#FF7D00"
DANGER    = "#F53F3F"
BG        = "#F5F7FA"
PANEL     = "#FFFFFF"
MUTED     = "#EDF1F5"
TEXT_PRI  = "#333333"
TEXT_SEC  = "#666666"
BORDER    = "#E5E6EB"
SIDEBAR   = "#24313D"
SIDEBAR_H = "#304150"
COMBO_ARROW_PATH = (Path(__file__).resolve().parents[1] / "assets" / "combo_arrow.svg").as_posix()

STYLESHEET = f"""
QMainWindow {{ background: {BG}; }}
QWidget#central {{ background: {BG}; }}
QLabel {{ border: none; background: transparent; }}

/* ══════════════ 侧边栏 ══════════════ */
QWidget#sidebar {{ background: {SIDEBAR}; }}

/* ══════════════ 按钮 ══════════════ */
/* 灰色文件选择按钮 */
QPushButton#btn_file {{
    background: #E8ECF0; color: #344054;
    border: 1px solid #D0D5DD;
    border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-family: "Microsoft YaHei";
}}
QPushButton#btn_file:hover {{ background: #DDE3EA; }}

/* 蓝色主按钮（处理文件夹、分析Excel） */
QPushButton#btn_blue, QPushButton#btn_primary {{
    background: {PRIMARY}; color: #FFFFFF;
    border: none; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#btn_blue:hover, QPushButton#btn_primary:hover {{ background: {PRIMARY_H}; }}
QPushButton#btn_blue:pressed, QPushButton#btn_primary:pressed {{ background: {PRIMARY_P}; }}
QPushButton#btn_blue:disabled, QPushButton#btn_primary:disabled {{ background: #A8C0E0; color: #FFFFFF; }}

/* 橙色警告按钮 */
QPushButton#btn_warning_active {{
    background: {WARNING_C}; color: #FFFFFF;
    border: none; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#btn_warning_active:hover {{ background: #E06E00; }}
QPushButton#btn_warning_active:disabled {{ background: #F0C080; color: #FFFFFF; }}

/* 橙色空心警告按钮（无警告时） */
QPushButton#btn_warning {{
    background: #FFF1E8; color: #B54708;
    border: 1px solid #F0A060; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#btn_warning:hover {{ background: #FFE2CC; }}
QPushButton#btn_warning:disabled {{ background: #F5EDE6; color: #C4A484; border-color: #E0C8B0; }}

/* 绿色一键保存按钮 */
QPushButton#btn_accent, QPushButton#btn_green {{
    background: {ACCENT}; color: #FFFFFF;
    border: none; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#btn_accent:hover, QPushButton#btn_green:hover {{ background: {ACCENT_H}; }}

/* 次要按钮 */
QPushButton#btn_secondary {{
    background: #F0F3F7; color: {TEXT_PRI};
    border: 1px solid #D0D5DD; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-family: "Microsoft YaHei";
}}
QPushButton#btn_secondary:hover {{ background: #E4E8ED; }}

/* 红色按钮 */
QPushButton#btn_red {{
    background: {DANGER}; color: #FFFFFF;
    border: none; border-radius: 4px; padding: 5px 14px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#btn_red:hover {{ background: #D93030; }}

/* 对话框内按钮 */
QPushButton#dlg_primary {{
    background: {PRIMARY}; color: #FFFFFF;
    border: none; border-radius: 4px; padding: 6px 20px;
    font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei";
}}
QPushButton#dlg_primary:hover {{ background: {PRIMARY_H}; }}
QPushButton#dlg_secondary {{
    background: #F0F3F7; color: {TEXT_PRI};
    border: 1px solid #D0D5DD; border-radius: 4px; padding: 6px 20px;
    font-size: 13px; font-family: "Microsoft YaHei";
}}
QPushButton#dlg_secondary:hover {{ background: #E4E8ED; }}

/* ══════════════ 卡片/面板 ══════════════ */
QFrame#hero_card {{
    background: {PANEL}; border: 1px solid {BORDER};
    border-radius: 0px;
}}
QFrame#overview_card {{
    border-radius: 4px;
}}
QFrame#table_panel, QFrame#chart_panel {{
    background: {PANEL}; border: 1px solid {BORDER};
    border-radius: 0px;
}}
QFrame#panel_header {{
    background: #F7F8FA; border-bottom: 1px solid {BORDER};
}}

/* ══════════════ 表格 ══════════════ */
QTableWidget {{
    background: {PANEL}; border: none;
    gridline-color: #EAECF0;
    font-size: 12px; font-family: "Microsoft YaHei";
    selection-background-color: #EEF4FB;
    selection-color: {TEXT_PRI};
    alternate-background-color: #F9FAFB;
}}
QTableWidget::item {{ padding: 4px 6px; border: none; }}
QTableWidget::item:selected {{ background: #D6E8FF; color: {TEXT_PRI}; }}
QTableWidget::item:focus {{ border: none; outline: none; }}
QAbstractItemView {{ outline: none; }}
QHeaderView::section {{
    background: #F7F8FA; color: #475467;
    padding: 7px 6px; border: none;
    border-bottom: 2px solid #D0D5DD;
    border-right: 1px solid #EAECF0;
    font-weight: bold; font-size: 12px;
    font-family: "Microsoft YaHei";
}}
QHeaderView::section:hover {{ background: #EEF2F7; }}
QHeaderView::section:last {{ border-right: none; }}

/* ══════════════ 输入框 ══════════════ */
QLineEdit {{
    background: {PANEL}; border: 1px solid {BORDER};
    border-radius: 4px; padding: 4px 8px;
    font-size: 12px; color: {TEXT_PRI};
    font-family: "Microsoft YaHei";
}}
QLineEdit:focus {{ border-color: {PRIMARY}; background: #FAFCFF; }}
QLineEdit:read-only {{ background: #F9FAFB; color: #667085; }}

/* ══════════════ 下拉框 ══════════════ */
QComboBox {{
    background: {PANEL}; border: 1px solid #BFC7D1;
    border-radius: 0px; padding: 2px 24px 2px 7px;
    font-size: 12px; color: {TEXT_PRI};
    font-family: "Microsoft YaHei";
    min-height: 24px;
}}
QComboBox:hover {{ border-color: #8EA6C4; }}
QComboBox:focus {{ border-color: #6E8FB3; }}
QComboBox::drop-down {{
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 20px;
    border-left: 1px solid #BFC7D1;
    background: #EDEFF2;
}}
QComboBox::down-arrow {{
    image: url("{COMBO_ARROW_PATH}");
    width: 8px;
    height: 6px;
}}
QComboBox QAbstractItemView {{
    background: {PANEL}; border: 1px solid {BORDER};
    selection-background-color: #EEF4FB;
    selection-color: {TEXT_PRI};
    color: {TEXT_PRI};
    font-family: "Microsoft YaHei"; font-size: 12px;
    outline: none;
}}
QComboBox QAbstractItemView::item {{ min-height: 24px; padding: 4px 8px; }}
QComboBox QAbstractItemView::item:selected {{ background: #D6E8FF; color: {TEXT_PRI}; }}

QDialog#filter_popup {{
    background: #FFFFFF;
    border: 1px solid #AEB7C2;
}}
QListWidget#filter_list {{
    background: #FFFFFF;
    border: 1px solid #D0D5DD;
    font-family: "Microsoft YaHei";
    font-size: 12px;
    outline: none;
}}
QListWidget#filter_list::item {{ min-height: 22px; padding: 2px 4px; }}
QListWidget#filter_list::item:selected {{ background: #E8F1FF; color: {TEXT_PRI}; }}

/* ══════════════ 进度条 ══════════════ */
QProgressBar {{
    background: #E4E7EC; border: none; border-radius: 4px;
    height: 8px; text-align: center;
}}
QProgressBar::chunk {{ background: {PRIMARY}; border-radius: 4px; }}

/* ══════════════ 状态栏 ══════════════ */
QStatusBar {{
    background: #FFFFFF; color: #475467;
    font-size: 11px; font-family: "Microsoft YaHei";
    border-top: 1px solid {BORDER};
}}

/* ══════════════ 滚动条 ══════════════ */
QScrollBar:vertical {{
    background: #F2F4F7; width: 8px; border-radius: 4px; margin: 0;
}}
QScrollBar::handle:vertical {{
    background: #C1C9D2; border-radius: 4px; min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{ background: #98A2B3; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; border: none; }}
QScrollBar:horizontal {{
    background: #F2F4F7; height: 8px; border-radius: 4px; margin: 0;
}}
QScrollBar::handle:horizontal {{
    background: #C1C9D2; border-radius: 4px; min-width: 24px;
}}
QScrollBar::handle:horizontal:hover {{ background: #98A2B3; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; border: none; }}

/* ══════════════ 分割器 ══════════════ */
QSplitter::handle {{ background: {BORDER}; }}
QSplitter::handle:horizontal {{ width: 5px; }}

/* ══════════════ 对话框 ══════════════ */
QDialog {{ background: #F7F9FC; font-family: "Microsoft YaHei"; }}
QTextEdit {{
    background: {PANEL}; border: none; border-radius: 0px;
    color: {TEXT_PRI}; font-family: "Microsoft YaHei"; font-size: 12px;
    padding: 6px;
}}

/* ══════════════ 菜单 ══════════════ */
QMenuBar {{
    background: {BG}; font-family: "Microsoft YaHei"; font-size: 13px;
    border-bottom: 1px solid {BORDER};
}}
QMenuBar::item {{ padding: 5px 12px; background: transparent; color: {TEXT_PRI}; }}
QMenuBar::item:selected {{ background: #E4EAF2; border-radius: 3px; }}
QMenu {{
    background: {PANEL}; border: 1px solid {BORDER};
    font-family: "Microsoft YaHei"; font-size: 13px;
    padding: 4px 0;
}}
QMenu::item {{ padding: 7px 28px 7px 14px; color: {TEXT_PRI}; }}
QMenu::item:selected {{ background: #EEF4FB; color: {PRIMARY}; }}
QMenu::separator {{ height: 1px; background: {BORDER}; margin: 4px 0; }}
"""


# ─────────────────────────── Worker ───────────────────────────
class _WorkerSignals(QObject):
    progress    = pyqtSignal(str, float)   # text, percent 0-100
    success     = pyqtSignal(dict)
    chart_ready = pyqtSignal(dict)
    error       = pyqtSignal(str)
    sum_done    = pyqtSignal(dict)
    cancelled   = pyqtSignal(str)


class TaskCancelled(Exception):
    """用户取消了后台任务。"""


class _Worker(QRunnable):
    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.signals = _WorkerSignals()
        self._fn = fn
        self._args = args
        self._kwargs = kwargs
        self._cancelled = threading.Event()
        self.setAutoDelete(True)

    def cancel(self):
        self._cancelled.set()

    def is_cancelled(self) -> bool:
        return self._cancelled.is_set()

    def raise_if_cancelled(self):
        if self.is_cancelled():
            raise TaskCancelled("用户已取消当前任务")

    @pyqtSlot()
    def run(self):
        try:
            self._fn(self.signals, self, *self._args, **self._kwargs)
        except TaskCancelled as exc:
            self.signals.cancelled.emit(str(exc))
        except Exception as exc:
            logger.exception("Worker error")
            self.signals.error.emit(f"{exc}\n\n{traceback.format_exc()}")


# ─────────────────────────── 进度对话框 ───────────────────────────
class ProgressDialog(QDialog):
    cancel_requested = pyqtSignal()

    def __init__(self, title: str, parent=None, cancellable: bool = False):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedSize(480, 156 if cancellable else 122)
        self.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint | Qt.WindowTitleHint)
        self.setStyleSheet(f"background: {PANEL};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        self._label = QLabel("准备中...")
        self._label.setFont(QFont("Microsoft YaHei", 10))
        self._label.setStyleSheet(f"color: #344054;")
        self._label.setWordWrap(True)
        layout.addWidget(self._label)

        self._bar = QProgressBar()
        self._bar.setRange(0, 100)
        self._bar.setValue(0)
        self._bar.setTextVisible(False)
        self._bar.setFixedHeight(9)
        layout.addWidget(self._bar)
        self._cancel_btn = None
        if cancellable:
            btn_row = QHBoxLayout()
            btn_row.addStretch(1)
            self._cancel_btn = QPushButton("取消任务")
            self._cancel_btn.setObjectName("dlg_secondary")
            self._cancel_btn.setCursor(QCursor(Qt.PointingHandCursor))
            _apply_btn_style(self._cancel_btn, "dlg_secondary")
            self._cancel_btn.clicked.connect(self._request_cancel)
            btn_row.addWidget(self._cancel_btn)
            layout.addLayout(btn_row)

        # 居中于父窗口
        if parent:
            pg = parent.geometry()
            self.move(
                pg.x() + (pg.width() - self.width()) // 2,
                pg.y() + (pg.height() - self.height()) // 2,
            )

    def set_progress(self, text: str, value: float):
        self._label.setText(text)
        self._bar.setValue(int(max(0, min(100, value))))
        self.repaint()
        QApplication.processEvents()

    def show_and_paint(self, text: str = "准备中...", value: float = 0):
        self.set_progress(text, value)
        self.show()
        self.raise_()
        self.activateWindow()
        self.repaint()
        QApplication.processEvents()

    def _request_cancel(self):
        if self._cancel_btn:
            self._cancel_btn.setEnabled(False)
            self._cancel_btn.setText("正在取消...")
        self._label.setText("正在取消，请稍候...")
        self.cancel_requested.emit()


# ─────────────────────────── 辅助函数 ───────────────────────────
def _lbl(text: str, bold=False, size=9, color=TEXT_PRI, parent=None) -> QLabel:
    w = QLabel(text, parent)
    f = QFont("Microsoft YaHei", size)
    f.setBold(bold)
    w.setFont(f)
    w.setStyleSheet(f"color: {color}; background: transparent;")
    return w


def _btn(text: str, obj_name: str, parent=None) -> QPushButton:
    b = QPushButton(text, parent)
    b.setObjectName(obj_name)
    b.setCursor(QCursor(Qt.PointingHandCursor))
    _apply_btn_style(b, obj_name)
    return b


# 按钮样式映射（直接内联，不依赖QSS继承）
_BTN_STYLES = {
    "btn_file": f"""
        QPushButton {{ background: #E8ECF0; color: #344054; border: 1px solid #D0D5DD;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #DDE3EA; }}
        QPushButton:pressed {{ background: #D0D8E2; }}
    """,
    "btn_primary": f"""
        QPushButton {{ background: {PRIMARY}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: {PRIMARY_H}; }}
        QPushButton:pressed {{ background: {PRIMARY_P}; }}
        QPushButton:disabled {{ background: #A8C0E0; color: #FFFFFF; }}
    """,
    "btn_blue": f"""
        QPushButton {{ background: {PRIMARY}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: {PRIMARY_H}; }}
        QPushButton:pressed {{ background: {PRIMARY_P}; }}
        QPushButton:disabled {{ background: #A8C0E0; color: #FFFFFF; }}
    """,
    "btn_accent": f"""
        QPushButton {{ background: {ACCENT}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: {ACCENT_H}; }}
        QPushButton:pressed {{ background: #009A29; }}
    """,
    "btn_green": f"""
        QPushButton {{ background: {ACCENT}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: {ACCENT_H}; }}
    """,
    "btn_warning": f"""
        QPushButton {{ background: #FFF1E8; color: #B54708; border: 1px solid #F0A060;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #FFE2CC; }}
        QPushButton:disabled {{ background: #F5EDE6; color: #C4A484; border-color: #E0C8B0; }}
    """,
    "btn_warning_active": f"""
        QPushButton {{ background: {WARNING_C}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #E06E00; }}
        QPushButton:disabled {{ background: #F0C080; color: #FFFFFF; }}
    """,
    "btn_secondary": f"""
        QPushButton {{ background: #F0F3F7; color: {TEXT_PRI}; border: 1px solid #D0D5DD;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #E4E8ED; }}
    """,
    "btn_red": f"""
        QPushButton {{ background: {DANGER}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 5px 14px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #D93030; }}
    """,
    "dlg_primary": f"""
        QPushButton {{ background: {PRIMARY}; color: #FFFFFF; border: none;
            border-radius: 4px; padding: 6px 20px; font-size: 13px; font-weight: bold; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: {PRIMARY_H}; }}
    """,
    "dlg_secondary": f"""
        QPushButton {{ background: #F0F3F7; color: {TEXT_PRI}; border: 1px solid #D0D5DD;
            border-radius: 4px; padding: 6px 20px; font-size: 13px; font-family: "Microsoft YaHei"; }}
        QPushButton:hover {{ background: #E4E8ED; }}
    """,
}


def _apply_btn_style(btn: QPushButton, style_name: str):
    """直接给按钮设置内联样式，不依赖QSS继承"""
    ss = _BTN_STYLES.get(style_name, "")
    if ss:
        btn.setStyleSheet(ss)


def _relative_path_text(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def _path_is_inside(path: Path, root: Path) -> bool:
    try:
        path_text = os.path.normcase(os.path.abspath(os.fspath(path)))
        root_text = os.path.normcase(os.path.abspath(os.fspath(root)))
        return os.path.commonpath([path_text, root_text]) == root_text
    except (OSError, ValueError):
        return False


def _find_sum_data_folders(root: Path) -> List[Path]:
    folders = set()
    for file_path in root.rglob('*'):
        if file_path.is_file() and file_path.suffix.lower() == '.sum':
            folders.add(file_path.parent)
    return sorted(folders, key=lambda p: str(p).lower())


def _sep_v(parent=None) -> QFrame:
    f = QFrame(parent)
    f.setFrameShape(QFrame.VLine)
    f.setStyleSheet(f"color: {BORDER};")
    f.setFixedWidth(1)
    f.setFixedHeight(22)
    return f


# ═══════════════════════════════════════════════════════════════
#  主窗口
# ═══════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    """涡流检测探头信息提取软件 - PyQt5 主窗口（1:1 还原旧版）"""

    def __init__(self, logger: logging.Logger | None = None):
        super().__init__()
        self._log = logger or logging.getLogger(__name__)
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(*MIN_WINDOW_SIZE)
        self._apply_default_geometry()
        self._set_app_icon()
        self._modeless_dialogs: dict[str, QDialog] = {}
        self._pending_chart_refresh = False
        self._filtered_stats_cache: dict[tuple, tuple[list, dict]] = {}
        self._filtered_warning_cache: dict[tuple, tuple[str, dict, list]] = {}
        self._chart_request_token: int = 0
        self._chart_worker: Optional[_Worker] = None

        # ── 业务组件 ──
        self.extractor              = SummaryFileExtractor()
        self.analyzer               = ProbeAnalyzer()
        self.exporter               = DataExporter()
        self.visualizer             = DataVisualizer()
        self.sum_parser             = SumFileParser()
        self.batch_lifetime_analyzer = BatchLifetimeAnalyzer()

        # ── 数据状态 ──
        self.session_records:    List[ProbeRecord]          = []
        self.session_statistics: Dict[str, ProbeStatistics] = {}
        self.current_records:    List[ProbeRecord]          = []
        self.current_statistics: Dict[str, ProbeStatistics] = {}
        self.history_records:    List[ProbeRecord]          = []
        self.history_statistics: Dict[str, ProbeStatistics] = {}
        self._history_statistics_dirty: bool                 = False
        self.history_warning_messages: List[str]            = []
        self.history_warning_by_probe: Dict                 = {}
        self.history_error_records: List[Dict]              = []
        self.history_store_path = HISTORY_STORE_PATH
        self._warning_messages:  List[str]                  = []
        self._warning_by_probe:  Dict                       = {}
        self.session_error_records: List[Dict]              = []
        self.session_deduplication_info: Dict               = {}
        self.history_deduplication_info: Dict               = {}
        self.history_import_summaries:   List[Dict]         = []

        # ── 筛选状态 ──
        self.filter_values:         Dict[str, Any] = {}
        self.summary_filter_values: Dict[str, Any] = {}
        self._global_keyword_filter: str           = ""

        # ── UI 状态 ──
        self._current_scope:  str  = "current"
        self._current_view:   str  = "split"
        self._history_enabled: bool = True
        self._current_chart_figure  = None
        self._current_chart_canvas: Optional[FigureCanvas] = None
        self._current_chart_type:   str = ""
        self._is_refreshing:        bool = False
        self._safety_shown:         bool = False
        self._current_worker: Optional[_Worker] = None
        self._resize_timer: Optional[QTimer] = None
        self._chart_refresh_timer: Optional[QTimer] = None

        # ── 线程池 ──
        self._pool = QThreadPool.globalInstance()
        self._pool.setMaxThreadCount(max(2, min(4, self._pool.maxThreadCount())))

        # ── 构建 UI ──
        self._build_ui()
        self._build_menu()
        self._build_status_bar()

        # ── 加载历史 ──
        self._load_history_store()
        self._apply_active_dataset(refresh_ui=True)
        self._refresh_warning_button_state()

        # 安全提示保留在菜单中手动查看，不在启动时弹窗打断主界面。

    # ═══════════════════════════════════════════════════════════
    #  初始化辅助
    # ═══════════════════════════════════════════════════════════
    def _apply_default_geometry(self):
        screen = QApplication.primaryScreen().availableGeometry()
        w = min(DEFAULT_WINDOW_SIZE[0], max(MIN_WINDOW_SIZE[0], screen.width() - 90))
        h = min(DEFAULT_WINDOW_SIZE[1], max(MIN_WINDOW_SIZE[1], screen.height() - 120))
        x = max(20, (screen.width() - w) // 2)
        y = max(20, (screen.height() - h) // 2 - 18)
        self.setGeometry(x, y, w, h)

    def _set_app_icon(self):
        if LOGO_PATH.exists():
            self.setWindowIcon(QIcon(str(LOGO_PATH)))

    def _configure_dialog_window(self, dlg: QDialog, *, min_size: tuple[int, int] | None = None):
        flags = dlg.windowFlags()
        flags &= ~Qt.WindowContextHelpButtonHint
        flags &= ~Qt.MSWindowsFixedSizeDialogHint
        dlg.setWindowFlags(flags | Qt.Window)
        dlg.setModal(False)
        if min_size:
            dlg.setMinimumSize(*min_size)
        if LOGO_PATH.exists():
            dlg.setWindowIcon(QIcon(str(LOGO_PATH)))

    def _present_dialog(self, dlg: QDialog, key: str | None = None):
        dialog_key = key or f"dialog:{id(dlg)}"
        existing = self._modeless_dialogs.get(dialog_key)
        if existing is not None and existing is not dlg:
            try:
                if existing.isVisible():
                    existing.close()
            except RuntimeError:
                pass
            self._modeless_dialogs.pop(dialog_key, None)

        self._modeless_dialogs[dialog_key] = dlg
        dlg.setAttribute(Qt.WA_DeleteOnClose, True)

        def _cleanup(*_):
            if self._modeless_dialogs.get(dialog_key) is dlg:
                self._modeless_dialogs.pop(dialog_key, None)

        dlg.destroyed.connect(_cleanup)
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()
        return dlg

    def _get_protected_source_roots(self) -> list[Path]:
        roots: list[Path] = []
        for candidate in (
            self._folder_edit.text().strip(),
            self._file_edit.text().strip(),
        ):
            if not candidate:
                continue
            try:
                path = Path(candidate)
                if path.exists():
                    roots.append(path if path.is_dir() else path.parent)
            except OSError:
                continue
        return roots

    def _ensure_output_path_allowed(self, output_path: Path, *, title: str = "禁止写入源目录") -> bool:
        for root in self._get_protected_source_roots():
            if _path_is_inside(output_path, root):
                QMessageBox.warning(
                    self,
                    title,
                    "为保护原始数据，输出文件不能保存到当前导入的数据源目录或其子目录。"
                    "请保存到本地输出目录。",
                )
                self._set_status("已取消保存 - 禁止写入源目录")
                return False
        return True

    def _build_summary_export_rows(self, records: list, statistics: dict | None = None) -> list[dict]:
        statistics = statistics or {}
        if not statistics:
            if self._get_active_filter_items():
                _, statistics = self._get_filtered_records_and_stats()
            else:
                statistics = self.current_statistics
        rows: list[dict] = []
        for idx, record in enumerate(records, 1):
            stat = statistics.get(record.stat_key)
            if not stat:
                continue
            start_time = getattr(record, 'start_time', None)
            end_time = getattr(record, 'end_time', None)
            start_str = start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else ''
            end_str = end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else ''
            if start_time and end_time:
                diff_s = (end_time - start_time).total_seconds()
                sh, sm = diff_s / 3600.0, diff_s / 60.0
            else:
                sh = sm = 0.0
            total_uses = int(getattr(stat, 'total_uses', 0) or 0)
            total_duration_minutes = float(getattr(stat, 'total_duration_minutes', 0.0) or 0.0)
            longest = float(getattr(stat, 'longest_continuous_duration_minutes', 0.0) or 0.0)
            detection_speed = float(getattr(stat, 'detection_speed', 0.0) or 0.0)
            unique_tube_count = int(getattr(stat, 'unique_tube_count', 0) or 0)
            first_use_time = getattr(stat, 'first_use_time', None)
            last_use_time = getattr(stat, 'last_use_time', None)
            rows.append({
                '序号': idx,
                '大修': getattr(record, 'outage', '') or '',
                '蒸汽发生器编号': getattr(record, 'sg_id', '') or '',
                '数据组': getattr(record, 'data_group', '') or '',
                '操作员': getattr(record, 'operator', '') or '',
                '探头类型': getattr(record, 'probe_type_raw', None) or getattr(stat, 'probe_type', ''),
                '探头编码': getattr(record, 'probe_sn', '') or '',
                '探头型号': getattr(record, 'model', '') or '',
                '管道数量': getattr(record, 'tube_number', '') if getattr(record, 'tube_number', None) is not None else '',
                '累计管道数量': unique_tube_count,
                '开始时间': start_str,
                '结束时间': end_str,
                '单次使用时间(小时)': f"{sh:.2f}",
                '单次使用时间(分钟)': f"{sm:.2f}",
                '总使用次数': total_uses,
                '总使用时间(小时)': f"{total_duration_minutes / 60.0:.2f}",
                '最长连续使用(小时)': f"{longest / 60.0:.2f}",
                '检测速度(管道/小时)': f"{detection_speed:.2f}",
                '首次使用日期': first_use_time.strftime('%Y-%m-%d') if first_use_time else '',
                '末次使用日期': last_use_time.strftime('%Y-%m-%d') if last_use_time else '',
            })
        return rows

    def _invalidate_filtered_cache(self):
        self._filtered_stats_cache.clear()
        self._filtered_warning_cache.clear()

    def _make_filter_cache_key(self) -> tuple:
        filter_items = tuple(sorted(
            (col, self.summary_filter_values.get(col))
            for col in TABLE_HEADER_FILTERABLE_COLUMNS
            if self._filter_value_is_active(self.summary_filter_values.get(col))
        ))
        return (
            self._current_scope,
            id(self.current_records),
            id(self.current_statistics),
            len(self.current_records),
            filter_items,
            self._global_keyword_filter.strip().lower(),
        )

    def _get_filtered_records_and_stats(self) -> tuple[list, dict]:
        key = self._make_filter_cache_key()
        cached = self._filtered_stats_cache.get(key)
        if cached is not None:
            return cached
        filtered = [r for r in self.current_records if self._summary_record_matches_filters(r)]
        stats = self._rebuild_statistics_from_records(filtered) if filtered else {}
        cached_result = (filtered, stats)
        self._filtered_stats_cache[key] = cached_result
        return cached_result

    # ═══════════════════════════════════════════════════════════
    #  菜单栏
    # ═══════════════════════════════════════════════════════════
    def _build_menu(self):
        mb = self.menuBar()
        fm = mb.addMenu("文件")
        fm.addAction("选择文件夹...", self._select_folder)
        fm.addAction("选择 Excel...", self._select_file)
        fm.addSeparator()
        fm.addAction("处理文件夹", self._process_sum_files)
        fm.addAction("分析Excel", self._start_processing)
        fm.addSeparator()
        fm.addAction("一键保存", self._save_all)
        fm.addAction("退出", self.close)

        vm = mb.addMenu("视图")
        vm.addAction("综合工作台", lambda: self._switch_view("split"))
        vm.addAction("信息表格", lambda: self._switch_view("table"))
        vm.addAction("数据可视化", lambda: self._switch_view("chart"))

        hm = mb.addMenu("历史")
        hm.addAction("开启/关闭累计历史", self._toggle_history_enabled)
        hm.addAction("切换到当前批次", lambda: self._set_data_scope("current"))
        hm.addAction("切换到历史累计", lambda: self._set_data_scope("history"))
        hm.addAction("历史详情", self._show_history_details)
        hm.addAction("清空数据", self._clear_history_records)

        em = mb.addMenu("导出与提醒")
        em.addAction("保存表格", self._export_summary_table)
        em.addAction("导出错误记录", self._export_error_records)
        em.addSeparator()
        em.addAction("查看警告", self._show_warning_details)
        em.addAction("去重信息", self._show_deduplication_info)

        hlm = mb.addMenu("帮助")
        hlm.addAction("软件说明", self._show_software_notes)
        hlm.addAction("数据安全提示", lambda: self._show_data_safety_warning(force=True))

    # ═══════════════════════════════════════════════════════════
    #  状态栏
    # ═══════════════════════════════════════════════════════════
    def _build_status_bar(self):
        sb = QStatusBar()
        self.setStatusBar(sb)
        self._status_lbl = QLabel("就绪")
        self._status_lbl.setFont(QFont("Microsoft YaHei", 9))
        sb.addWidget(self._status_lbl, 1)
        self._status_progress = QProgressBar()
        self._status_progress.setFixedWidth(180)
        self._status_progress.setFixedHeight(9)
        self._status_progress.setRange(0, 100)
        self._status_progress.setValue(0)
        self._status_progress.setTextVisible(False)
        self._status_progress.hide()
        sb.addPermanentWidget(self._status_progress)

    def _set_status(self, text: str, progress: float = -1):
        self._status_lbl.setText(text)
        if progress >= 0:
            self._status_progress.show()
            self._status_progress.setValue(int(progress))
        else:
            self._status_progress.hide()
        # 同时更新详细状态
        self._update_scope_status_summary()

    def _update_scope_status_summary(self):
        """更新状态栏的详细信息（与旧版完全一致）"""
        scope = self._current_scope_label()
        total_r = len(self.current_records)
        total_p = len(self.current_statistics)
        total_h = sum(s.total_duration_minutes for s in self.current_statistics.values()) / 60.0
        hist_count = len(self.history_records)

        active = self._get_active_filter_items()
        if active:
            self._status_lbl.setText(
                f"{scope}模式 | 筛选后 {total_r} 条记录，{total_p} 个探头，{total_h:.1f} 小时"
            )
        elif scope == "当前批次":
            self._status_lbl.setText(
                f"当前批次模式 | 本批次 {total_r} 条记录，{total_p} 个探头 | 历史仓 {hist_count} 条"
            )
        else:
            dedup = self.history_deduplication_info or {}
            orig = dedup.get('original_count', total_r)
            removed = max(0, orig - total_r)
            msg = f"历史累计模式 | 当前显示 {total_r} 条有效记录，{total_p} 个探头"
            if removed:
                msg += f" | 去重前 {orig} 条，累计去重 {removed} 条"
            self._status_lbl.setText(msg)

    def _schedule_status_reset(self, delay_ms: int = 3000):
        QTimer.singleShot(delay_ms, lambda: self._set_status("就绪"))

    # ═══════════════════════════════════════════════════════════
    #  主布局
    # ═══════════════════════════════════════════════════════════
    def _build_ui(self):
        central = QWidget()
        central.setObjectName("central")
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._build_sidebar())
        root.addWidget(self._build_workspace(), 1)

    # ═══════════════════════════════════════════════════════════
    #  侧边栏（完全还原截图）
    # ═══════════════════════════════════════════════════════════
    def _build_sidebar(self) -> QWidget:
        sb = QWidget()
        sb.setObjectName("sidebar")
        sb.setFixedWidth(150)
        sb.setStyleSheet(f"background: {SIDEBAR};")
        lay = QVBoxLayout(sb)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── 品牌区 ──
        brand = QWidget()
        brand.setStyleSheet(f"background: {SIDEBAR_H};")
        bl = QVBoxLayout(brand)
        bl.setContentsMargins(10, 12, 10, 12)
        bl.setSpacing(4)
        t = QLabel("PIES 工作台")
        t.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        t.setStyleSheet("color: #FFFFFF; background: transparent;")
        bl.addWidget(t)
        s = QLabel("探头信息提取、\n分析、可视化与\n导出")
        s.setFont(QFont("Microsoft YaHei", 8))
        s.setStyleSheet("color: #B0C4D8; background: transparent; line-height: 150%;")
        bl.addWidget(s)
        lay.addWidget(brand)

        # ── 导航区 ──
        nav = QWidget()
        nav.setStyleSheet(f"background: {SIDEBAR};")
        nl = QVBoxLayout(nav)
        nl.setContentsMargins(0, 8, 0, 0)
        nl.setSpacing(0)

        # 工作视图
        nl.addWidget(self._sb_group_label("工作视图"))
        self._sb_view_btns: Dict[str, QPushButton] = {}
        for key, label in [("split", "综合工作台"), ("table", "表格专注视图"), ("chart", "图形专注视图")]:
            b = self._make_sb_btn(label, lambda checked=False, k=key: self._switch_view(k))
            self._sb_view_btns[key] = b
            nl.addWidget(b)

        # 数据范围
        nl.addWidget(self._sb_group_label("数据范围"))
        self._sb_scope_btns: Dict[str, QPushButton] = {}
        for key, label in [("current", "当前批次"), ("history", "历史累计")]:
            b = self._make_sb_btn(label, lambda checked=False, k=key: self._set_data_scope(k))
            self._sb_scope_btns[key] = b
            nl.addWidget(b)

        # 历史记录
        nl.addWidget(self._sb_group_label("历史记录"))
        self._history_toggle_btn = self._make_sb_btn("历史记录：开", self._toggle_history_enabled)
        self._history_toggle_btn.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
        nl.addWidget(self._history_toggle_btn)

        # 快捷操作
        nl.addWidget(self._sb_group_label("快捷操作"))
        nl.addWidget(self._make_sb_btn("查看历史详情", self._show_history_details))
        nl.addWidget(self._make_sb_btn("清空数据", self._clear_history_records))

        nl.addStretch(1)
        lay.addWidget(nav, 1)

        # ── 运行状态 ──
        status_area = QWidget()
        status_area.setStyleSheet(f"background: {SIDEBAR_H};")
        sl = QVBoxLayout(status_area)
        sl.setContentsMargins(10, 8, 10, 10)
        sl.setSpacing(3)
        run_lbl = QLabel("运行状态")
        run_lbl.setFont(QFont("Microsoft YaHei", 8, QFont.Bold))
        run_lbl.setStyleSheet("color: #FFFFFF; background: transparent;")
        sl.addWidget(run_lbl)
        self._scope_badge_lbl = QLabel("当前批次")
        self._scope_badge_lbl.setFont(QFont("Microsoft YaHei", 8))
        self._scope_badge_lbl.setStyleSheet("color: #AFC3D9; background: transparent;")
        sl.addWidget(self._scope_badge_lbl)
        lay.addWidget(status_area)
        return sb

    def _sb_group_label(self, text: str) -> QLabel:
        """侧边栏分组标签"""
        lbl = QLabel(text)
        lbl.setFont(QFont("Microsoft YaHei", 8, QFont.Bold))
        lbl.setStyleSheet(f"""
            color: #7A9BBF; background: transparent;
            padding: 10px 10px 3px 10px;
            font-size: 11px;
        """)
        return lbl

    def _make_sb_btn(self, text: str, slot) -> QPushButton:
        """侧边栏导航按钮"""
        b = QPushButton(text)
        b.setFont(QFont("Microsoft YaHei", 9))
        b.setCursor(QCursor(Qt.PointingHandCursor))
        b.setFixedHeight(32)
        b.setMinimumWidth(138)
        b.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: #C8D8E8;
                border: none; border-radius: 0px;
                padding: 0 12px; text-align: left;
                font-size: 12px;
            }}
            QPushButton:hover {{ background: #2E3F50; color: #FFFFFF; }}
        """)
        b.clicked.connect(slot)
        return b

    def _update_sidebar_nav(self):
        """更新侧边栏导航按钮的选中状态"""
        for key, btn in self._sb_view_btns.items():
            sel = key == self._current_view
            if sel:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {PRIMARY}; color: #FFFFFF;
                        border: none; border-radius: 0px;
                        padding: 0 10px; text-align: left; font-size: 12px;
                    }}
                    QPushButton:hover {{ background: {PRIMARY_H}; }}
                """)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: transparent; color: #C8D8E8;
                        border: none; border-radius: 0px;
                        padding: 0 10px; text-align: left; font-size: 12px;
                    }}
                    QPushButton:hover {{ background: #2E3F50; color: #FFFFFF; }}
                """)

        for key, btn in self._sb_scope_btns.items():
            sel = key == self._current_scope
            if sel:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {PRIMARY}; color: #FFFFFF;
                        border: none; border-radius: 0px;
                        padding: 0 10px; text-align: left; font-size: 12px;
                    }}
                    QPushButton:hover {{ background: {PRIMARY_H}; }}
                """)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: transparent; color: #C8D8E8;
                        border: none; border-radius: 0px;
                        padding: 0 10px; text-align: left; font-size: 12px;
                    }}
                    QPushButton:hover {{ background: #2E3F50; color: #FFFFFF; }}
                """)

        # 历史记录开关按钮颜色
        if self._history_enabled:
            self._history_toggle_btn.setStyleSheet(f"""
                QPushButton {{
                    background: {ACCENT}; color: #FFFFFF;
                    border: none; border-radius: 0px;
                    padding: 0 10px; text-align: left;
                    font-size: 12px; font-weight: bold;
                }}
                QPushButton:hover {{ background: {ACCENT_H}; }}
            """)
            self._history_toggle_btn.setText("历史记录：开")
        else:
            self._history_toggle_btn.setStyleSheet(f"""
                QPushButton {{
                    background: {DANGER}; color: #FFFFFF;
                    border: none; border-radius: 0px;
                    padding: 0 10px; text-align: left;
                    font-size: 12px; font-weight: bold;
                }}
                QPushButton:hover {{ background: #D93030; }}
            """)
            self._history_toggle_btn.setText("历史记录：关")

        scope_text = "历史累计" if self._current_scope == "history" else "当前批次"
        self._scope_badge_lbl.setText(scope_text)

    # ═══════════════════════════════════════════════════════════
    #  工作区（完全还原截图布局）
    # ═══════════════════════════════════════════════════════════
    def _build_workspace(self) -> QWidget:
        ws = QWidget()
        ws.setStyleSheet(f"background: {BG};")
        lay = QVBoxLayout(ws)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # 顶部工具栏（白色背景，带底部边框）
        lay.addWidget(self._build_hero_card())

        # 概览指标行
        lay.addWidget(self._build_overview_row())

        # 内容区（表格+图表分割）
        content = QWidget()
        content.setStyleSheet(f"background: {BG};")
        cl = QVBoxLayout(content)
        cl.setContentsMargins(8, 6, 8, 8)
        cl.setSpacing(0)

        self._splitter = QSplitter(Qt.Horizontal)
        self._splitter.setHandleWidth(5)
        self._table_panel = self._build_table_panel()
        self._chart_panel = self._build_chart_panel()
        self._splitter.addWidget(self._table_panel)
        self._splitter.addWidget(self._chart_panel)
        self._splitter.setSizes([680, 720])
        cl.addWidget(self._splitter, 1)
        lay.addWidget(content, 1)
        return ws

    # ─── 顶部工具栏（Hero Card）完全还原截图 ───
    def _build_hero_card(self) -> QWidget:
        """顶部工具栏：白色背景，包含按钮行和路径行"""
        card = QWidget()
        card.setObjectName("hero_card")
        card.setStyleSheet(f"background: {PANEL}; border-bottom: 1px solid {BORDER};")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 6, 10, 6)
        lay.setSpacing(5)

        # ── 按钮行 ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)

        # 文件夹按钮（灰色）
        self._btn_folder = QPushButton("文件夹")
        self._btn_folder.setObjectName("btn_file")
        self._btn_folder.setFixedHeight(30)
        self._btn_folder.setMinimumWidth(78)
        self._btn_folder.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_folder.clicked.connect(self._select_folder)
        _apply_btn_style(self._btn_folder, "btn_file")
        btn_row.addWidget(self._btn_folder)

        # Excel按钮（灰色）
        self._btn_excel = QPushButton("Excel")
        self._btn_excel.setObjectName("btn_file")
        self._btn_excel.setFixedHeight(30)
        self._btn_excel.setMinimumWidth(98)
        self._btn_excel.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_excel.clicked.connect(self._select_file)
        _apply_btn_style(self._btn_excel, "btn_file")
        btn_row.addWidget(self._btn_excel)

        # 处理文件夹按钮（蓝色）
        self._btn_sum = QPushButton("处理文件夹")
        self._btn_sum.setObjectName("btn_blue")
        self._btn_sum.setFixedHeight(30)
        self._btn_sum.setMinimumWidth(112)
        self._btn_sum.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_sum.clicked.connect(self._process_sum_files)
        _apply_btn_style(self._btn_sum, "btn_blue")
        btn_row.addWidget(self._btn_sum)

        # 分析Excel按钮（蓝色）
        self._btn_process = QPushButton("分析Excel")
        self._btn_process.setObjectName("btn_primary")
        self._btn_process.setFixedHeight(30)
        self._btn_process.setMinimumWidth(122)
        self._btn_process.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_process.clicked.connect(self._start_processing)
        _apply_btn_style(self._btn_process, "btn_primary")
        btn_row.addWidget(self._btn_process)

        # 弹性空间
        btn_row.addStretch(1)

        # 异常提醒按钮（橙色，右侧）
        self._btn_warnings = QPushButton("异常提醒")
        self._btn_warnings.setObjectName("btn_warning")
        self._btn_warnings.setFixedHeight(30)
        self._btn_warnings.setMinimumWidth(104)
        self._btn_warnings.setMaximumWidth(128)
        self._btn_warnings.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_warnings.setEnabled(False)
        self._btn_warnings.clicked.connect(self._show_warning_details)
        _apply_btn_style(self._btn_warnings, "btn_warning")
        btn_row.addWidget(self._btn_warnings)

        # 一键保存按钮（绿色，右侧）
        self._btn_save_all = QPushButton("一键保存")
        self._btn_save_all.setObjectName("btn_accent")
        self._btn_save_all.setFixedHeight(30)
        self._btn_save_all.setMinimumWidth(92)
        self._btn_save_all.setCursor(QCursor(Qt.PointingHandCursor))
        self._btn_save_all.clicked.connect(self._save_all)
        _apply_btn_style(self._btn_save_all, "btn_accent")
        btn_row.addWidget(self._btn_save_all)

        lay.addLayout(btn_row)

        # ── 路径行 ──
        path_row = QHBoxLayout()
        path_row.setSpacing(6)

        lbl_dir = QLabel("目录")
        lbl_dir.setStyleSheet("color: #667085; font-size: 12px; font-weight: bold; font-family: 'Microsoft YaHei';")
        lbl_dir.setFixedWidth(28)
        path_row.addWidget(lbl_dir)

        self._folder_edit = QLineEdit()
        self._folder_edit.setReadOnly(True)
        self._folder_edit.setFixedHeight(24)
        self._folder_edit.setPlaceholderText("未选择 SUM 目录")
        self._folder_edit.setFont(QFont("Consolas", 9))
        path_row.addWidget(self._folder_edit, 1)

        lbl_excel = QLabel("Excel")
        lbl_excel.setStyleSheet("color: #667085; font-size: 12px; font-weight: bold; font-family: 'Microsoft YaHei';")
        lbl_excel.setFixedWidth(36)
        path_row.addWidget(lbl_excel)

        self._file_edit = QLineEdit()
        self._file_edit.setReadOnly(True)
        self._file_edit.setFixedHeight(24)
        self._file_edit.setPlaceholderText("未选择 Excel 文件")
        self._file_edit.setFont(QFont("Consolas", 9))
        path_row.addWidget(self._file_edit, 1)

        lay.addLayout(path_row)
        return card

    # ─── 概览指标行（完全还原截图中的4个大卡片）───
    def _build_overview_row(self) -> QWidget:
        """4个概览指标卡片，横向排列，高度固定"""
        row = QWidget()
        row.setStyleSheet(f"background: {BG};")
        row.setFixedHeight(68)
        lay = QHBoxLayout(row)
        lay.setContentsMargins(8, 4, 8, 4)
        lay.setSpacing(6)

        self._metric_cards: Dict[str, Dict] = {}
        specs = [
            ("records", "当前记录",  "#1677FF", "#EBF3FF", "#1677FF"),
            ("probes",  "探头数量",  "#1677FF", "#EBF3FF", "#1677FF"),
            ("hours",   "累计时长",  "#00B42A", "#E8F8EE", "#00B42A"),
            ("warnings","异常提醒",  "#FF7D00", "#FFF3E8", "#FF7D00"),
        ]
        for key, title, accent, bg, fg in specs:
            lay.addWidget(self._build_metric_card(key, title, accent, bg, fg), 1)
        return row

    def _build_metric_card(self, key: str, title: str, accent: str, bg: str, fg: str = "#1F2937") -> QFrame:
        """单个概览指标卡片"""
        card = QFrame()
        card.setObjectName("overview_card")
        card.setStyleSheet(f"""
            QFrame#overview_card {{
                background: {bg};
                border: 1px solid {accent}40;
                border-left: 3px solid {accent};
                border-radius: 4px;
            }}
        """)
        lay = QHBoxLayout(card)
        lay.setContentsMargins(12, 6, 12, 6)
        lay.setSpacing(8)

        # 左侧：标题+备注
        left = QVBoxLayout()
        left.setSpacing(2)
        title_lbl = QLabel(title)
        title_lbl.setFont(QFont("Microsoft YaHei", 9))
        title_lbl.setStyleSheet(f"color: #667085; background: transparent;")
        left.addWidget(title_lbl)
        note_lbl = QLabel("")
        note_lbl.setFont(QFont("Microsoft YaHei", 8))
        note_lbl.setStyleSheet("color: #98A2B3; background: transparent;")
        left.addWidget(note_lbl)
        lay.addLayout(left, 1)

        # 右侧：大数字
        value_lbl = QLabel("0")
        value_lbl.setFont(QFont("Microsoft YaHei", 18, QFont.Bold))
        value_lbl.setStyleSheet(f"color: {fg}; background: transparent;")
        value_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lay.addWidget(value_lbl)

        self._metric_cards[key] = {
            "card": card, "title": title_lbl, "value": value_lbl, "note": note_lbl,
            "default_bg": bg, "default_accent": accent, "default_fg": fg,
        }
        return card

    def _refresh_overview_metrics(self):
        recs, stats, filtered = self._get_metric_source_data()
        total_r = len(recs)
        total_p = len(stats)
        total_h = sum(float(getattr(s, 'total_duration_minutes', 0.0) or 0.0) for s in stats.values()) / 60.0
        _, warning_groups, warning_messages = self._get_filtered_warning_snapshot()
        warn_n = self._count_warning_items(warning_groups, warning_messages)
        reuse_n = sum(1 for s in stats.values() if int(getattr(s, 'continuous_use_count', 0) or 0) > 1)
        scope   = self._current_scope_label()

        # 当前记录卡片
        self._metric_cards["records"]["value"].setText(str(total_r))
        self._metric_cards["records"]["note"].setText(
            f"{scope}下共 {total_r} 条记录" if total_r else f"{scope}下暂无记录"
        )

        # 探头数量卡片
        self._metric_cards["probes"]["value"].setText(str(total_p))
        self._metric_cards["probes"]["note"].setText(
            f"统计对象共 {total_p} 个探头" if total_p else "统计对象共 0 个探头"
        )

        # 累计时长卡片
        self._metric_cards["hours"]["value"].setText(f"{total_h:.1f} h")
        if total_p:
            self._metric_cards["hours"]["note"].setText(f"再次使用探头 {reuse_n} 个")
        else:
            self._metric_cards["hours"]["note"].setText("等待分析结果生成")

        # 异常提醒卡片
        if warn_n:
            self._metric_cards["warnings"]["value"].setText(f"{warn_n} 项")
            self._metric_cards["warnings"]["note"].setText(f"{scope}需关注的数据异常")
            self._metric_cards["warnings"]["card"].setStyleSheet("""
                QFrame#overview_card {
                    background: #FFF3E8;
                    border: 1px solid #FF7D0040;
                    border-left: 3px solid #FF7D00;
                    border-radius: 4px;
                }
            """)
            self._metric_cards["warnings"]["value"].setStyleSheet("color: #FF7D00; background: transparent; font-size: 18px; font-weight: bold;")
            self._metric_cards["warnings"]["title"].setStyleSheet("color: #B45309; background: transparent;")
        else:
            self._metric_cards["warnings"]["value"].setText("无异常")
            self._metric_cards["warnings"]["note"].setText(f"{scope}未发现异常")
            self._metric_cards["warnings"]["card"].setStyleSheet("""
                QFrame#overview_card {
                    background: #E8F8EE;
                    border: 1px solid #00B42A40;
                    border-left: 3px solid #00B42A;
                    border-radius: 4px;
                }
            """)
            self._metric_cards["warnings"]["value"].setStyleSheet("color: #00B42A; background: transparent; font-size: 18px; font-weight: bold;")
            self._metric_cards["warnings"]["title"].setStyleSheet("color: #067647; background: transparent;")

    def _get_metric_source_data(self):
        active = self._get_active_filter_items()
        if not active:
            return self.current_records, self.current_statistics, False
        return (*self._get_filtered_records_and_stats(), True)

    def _update_filter_summary_label(self):
        label = getattr(self, '_filter_summary_label', None)
        if label is None:
            return
        items = self._get_active_filter_items()
        if not items:
            label.setText("当前筛选：全部数据")
            return
        summary = "；".join(items[:3])
        if len(items) > 3:
            summary += f"；其余{len(items) - 3}项"
        label.setText(f"当前筛选：{summary}")

    # ─── 表格面板（完全还原截图）───
    def _build_table_panel(self) -> QFrame:
        panel = QFrame()
        panel.setObjectName("table_panel")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # 标题栏（灰色背景，带底部边框）
        hdr = QWidget()
        hdr.setStyleSheet(f"background: #F7F8FA; border-bottom: 1px solid {BORDER};")
        hdr.setFixedHeight(36)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        hl.setSpacing(8)

        title_lbl = QLabel("探头使用信息表格")
        title_lbl.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        title_lbl.setStyleSheet(f"color: {TEXT_PRI};")
        hl.addWidget(title_lbl, 1)

        self._filter_summary_label = QLabel("当前筛选：全部数据")
        self._filter_summary_label.setFont(QFont("Microsoft YaHei", 9))
        self._filter_summary_label.setStyleSheet(f"color: {TEXT_SEC};")
        self._filter_summary_label.setMinimumWidth(240)
        self._filter_summary_label.setAlignment(Qt.AlignVCenter | Qt.AlignRight)
        hl.addWidget(self._filter_summary_label)

        bc = QPushButton("清除筛选"); bc.setObjectName("btn_warning")
        bc.setFixedHeight(26); bc.setMinimumWidth(92)
        bc.setCursor(QCursor(Qt.PointingHandCursor)); bc.clicked.connect(self._clear_all_filters)
        _apply_btn_style(bc, "btn_warning")
        hl.addWidget(bc)

        bs = QPushButton("保存表格"); bs.setObjectName("btn_blue")
        bs.setFixedHeight(26); bs.setMinimumWidth(90)
        bs.setCursor(QCursor(Qt.PointingHandCursor)); bs.clicked.connect(self._export_summary_table)
        _apply_btn_style(bs, "btn_blue")
        hl.addWidget(bs)

        lay.addWidget(hdr)

        # 表格
        self._summary_table = QTableWidget()
        self._summary_table.setAlternatingRowColors(True)
        self._summary_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._summary_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._summary_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._summary_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._summary_table.setColumnCount(len(SUMMARY_TABLE_HEADERS))
        self._summary_table.setHorizontalHeaderLabels(self._summary_header_labels())
        self._summary_table.verticalHeader().setVisible(False)
        self._summary_table.verticalHeader().setDefaultSectionSize(28)
        self._summary_table.setShowGrid(True)
        self._summary_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._summary_table.customContextMenuRequested.connect(self._on_table_context_menu)
        self._summary_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        header = self._summary_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(76)
        header.setDefaultSectionSize(126)
        header.setSectionsMovable(False)
        header.customContextMenuRequested.connect(self._on_header_context_menu)
        header.sectionClicked.connect(self._on_header_clicked)
        self._summary_table.doubleClicked.connect(self._on_table_double_click)
        self._apply_summary_column_widths()
        lay.addWidget(self._summary_table, 1)
        return panel

    def _summary_header_labels(self) -> list:
        return [
            f"{name} ▼" if name in TABLE_HEADER_FILTERABLE_COLUMNS else name
            for name in SUMMARY_TABLE_HEADERS
        ]

    def _apply_summary_column_widths(self):
        widths = {
            "序号": 52, "大修": 74, "蒸汽发生器编号": 132, "数据组": 178,
            "操作员": 88, "探头类型": 112, "探头编码": 132, "探头型号": 160,
            "管道数量": 82, "累计管道数量": 112, "开始时间": 158, "结束时间": 158,
            "单次使用时间(小时)": 130, "单次使用时间(分钟)": 130,
            "总使用次数": 90, "总使用时间(小时)": 120, "最长连续使用(小时)": 130,
            "检测速度(管道/小时)": 130, "首次使用日期": 110, "末次使用日期": 110,
        }
        for i, col in enumerate(SUMMARY_TABLE_HEADERS):
            width = max(90, widths.get(col, 118))
            self._summary_table.setColumnWidth(i, width)
            item = self._summary_table.horizontalHeaderItem(i)
            if item:
                item.setToolTip(col)

    # ─── 图表面板（完全还原截图）───
    def _build_chart_panel(self) -> QFrame:
        panel = QFrame()
        panel.setObjectName("chart_panel")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # 标题栏（灰色背景，带底部边框）
        hdr = QWidget()
        hdr.setStyleSheet(f"background: #F7F8FA; border-bottom: 1px solid {BORDER};")
        hdr.setFixedHeight(36)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        hl.setSpacing(6)

        title_lbl = QLabel("数据可视化")
        title_lbl.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        title_lbl.setStyleSheet(f"color: {TEXT_PRI};")
        hl.addWidget(title_lbl, 1)

        # 图表类型下拉
        self._chart_type_combo = QComboBox()
        self._chart_type_combo.setFont(QFont("Microsoft YaHei", 9))
        self._chart_type_combo.setFixedHeight(26)
        self._chart_type_combo.setMinimumWidth(130)
        self._chart_type_combo.setMaxVisibleItems(8)
        for g in CHART_GROUPS:
            self._chart_type_combo.addItem(g)
        self._chart_type_combo.activated[str].connect(self._on_chart_type_changed)
        hl.addWidget(self._chart_type_combo)

        # 图表选项下拉
        self._chart_option_combo = QComboBox()
        self._chart_option_combo.setFont(QFont("Microsoft YaHei", 9))
        self._chart_option_combo.setFixedHeight(26)
        self._chart_option_combo.setMinimumWidth(150)
        self._chart_option_combo.setMaxVisibleItems(8)
        self._chart_option_combo.activated[str].connect(lambda _: self._schedule_chart_refresh())
        hl.addWidget(self._chart_option_combo)

        # 刷新按钮（蓝色）
        btn_refresh = QPushButton("刷新")
        btn_refresh.setObjectName("btn_blue")
        btn_refresh.setFixedHeight(26); btn_refresh.setMinimumWidth(60)
        btn_refresh.setCursor(QCursor(Qt.PointingHandCursor))
        btn_refresh.clicked.connect(self._refresh_chart)
        _apply_btn_style(btn_refresh, "btn_blue")
        hl.addWidget(btn_refresh)

        # 导出图表按钮（绿色）
        btn_export = QPushButton("导出图表")
        btn_export.setObjectName("btn_accent")
        btn_export.setFixedHeight(26); btn_export.setMinimumWidth(90)
        btn_export.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export.clicked.connect(self._export_chart)
        _apply_btn_style(btn_export, "btn_accent")
        hl.addWidget(btn_export)

        lay.addWidget(hdr)

        # 图表滚动区
        self._chart_scroll = QScrollArea()
        self._chart_scroll.setWidgetResizable(True)
        self._chart_scroll.setStyleSheet("QScrollArea { border: none; background: #FFFFFF; }")
        self._chart_container = QWidget()
        self._chart_container.setStyleSheet("background: #FFFFFF;")
        self._chart_container_layout = QVBoxLayout(self._chart_container)
        self._chart_container_layout.setContentsMargins(0, 0, 0, 0)
        self._chart_container_layout.setAlignment(Qt.AlignTop)
        self._chart_scroll.setWidget(self._chart_container)
        lay.addWidget(self._chart_scroll, 1)

        self._update_chart_option_combo()
        return panel

    # ═══════════════════════════════════════════════════════════
    #  视图切换
    # ═══════════════════════════════════════════════════════════
    def _switch_view(self, view: str):
        self._current_view = view
        total = max(self._splitter.width(), 1000)
        self._splitter.setUpdatesEnabled(False)
        if view == "table":
            self._table_panel.show(); self._chart_panel.hide()
            self._splitter.setSizes([total, 0])
        elif view == "chart":
            self._table_panel.hide(); self._chart_panel.show()
            self._splitter.setSizes([0, total])
        else:
            self._table_panel.show(); self._chart_panel.show()
            self._splitter.setSizes([total // 2, total // 2])
        self._splitter.setUpdatesEnabled(True)
        self._update_sidebar_nav()
        self._table_panel.update()
        self._chart_panel.update()
        if view == "chart" and self.current_records and self.current_statistics:
            self._show_chart_placeholder("正在生成默认图表...")
            self._pending_chart_refresh = True
            self._schedule_chart_render_ready_checks()

    # ═══════════════════════════════════════════════════════════
    #  数据范围 & 历史
    # ═══════════════════════════════════════════════════════════
    def _set_data_scope(self, scope: str):
        if scope == "history" and not self.history_records:
            QMessageBox.information(self, "提示", "当前还没有累计的历史数据。")
            return
        self._current_scope = scope
        self._apply_active_dataset(refresh_ui=True)

    def _toggle_history_enabled(self):
        self._history_enabled = not self._history_enabled
        self._update_sidebar_nav()
        self._set_status(f"历史记录已{'开启' if self._history_enabled else '关闭'}")

    def _apply_active_dataset(self, refresh_ui: bool = False):
        if self._current_scope == "history" and self.history_records:
            self._ensure_history_statistics()
            self.current_records = self.history_records
            self.current_statistics = self.history_statistics
        else:
            self._current_scope = "current"
            self.current_records = self.session_records
            if self.session_statistics:
                self.current_statistics = self.session_statistics
            else:
                self.current_statistics = self._rebuild_statistics_from_records(self.current_records)
                self.session_statistics = self.current_statistics
        self._invalidate_filtered_cache()
        self._update_sidebar_nav()
        if refresh_ui:
            self._update_summary_table(self.current_statistics)
            if not self.current_records or not self.current_statistics:
                scope = self._current_scope_label()
                self._show_chart_placeholder(
                    "当前批次暂无分析内容" if scope == "当前批次" else "历史累计暂无可视化内容"
                )
                self._pending_chart_refresh = False
            else:
                self._show_chart_placeholder("正在生成默认图表...")
                self._pending_chart_refresh = True
                self._schedule_chart_render_ready_checks()
        self._refresh_warning_button_state()
        self._refresh_overview_metrics()
        self._update_scope_status_summary()

    def _current_scope_label(self) -> str:
        return "历史累计" if self._current_scope == "history" else "当前批次"

    def _ensure_history_statistics(self):
        if self._history_statistics_dirty or (self.history_records and not self.history_statistics):
            self.history_statistics = self._rebuild_statistics_from_records(self.history_records)
            self._history_statistics_dirty = False

    # ═══════════════════════════════════════════════════════════
    #  文件选择
    # ═══════════════════════════════════════════════════════════
    def _select_folder(self):
        path = QFileDialog.getExistingDirectory(self, "选择 SUM 文件夹（支持本地或网络共享，只读解析）")
        if path:
            self._folder_edit.setText(path)
            self._reset_processing_controls_for_new_input(clear_excel_state=False)
            self._set_status(f"已选择目录: {path}")

    def _select_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件（请确保已从服务器下载到本地）",
            "", "Excel 文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if path:
            self._file_edit.setText(path)
            self._reset_processing_controls_for_new_input(clear_excel_state=True)
            self._set_status(f"已选择文件: {path}")

    def _reset_processing_controls_for_new_input(self, clear_excel_state: bool = False):
        self._btn_sum.setEnabled(True)
        self._btn_sum.setText("处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_blue")
        self._btn_process.setEnabled(True)
        self._btn_process.setText("分析Excel")
        _apply_btn_style(self._btn_process, "btn_primary")

        if not clear_excel_state:
            return

        self.session_records = []
        self.session_statistics = {}
        self.session_error_records = []
        self.session_deduplication_info = {}
        self.current_records = []
        self.current_statistics = {}
        self._invalidate_filtered_cache()
        self._warning_messages = []
        self._warning_by_probe = {}
        self._current_scope = "current"
        self._update_summary_table({})
        self._refresh_overview_metrics()
        self._refresh_warning_button_state()
        self._show_chart_placeholder("当前批次暂无分析内容")
        self._set_status("已选择新的 Excel 文件，等待分析")

    # ═══════════════════════════════════════════════════════════
    #  处理文件夹（SUM）
    # ═══════════════════════════════════════════════════════════
    def _set_sum_processing_controls(self, busy_text: str):
        self._btn_sum.setEnabled(False)
        self._btn_process.setEnabled(False)
        self._btn_sum.setText(busy_text)
        self._btn_process.setText("分析Excel")
        _apply_btn_style(self._btn_sum, "btn_blue")
        _apply_btn_style(self._btn_process, "btn_primary")

    def _restore_sum_processing_controls(self, completed: bool = False):
        self._btn_sum.setEnabled(True)
        self._btn_process.setEnabled(True)
        self._btn_process.setText("分析Excel")
        _apply_btn_style(self._btn_process, "btn_primary")

        self._btn_sum.setText("处理完成" if completed else "处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_green" if completed else "btn_blue")

    def _worker_is_active(self) -> bool:
        worker = getattr(self, '_current_worker', None)
        return worker is not None

    def _process_sum_files(self):
        if self._worker_is_active():
            QMessageBox.information(self, "提示", "当前已有任务正在执行，请等待完成或先取消当前任务。")
            return
        folder = self._folder_edit.text().strip()
        if not folder:
            QMessageBox.warning(self, "警告", "请先选择 SUM 文件夹")
            return
        if not Path(folder).exists():
            QMessageBox.warning(self, "警告", "所选 SUM 文件夹不存在")
            return
        self._set_sum_processing_controls("处理中...")
        self._set_status("正在扫描并处理 SUM 文件夹...", 0)
        w = _Worker(self._do_process_sum_auto, folder)
        self._current_worker = w
        self._prog_dlg = ProgressDialog("处理文件夹进度", self, cancellable=True)
        self._prog_dlg.cancel_requested.connect(w.cancel)
        self._prog_dlg.show_and_paint("正在准备处理文件夹...", 0)
        w.signals.progress.connect(lambda t, v: (self._set_status(t, v), self._prog_dlg.set_progress(t, v)))
        w.signals.sum_done.connect(self._on_sum_complete)
        w.signals.error.connect(self._on_worker_error)
        w.signals.cancelled.connect(self._on_worker_cancelled)
        self._pool.start(w)

    def _do_process_sum_auto(self, signals: _WorkerSignals, worker: _Worker, root: str):
        worker.raise_if_cancelled()
        root_path = Path(root)
        signals.progress.emit("正在扫描包含 SUM 文件的文件夹...", 2)
        folders = _find_sum_data_folders(root_path)
        worker.raise_if_cancelled()
        if not folders:
            signals.sum_done.emit({
                "data": [],
                "total_count": 0,
                "folder_count": 0,
                "default_filename": f"SUM批量处理结果_{root_path.name}.xlsx",
                "skipped_invalid_groups": [],
                "folder_summaries": [],
            })
            return

        all_rows = []
        all_skipped = []
        folder_summaries = []
        total_folders = len(folders)

        for index, folder in enumerate(folders, 1):
            worker.raise_if_cancelled()
            parser = SumFileParser()
            folder_label = _relative_path_text(folder, root_path)
            base_pct = 4 + ((index - 1) / max(total_folders, 1)) * 92
            signals.progress.emit(f"处理文件夹 {index}/{total_folders}: {folder_label}", base_pct)

            def _cb(current, total, name, idx=index, label=folder_label):
                worker.raise_if_cancelled()
                folder_pct = (current / max(total, 1)) * (92 / max(total_folders, 1))
                pct = 4 + ((idx - 1) / max(total_folders, 1)) * 92 + folder_pct
                signals.progress.emit(f"{label} | 解析 {name} ({current}/{total})", pct)

            records = parser.parse_directory(folder, progress_callback=_cb, recursive=False)
            worker.raise_if_cancelled()
            rows = parser.records_to_dict_list(records)
            for row in rows:
                row["来源文件夹"] = str(folder)
                row["相对文件夹"] = _relative_path_text(folder, root_path)
            all_rows.extend(rows)
            for item in parser.invalid_group_errors:
                enriched_item = dict(item)
                group_dir = Path(str(enriched_item.get('group_dir', '')))
                enriched_item["relative_group_dir"] = _relative_path_text(group_dir, root_path) if str(group_dir) else ""
                all_skipped.append(enriched_item)
            folder_summaries.append({
                "folder": str(folder),
                "relative_folder": str(folder_label),
                "record_count": len(rows),
                "skipped_count": len(parser.invalid_group_errors),
            })

        signals.progress.emit("批量处理完成，等待保存结果...", 100)
        signals.sum_done.emit({
            "data": all_rows,
            "total_count": len(all_rows),
            "folder_count": total_folders,
            "default_filename": f"SUM文件处理结果_{root_path.name}.xlsx",
            "skipped_invalid_groups": all_skipped,
            "folder_summaries": folder_summaries,
        })

    def _on_sum_complete(self, result: dict):
        self._current_worker = None
        if hasattr(self, '_prog_dlg'):
            try: self._prog_dlg.accept()
            except: pass
        self._restore_sum_processing_controls(True)
        self._set_status("SUM 文件处理完成", 100)
        self._display_sum_results(result)

    def _display_sum_results(self, result: dict):
        data    = result.get("data", [])
        total   = result.get("total_count", len(data))
        skipped = result.get("skipped_invalid_groups", [])
        default = result.get("default_filename", "SUM文件处理结果.xlsx")
        default_save_path = str(HISTORY_STORE_PATH.parent / default)
        folder_summaries = result.get("folder_summaries", [])
        folder_count = result.get("folder_count", 1 if data else 0)

        if not data and not skipped and not folder_summaries:
            QMessageBox.warning(self, "警告", "没有可显示的数据。请确认选择目录下存在可解析的 SUM 文件及完整的 999 标记 ECT 文件。")
            return
        if skipped:
            lines = "\n".join(
                f"- {item.get('data_group') or Path(item.get('group_dir', '')).name}: {item.get('error_type') or item.get('reason') or '解析失败'}"
                for item in skipped[:12]
            )
            if len(skipped) > 12:
                lines += f"\n... 其余 {len(skipped) - 12} 个异常数据组已省略"
            QMessageBox.warning(self, "异常数据已跳过",
                "以下 SUM 文件或数据组存在异常并已跳过，完整明细会写入导出文件：\n" + lines)

        path, _ = QFileDialog.getSaveFileName(
            self, "保存 SUM 文件处理结果", default_save_path,
            "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if not path:
            QMessageBox.information(self, "提示", f"已解析 {folder_count} 个文件夹，生成 {total} 条记录，跳过 {len(skipped)} 个异常组（未保存）")
            return
        source_folder = self._folder_edit.text().strip()
        if source_folder and _path_is_inside(Path(path), Path(source_folder)):
            QMessageBox.warning(self, "禁止保存到源目录", "为保护原始数据，处理结果不能保存到当前导入的文件夹或其子文件夹。请保存到本地 SaveDate 或其他输出目录。")
            self._set_status("已取消保存 - 禁止写入源目录")
            return

        try:
            self._export_sum_result_workbook(path, data, folder_summaries, skipped)
        except PermissionError:
            QMessageBox.critical(self, "保存失败", "目标文件可能正在被 Excel/WPS 打开，请关闭后重试。")
            self._set_status("保存失败 - 文件被占用")
            return
        except Exception as exc:
            logger.exception("保存 SUM 处理结果失败")
            QMessageBox.critical(self, "保存失败", f"保存处理结果时发生错误:\n{exc}")
            self._set_status("保存失败")
            return
        if total:
            QMessageBox.information(self, "成功", f"成功处理 {folder_count} 个文件夹，生成 {total} 条记录，跳过 {len(skipped)} 个异常组\n已保存到:\n{path}")
        else:
            QMessageBox.warning(self, "处理完成但无有效记录", f"未生成有效记录，已导出 {len(skipped)} 个异常组诊断信息:\n{path}")
        self._set_status(f"处理完成 - {folder_count} 个文件夹，{total} 条记录，跳过 {len(skipped)} 个异常组")

    @staticmethod
    def _fit_worksheet_columns(worksheet, max_width: int):
        for col in worksheet.columns:
            ml = max((len(str(c.value or '')) for c in col), default=0)
            worksheet.column_dimensions[col[0].column_letter].width = min(ml + 2, max_width)

    @staticmethod
    def _export_sum_result_workbook(
        path: str,
        data: List[Dict],
        folder_summaries: List[Dict] | None = None,
        skipped: List[Dict] | None = None,
    ):
        import pandas as pd

        folder_summaries = folder_summaries or []
        skipped = skipped or []
        cols = ['Outage', 'SG_ID', 'Data Group', 'Operator',
                'Probe Type', 'Probe SN', 'Model', 'Tube Number', 'Start Time', 'End Time']
        for optional_col in ('相对文件夹', '来源文件夹'):
            if any(optional_col in row for row in data) or folder_summaries:
                cols.append(optional_col)
        filtered = [{c: r.get(c, '') for c in cols} for r in data]
        df = pd.DataFrame(filtered, columns=cols)
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='SUM文件数据', index=False)
            ws = writer.sheets['SUM文件数据']
            MainWindow._fit_worksheet_columns(ws, 50)
            if folder_summaries:
                summary_rows = [
                    {
                        '序号': i,
                        '相对文件夹': item.get('relative_folder', ''),
                        '完整路径': item.get('folder', ''),
                        '有效记录数': item.get('record_count', 0),
                        '跳过异常组数': item.get('skipped_count', 0),
                    }
                    for i, item in enumerate(folder_summaries, 1)
                ]
                summary_df = pd.DataFrame(summary_rows)
                summary_df.to_excel(writer, sheet_name='批量处理摘要', index=False)
                sws = writer.sheets['批量处理摘要']
                MainWindow._fit_worksheet_columns(sws, 70)
            if skipped:
                skipped_rows = [
                    {
                        '序号': i,
                        '数据组': item.get('data_group', ''),
                        '相对异常目录': item.get('relative_group_dir', ''),
                        '异常目录': item.get('group_dir', ''),
                        'SUM文件': item.get('sum_file', ''),
                        '错误类型': item.get('error_type', ''),
                        '错误范围': item.get('error_scope', ''),
                        '原因': item.get('reason', ''),
                    }
                    for i, item in enumerate(skipped, 1)
                ]
                skipped_df = pd.DataFrame(skipped_rows)
                skipped_df.to_excel(writer, sheet_name='跳过异常组', index=False)
                ews = writer.sheets['跳过异常组']
                MainWindow._fit_worksheet_columns(ews, 80)

    # ═══════════════════════════════════════════════════════════
    #  分析 Excel
    # ═══════════════════════════════════════════════════════════
    def _start_processing(self):
        if self._worker_is_active():
            QMessageBox.information(self, "提示", "当前已有任务正在执行，请等待完成或先取消当前任务。")
            return
        file_path = self._file_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "警告", "请先选择 Excel 文件")
            return
        if not Path(file_path).exists():
            QMessageBox.warning(self, "警告", "所选 Excel 文件不存在")
            return
        self._warning_messages.clear()
        self._warning_by_probe.clear()
        self.session_error_records = []
        self._btn_process.setEnabled(False)
        self._btn_process.setText("分析中...")
        _apply_btn_style(self._btn_process, "btn_primary")
        self._btn_sum.setEnabled(False)
        self._btn_sum.setText("处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_blue")
        self._btn_warnings.setEnabled(False)
        self._btn_warnings.setText("异常提醒")
        _apply_btn_style(self._btn_warnings, "btn_warning")
        self._set_status("正在分析...", 0)
        w = _Worker(self._do_process_excel, file_path)
        self._current_worker = w
        self._prog_dlg = ProgressDialog("分析Excel进度", self, cancellable=True)
        self._prog_dlg.cancel_requested.connect(w.cancel)
        self._prog_dlg.show_and_paint("正在准备分析 Excel...", 0)
        w.signals.progress.connect(lambda t, v: (self._set_status(t, v), self._prog_dlg.set_progress(t, v)))
        w.signals.success.connect(self._on_excel_success)
        w.signals.error.connect(self._on_worker_error)
        w.signals.cancelled.connect(self._on_worker_cancelled)
        self._pool.start(w)

    def _do_process_excel(self, signals: _WorkerSignals, worker: _Worker, file_path: str):
        worker.raise_if_cancelled()
        signals.progress.emit("正在打开 Excel 文件...", 0)
        extractor = SummaryFileExtractor()
        analyzer = ProbeAnalyzer()
        warning_logs = []

        class _ThreadWarningCollector(logging.Handler):
            def __init__(self):
                super().__init__(level=logging.WARNING)
                self.setFormatter(logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
                ))

            def emit(self, record):
                try:
                    warning_logs.append({
                        'formatted': self.format(record),
                        'message': record.getMessage(),
                        'logger': record.name,
                    })
                except Exception:
                    pass

        def _ext_cb(current, total, message):
            worker.raise_if_cancelled()
            signals.progress.emit(message or f"解析 Excel {current}/{total}", (current / max(total, 1)) * 40)

        def _ana_cb(stage, current, total, message):
            worker.raise_if_cancelled()
            base = 40 if stage == 'deduplicate' else 70
            signals.progress.emit(message or "正在分析探头统计...", base + (current / max(total, 1)) * 30)

        loggers = [logging.getLogger('pies_pyqt.core.extractor'), logging.getLogger('pies_pyqt.core.analyzer')]
        warning_handler = _ThreadWarningCollector()
        for lg in loggers:
            lg.addHandler(warning_handler)
        try:
            records = extractor.extract_probe_records(file_path, progress_callback=_ext_cb)
            worker.raise_if_cancelled()
            extracted_count = len(records)
            signals.progress.emit("正在整理提取结果...", 40)
            analyzer.add_records(records)
            statistics = analyzer.analyze(progress_callback=_ana_cb)
            worker.raise_if_cancelled()
        finally:
            for lg in loggers:
                try:
                    lg.removeHandler(warning_handler)
                except Exception:
                    pass
        unique_records = list(analyzer.records)
        dedup_info = dict(analyzer.deduplication_info or {})
        if dedup_info:
            dedup_info['original_count'] = extracted_count
        signals.progress.emit("正在刷新界面...", 97)
        signals.success.emit({
            'records': unique_records,
            'statistics': statistics,
            'deduplication_info': dedup_info,
            'error_records': list(getattr(extractor, 'error_records', []) or []),
            'warning_logs': warning_logs,
        })

    def _on_excel_success(self, data: dict):
        self._current_worker = None
        try:
            self._on_excel_success_impl(data)
        except Exception as exc:
            logger.exception("Excel 成功回调刷新界面失败")
            if hasattr(self, '_prog_dlg'):
                try: self._prog_dlg.accept()
                except Exception: pass
            self._btn_process.setEnabled(True)
            self._btn_process.setText("分析Excel")
            _apply_btn_style(self._btn_process, "btn_primary")
            self._btn_sum.setEnabled(True)
            self._btn_sum.setText("处理文件夹")
            _apply_btn_style(self._btn_sum, "btn_blue")
            self._set_status("刷新界面失败")
            QMessageBox.critical(self, "错误", f"Excel 已解析，但刷新界面失败:\n{exc}")

    def _on_excel_success_impl(self, data: dict):
        if hasattr(self, '_prog_dlg'):
            try: self._prog_dlg.accept()
            except: pass
        records    = data['records']
        statistics = data['statistics']
        dedup_info = data.get('deduplication_info', {})

        self.session_records    = list(records)
        self.session_statistics = dict(statistics)
        self.session_deduplication_info = dict(dedup_info)
        self.session_error_records = list(data.get('error_records') or [])
        self._warning_messages = []
        self._warning_by_probe = {}

        # 从 extractor.error_records 收集结构化警告
        for err in self.session_error_records:
            probe_sn = str(err.get('探头编号', '') or err.get('probe_sn', '') or '未知探头')
            msg = str(err.get('错误信息', '') or err.get('错误类型', '') or '')
            if not msg:
                continue
            if msg not in self._warning_messages:
                self._warning_messages.append(msg)
            operator = str(err.get('操作员', '') or '')
            line_number = str(err.get('行号', '') or '')
            key = self._warning_merge_key(probe_sn, operator, line_number)
            if key not in self._warning_by_probe:
                self._warning_by_probe[key] = {
                    'probe_sn': probe_sn, 'warnings': [], 'details': [],
                    'warning_types': set(),
                    'outage': str(err.get('大修编号', '') or ''),
                    'sg_id':  str(err.get('SG ID', '') or ''),
                    'data_group': str(err.get('数据组', '') or ''),
                    'operator':   str(err.get('操作员', '') or ''),
                    'probe_type': str(err.get('探头类型', '') or ''),
                    'model':      str(err.get('探头型号', '') or ''),
                    'tube_number': str(err.get('管道数量', '') or ''),
                    'start_time': str(err.get('开始时间', '') or ''),
                    'end_time':   str(err.get('结束时间', '') or ''),
                    'line_number': str(err.get('行号', '') or ''),
                }
            probe_info = self._warning_by_probe[key]
            if not self._normalize_warning_context_value(probe_info.get('line_number', '')):
                probe_info['line_number'] = line_number
            probe_info.setdefault('warning_types', set()).add(msg)
            if msg not in probe_info['warnings']:
                probe_info['warnings'].append(msg)
            probe_info['details'].append({'raw_text': str(err), 'message': msg, 'type': msg})

        # 从 records 的 warnings 字段收集
        for r in records:
            for w in (getattr(r, 'warnings', []) or []):
                if not w:
                    continue
                if w not in self._warning_messages:
                    self._warning_messages.append(w)
                probe_sn = r.probe_sn
                operator = getattr(r, 'operator', '') or ''
                line_number = str(getattr(r, 'warning_line_number', '') or '')
                data_group = getattr(r, 'data_group', '') or ''
                context = {
                    'probe_sn': probe_sn,
                    'operator': operator,
                    'line_number': line_number,
                    'outage': getattr(r, 'outage', ''),
                    'sg_id': getattr(r, 'sg_id', ''),
                    'data_group': data_group,
                    'probe_type': getattr(r, 'probe_type_raw', None) or getattr(r, 'probe_type', ''),
                    'model': getattr(r, 'model', ''),
                    'tube_number': str(getattr(r, 'tube_number', '') or ''),
                    'start_time': r.start_time.strftime('%Y-%m-%d %H:%M:%S') if getattr(r, 'start_time', None) else '',
                    'end_time': r.end_time.strftime('%Y-%m-%d %H:%M:%S') if getattr(r, 'end_time', None) else '',
                }
                if not line_number:
                    continue
                key = self._find_warning_group_key(context, w) or self._warning_merge_key(probe_sn, operator, line_number)
                if key not in self._warning_by_probe:
                    self._warning_by_probe[key] = {
                        'probe_sn': probe_sn, 'warnings': [], 'details': [],
                        'warning_types': set(),
                        **context,
                    }
                probe_info = self._warning_by_probe[key]
                self._apply_warning_context(probe_info, context)
                probe_info.setdefault('warning_types', set()).add(w)
                if w not in probe_info['warnings']:
                    probe_info['warnings'].append(w)
                probe_info.setdefault('details', []).append({'raw_text': w, 'message': w, 'type': w})

        for item in (data.get('warning_logs') or []):
            self._capture_warning_message(
                str(item.get('message', '') or ''),
                str(item.get('formatted', '') or item.get('message', '') or ''),
            )

        # 累计历史
        if self._history_enabled:
            self._append_to_history_records(
                records,
                warning_groups=self._warning_by_probe,
                warning_messages=self._warning_messages,
                error_records=self.session_error_records,
            )

        self._current_scope = "current"
        self._apply_active_dataset(refresh_ui=False)
        self._update_summary_table(self.current_statistics)
        self._refresh_overview_metrics()
        self._refresh_warning_button_state()
        self._btn_process.setEnabled(True)
        self._btn_process.setText("处理完成")
        _apply_btn_style(self._btn_process, "btn_green")
        self._btn_sum.setEnabled(True)
        self._btn_sum.setText("处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_blue")
        self._set_status(f"分析完成 - {len(records)} 条记录，{len(statistics)} 个探头", 100)
        self._show_chart_placeholder("正在生成默认图表...")
        self._pending_chart_refresh = True
        QTimer.singleShot(180, lambda: self._ensure_chart_render_ready(force=True))

    def _on_worker_error(self, msg: str):
        self._current_worker = None
        if hasattr(self, '_prog_dlg'):
            try: self._prog_dlg.accept()
            except: pass
        self._uninstall_warning_handler()
        self._restore_sum_processing_controls()
        self._btn_process.setEnabled(True)
        self._btn_process.setText("分析Excel")
        _apply_btn_style(self._btn_process, "btn_primary")
        self._btn_sum.setEnabled(True)
        self._btn_sum.setText("处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_blue")
        self._refresh_warning_button_state()
        self._set_status("处理失败")
        QMessageBox.critical(self, "错误", f"处理失败:\n{msg}")

    def _on_worker_cancelled(self, msg: str):
        self._current_worker = None
        if hasattr(self, '_prog_dlg'):
            try: self._prog_dlg.accept()
            except Exception: pass
        self._restore_sum_processing_controls()
        self._btn_process.setEnabled(True)
        self._btn_process.setText("分析Excel")
        _apply_btn_style(self._btn_process, "btn_primary")
        self._btn_sum.setEnabled(True)
        self._btn_sum.setText("处理文件夹")
        _apply_btn_style(self._btn_sum, "btn_blue")
        self._btn_warnings.setEnabled(bool(self._warning_messages or self._warning_by_probe))
        self._set_status("任务已取消")
        QMessageBox.information(self, "已取消", msg or "当前任务已取消，未更新当前数据。")

    # ═══════════════════════════════════════════════════════════
    #  表格更新
    # ═══════════════════════════════════════════════════════════
    def _update_summary_table(self, statistics: dict | None = None):
        tbl = self._summary_table
        tbl.setSortingEnabled(False)
        tbl.setUpdatesEnabled(False)
        tbl.setRowCount(0)
        if self._get_active_filter_items():
            filtered, display_statistics = self._get_filtered_records_and_stats()
        else:
            filtered = self.current_records
            display_statistics = statistics or self.current_statistics or {}
        if not filtered or not display_statistics:
            tbl.setUpdatesEnabled(True)
            return
        tbl.setRowCount(len(filtered))

        for idx, record in enumerate(filtered):
            stat: ProbeStatistics = display_statistics.get(record.stat_key)
            if not stat:
                continue
            start_time = getattr(record, 'start_time', None)
            end_time = getattr(record, 'end_time', None)
            start_str = start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else ''
            end_str = end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else ''
            if start_time and end_time:
                diff_s = (end_time - start_time).total_seconds()
                sh, sm = diff_s / 3600.0, diff_s / 60.0
            else:
                sh = sm = 0.0
            first_use_time = getattr(stat, 'first_use_time', None)
            last_use_time = getattr(stat, 'last_use_time', None)
            first_use = first_use_time.strftime('%Y-%m-%d') if first_use_time else ''
            last_use = last_use_time.strftime('%Y-%m-%d') if last_use_time else ''
            probe_type_display = getattr(record, 'probe_type_raw', None) or getattr(stat, 'probe_type', '')
            longest = float(getattr(stat, 'longest_continuous_duration_minutes', 0.0) or 0.0)
            total_uses = int(getattr(stat, 'total_uses', 0) or 0)
            total_duration_minutes = float(getattr(stat, 'total_duration_minutes', 0.0) or 0.0)
            detection_speed = float(getattr(stat, 'detection_speed', 0.0) or 0.0)
            unique_tube_count = int(getattr(stat, 'unique_tube_count', 0) or 0)

            values = [
                str(idx + 1),
                getattr(record, 'outage', '') or '',
                getattr(record, 'sg_id', '') or '',
                record.data_group,
                record.operator,
                probe_type_display,
                record.probe_sn,
                record.model,
                str(record.tube_number) if record.tube_number is not None else '',
                str(unique_tube_count),
                start_str, end_str,
                f"{sh:.2f}", f"{sm:.2f}",
                str(total_uses),
                f"{total_duration_minutes / 60.0:.2f}",
                f"{longest / 60.0:.2f}",
                f"{detection_speed:.2f}",
                first_use, last_use,
            ]

            has_warning = bool(getattr(record, 'warnings', None))
            row_bg = QColor("#FDEDEC") if has_warning else (QColor("#FAFBFC") if idx % 2 == 1 else QColor("#FFFFFF"))
            row_fg = QColor("#D84A3A") if has_warning else QColor(TEXT_PRI)

            for col, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(row_bg)
                item.setForeground(row_fg)
                tbl.setItem(idx, col, item)

        tbl.setUpdatesEnabled(True)
        tbl.viewport().update()
        self._refresh_overview_metrics()

    def _on_table_double_click(self, index):
        row = index.row()
        try:
            col_probe = SUMMARY_TABLE_HEADERS.index("探头编码")
        except ValueError:
            col_probe = 6
        item = self._summary_table.item(row, col_probe)
        if item:
            probe_sn = item.text().strip()
            if probe_sn:
                self.summary_filter_values['探头编码'] = probe_sn
                self.filter_values['探头编码'] = probe_sn
                self._update_summary_table(self.current_statistics)
                self._schedule_chart_refresh()
                self._set_status(f"已按探头 {probe_sn} 聚焦")
                self._refresh_overview_metrics()
                self._update_filter_summary_label()

    def _on_table_context_menu(self, pos):
        menu = QMenu(self)
        menu.addAction("清除所有筛选", self._clear_all_filters)
        menu.addAction("重置列宽", self._apply_summary_column_widths)
        menu.exec_(self._summary_table.viewport().mapToGlobal(pos))

    def _on_header_context_menu(self, pos):
        col = self._summary_table.horizontalHeader().logicalIndexAt(pos)
        menu = QMenu(self)
        menu.addAction("重置所有列宽", self._apply_summary_column_widths)
        if col >= 0 and col < len(SUMMARY_TABLE_HEADERS):
            col_name = SUMMARY_TABLE_HEADERS[col]
            if col_name in TABLE_HEADER_FILTERABLE_COLUMNS:
                menu.addSeparator()
                menu.addAction(f"筛选 {col_name}...", lambda: self._show_column_filter_menu(col_name))
                if self._filter_value_is_active(self.summary_filter_values.get(col_name)):
                    menu.addAction(f"清除 {col_name} 筛选", lambda: self._set_summary_filter(col_name, None))
        menu.exec_(self._summary_table.horizontalHeader().mapToGlobal(pos))

    def _on_header_clicked(self, col: int):
        if col < 0 or col >= len(SUMMARY_TABLE_HEADERS):
            return
        col_name = SUMMARY_TABLE_HEADERS[col]
        if col_name not in TABLE_HEADER_FILTERABLE_COLUMNS:
            return
        header = self._summary_table.horizontalHeader()
        x = header.sectionViewportPosition(col)
        pos = header.mapToGlobal(header.rect().topLeft())
        pos.setX(pos.x() + x)
        pos.setY(pos.y() + header.height())
        self._show_column_filter_menu_at(col_name, pos)

    def _show_column_filter_menu(self, col_name: str):
        self._show_column_filter_menu_at(col_name, QCursor.pos())

    def _show_column_filter_menu_at(self, col_name: str, global_pos):
        values = self._get_available_filter_values(col_name)
        if not values:
            return
        dlg = QDialog(self, Qt.Popup)
        dlg.setObjectName("filter_popup")
        dlg.setWindowTitle(f"筛选 {col_name}")
        dlg.setWindowFlags((dlg.windowFlags() | Qt.FramelessWindowHint) & ~Qt.WindowContextHelpButtonHint)
        dlg.resize(220, 326)
        dlg.setMinimumSize(220, 280)
        dlg.setMaximumSize(420, 560)
        dlg.setStyleSheet(STYLESHEET)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(6)

        title = QLabel(f"{col_name} 筛选")
        title.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
        title.setStyleSheet(f"color: {TEXT_PRI}; background: transparent;")
        lay.addWidget(title)

        current_items = set(self._filter_value_to_list(self.summary_filter_values.get(col_name)))
        exact_values = set(values)
        manual_initial = self._format_filter_value_text(current_items) if current_items and not current_items.issubset(exact_values) else ""
        manual_edit = QLineEdit()
        manual_edit.setPlaceholderText("手动输入筛选关键字")
        manual_edit.setFixedHeight(26)
        manual_edit.setText(manual_initial)
        lay.addWidget(manual_edit)

        lst = QListWidget()
        lst.setObjectName("filter_list")
        lst.setSelectionMode(QAbstractItemView.NoSelection)
        all_item = QListWidgetItem("全部")
        all_item.setFlags(all_item.flags() | Qt.ItemIsUserCheckable)
        all_item.setCheckState(Qt.Checked if not current_items and not manual_initial else Qt.Unchecked)
        lst.addItem(all_item)
        for v in values:
            item = QListWidgetItem(v)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if not manual_initial and v in current_items else Qt.Unchecked)
            lst.addItem(item)
        lay.addWidget(lst, 1)

        def on_item_changed(item):
            if item.checkState() == Qt.Checked and manual_edit.text().strip():
                manual_edit.blockSignals(True)
                manual_edit.clear()
                manual_edit.blockSignals(False)
            if item is all_item and item.checkState() == Qt.Checked:
                for i in range(1, lst.count()):
                    lst.item(i).setCheckState(Qt.Unchecked)
            elif item is not all_item and item.checkState() == Qt.Checked:
                all_item.setCheckState(Qt.Unchecked)

        lst.itemChanged.connect(on_item_changed)

        def on_manual_text_edited(text: str):
            if not text.strip():
                return
            lst.blockSignals(True)
            all_item.setCheckState(Qt.Unchecked)
            for i in range(1, lst.count()):
                lst.item(i).setCheckState(Qt.Unchecked)
            lst.blockSignals(False)

        manual_edit.textEdited.connect(on_manual_text_edited)

        row = QHBoxLayout()
        row.setSpacing(6)
        clear_btn = QPushButton("清除")
        ok_btn = QPushButton("确定")
        clear_btn.setFixedHeight(26)
        ok_btn.setFixedHeight(26)
        clear_btn.setMinimumWidth(58)
        ok_btn.setMinimumWidth(58)
        _apply_btn_style(clear_btn, "btn_secondary")
        _apply_btn_style(ok_btn, "btn_primary")
        row.addWidget(clear_btn)
        row.addStretch(1)
        row.addWidget(ok_btn)
        lay.addLayout(row)

        def apply_filter():
            manual_text = manual_edit.text().strip()
            if manual_text:
                self._set_summary_filter(col_name, manual_text)
            elif all_item.checkState() == Qt.Checked:
                self._set_summary_filter(col_name, None)
            else:
                selected = [
                    lst.item(i).text()
                    for i in range(1, lst.count())
                    if lst.item(i).checkState() == Qt.Checked
                ]
                self._set_summary_filter(col_name, selected or None)
            dlg.accept()

        clear_btn.clicked.connect(lambda: (self._set_summary_filter(col_name, None), dlg.accept()))
        ok_btn.clicked.connect(apply_filter)
        dlg.move(global_pos)
        self._present_dialog(dlg, key=f"filter_popup:{col_name}")

    # ═══════════════════════════════════════════════════════════
    #  图表
    # ═══════════════════════════════════════════════════════════
    def _on_chart_type_changed(self, _group: str):
        self._update_chart_option_combo()
        if self.current_records and self.current_statistics:
            self._show_chart_placeholder("正在切换图表...")
        self._schedule_chart_refresh(50)

    def _update_chart_option_combo(self):
        group = self._chart_type_combo.currentText()
        options = CHART_GROUPS.get(group, [])
        self._chart_option_combo.blockSignals(True)
        self._chart_option_combo.clear()
        for label, _ in options:
            self._chart_option_combo.addItem(label)
        self._chart_option_combo.blockSignals(False)

    def _schedule_chart_refresh(self, delay_ms: int = 30):
        if not self.current_records or not self.current_statistics:
            return
        if self._current_view == "table" or not self._chart_panel.isVisible():
            self._pending_chart_refresh = True
            return
        self._chart_request_token += 1
        if self._chart_refresh_timer is None:
            self._chart_refresh_timer = QTimer(self)
            self._chart_refresh_timer.setSingleShot(True)
            self._chart_refresh_timer.timeout.connect(self._refresh_chart)
        self._chart_refresh_timer.stop()
        self._pending_chart_refresh = True
        self._chart_refresh_timer.start(delay_ms)

    def _schedule_chart_render_ready_checks(self):
        for delay in (0, 90, 220, 420):
            QTimer.singleShot(delay, lambda: self._ensure_chart_render_ready(force=True))

    def _ensure_chart_render_ready(self, force: bool = False):
        if not self.current_records or not self.current_statistics:
            self._pending_chart_refresh = False
            return
        if self._current_view == "table":
            self._pending_chart_refresh = True
            return
        if not force and not self._pending_chart_refresh and self._current_chart_figure:
            return
        if self._chart_scroll.viewport().width() <= 0 or self._chart_scroll.viewport().height() <= 0:
            self._pending_chart_refresh = True
            QTimer.singleShot(120, lambda: self._ensure_chart_render_ready(force=True))
            return
        self._pending_chart_refresh = False
        self._schedule_chart_refresh(80)

    def _get_current_chart_key(self) -> Tuple[str, str, str]:
        group = self._chart_type_combo.currentText()
        label = self._chart_option_combo.currentText()
        for lbl, key in CHART_GROUPS.get(group, []):
            if lbl == label:
                return group, label, key
        opts = CHART_GROUPS.get(group, [])
        if opts:
            return group, opts[0][0], opts[0][1]
        return group, label, label

    def _get_chart_source_data(self):
        if self._get_active_filter_items():
            return (*self._get_filtered_records_and_stats(), 0)
        return self.current_records, self.current_statistics, 0

    def _refresh_chart(self):
        try:
            self._refresh_chart_impl()
        except Exception as exc:
            logger.exception("刷新图表失败")
            self._is_refreshing = False
            self._show_chart_placeholder(f"图表生成失败：{exc}")

    def _refresh_chart_impl(self):
        if not self.current_statistics or not self.current_records:
            self._show_chart_placeholder("当前没有可用于绘制图表的数据")
            return
        if self._is_refreshing:
            return
        if self._chart_scroll.viewport().width() <= 0 or self._chart_scroll.viewport().height() <= 0:
            self._pending_chart_refresh = True
            QTimer.singleShot(120, lambda: self._ensure_chart_render_ready(force=True))
            return
        self._is_refreshing = True
        if self._chart_worker is not None:
            try:
                self._chart_worker.cancel()
            except Exception:
                pass
        request_token = self._chart_request_token
        group, label, chart_key = self._get_current_chart_key()
        cw = max(420, self._chart_scroll.viewport().width() - 20)
        ch = max(340, self._chart_scroll.viewport().height() - 20)
        self._show_chart_placeholder("正在生成图表...")
        worker = _Worker(self._prepare_chart_payload, request_token, group, label, chart_key, cw, ch)
        self._chart_worker = worker
        worker.signals.chart_ready.connect(self._on_chart_payload_ready)
        worker.signals.cancelled.connect(self._on_chart_payload_cancelled)
        worker.signals.error.connect(self._on_chart_payload_error)
        self._pool.start(worker)

    def _prepare_chart_payload(self, signals: _WorkerSignals, worker: _Worker, request_token: int, group: str, label: str, chart_key: str, cw: int, ch: int):
        worker.raise_if_cancelled()
        records, statistics, hidden = self._get_chart_source_data()
        worker.raise_if_cancelled()
        total_items = len(statistics)
        fig_kwargs = {
            'figure_width':  max(7.8, min(22, cw / 115.0 + total_items * 0.08)),
            'figure_height': max(5.2, min(12.0, ch / 88.0)),
            'batch_figure_width':  max(9.5, min(20, cw / 120.0 + total_items * 0.18)),
            'batch_figure_height': max(5.8, min(10, ch / 92.0)),
        }
        fig = self._create_chart_figure(chart_key, records, statistics, hidden, fig_kwargs)
        worker.raise_if_cancelled()
        signals.chart_ready.emit({
            'request_token': request_token,
            'group': group,
            'label': label,
            'chart_key': chart_key,
            'records_empty': not bool(records),
            'fig': fig,
        })

    def _on_chart_payload_ready(self, payload: dict):
        try:
            request_token = int(payload.get('request_token', -1))
            if request_token != self._chart_request_token:
                self._is_refreshing = False
                return

            chart_key = str(payload.get('chart_key', '') or '')
            group = str(payload.get('group', '') or '')
            label = str(payload.get('label', '') or '')
            records_empty = bool(payload.get('records_empty', False))
            fig = payload.get('fig')

            if records_empty:
                self._show_chart_placeholder("当前筛选条件下没有可用于绘制图表的数据")
                self._is_refreshing = False
                return

            for i in reversed(range(self._chart_container_layout.count())):
                w = self._chart_container_layout.itemAt(i).widget()
                if w:
                    w.setParent(None)
                    w.deleteLater()
            if self._current_chart_figure:
                try:
                    plt.close(self._current_chart_figure)
                except Exception:
                    pass
            self._current_chart_canvas = None
            self._current_chart_figure = None

            if not fig:
                self._show_chart_placeholder("当前图表类型下暂无可展示的数据")
                self._is_refreshing = False
                return
            if request_token != self._chart_request_token:
                try:
                    plt.close(fig)
                except Exception:
                    pass
                self._is_refreshing = False
                return

            display_title = self._get_chart_display_title(group, label)
            try:
                axes = getattr(fig, 'axes', [])
                if len(axes) == 1:
                    axes[0].set_title(display_title, fontsize=12, fontweight='bold', pad=10)
                elif len(axes) > 1:
                    fig.suptitle(display_title, fontsize=12, fontweight='bold', y=0.985)
                self._apply_chart_font_sizes(fig)
            except Exception:
                pass

            try:
                if len(getattr(fig, 'axes', []) or []) <= 1 and getattr(fig, '_pies_export_mode', '') == '':
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        fig.tight_layout(pad=0.8)
                    fig.subplots_adjust(left=0.09, right=0.98, top=0.90, bottom=0.16)
            except Exception:
                pass

            self._attach_chart_hover(fig, chart_key)
            canvas = FigureCanvas(fig)
            canvas.setMinimumHeight(int(fig.get_figheight() * fig.dpi))
            self._chart_container_layout.addWidget(canvas)
            self._current_chart_figure = fig
            self._current_chart_canvas = canvas
            self._current_chart_type = chart_key
            self._pending_chart_refresh = False
            canvas.draw_idle()
        finally:
            self._chart_worker = None
            self._is_refreshing = False

    def _on_chart_payload_cancelled(self, _msg: str):
        self._chart_worker = None
        self._is_refreshing = False

    def _on_chart_payload_error(self, msg: str):
        self._chart_worker = None
        self._is_refreshing = False
        logger.error("图表后台准备失败: %s", msg)
        self._show_chart_placeholder("图表生成失败，请重试")

    def _apply_chart_font_sizes(self, fig):
        for ax in getattr(fig, 'axes', []) or []:
            try:
                ax.title.set_fontsize(11)
                ax.xaxis.label.set_size(9)
                ax.yaxis.label.set_size(9)
                ax.tick_params(axis='both', labelsize=8)
                legend = ax.get_legend()
                if legend:
                    for text in legend.get_texts():
                        text.set_fontsize(8)
            except Exception:
                pass

    def _show_chart_placeholder(self, message: str):
        for i in reversed(range(self._chart_container_layout.count())):
            w = self._chart_container_layout.itemAt(i).widget()
            if w:
                w.setParent(None)
                w.deleteLater()
        # 创建占位符容器
        placeholder = QWidget()
        placeholder.setStyleSheet("background: #FFFFFF;")
        pl = QVBoxLayout(placeholder)
        pl.setAlignment(Qt.AlignCenter)
        lbl = QLabel(message)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setFont(QFont("Microsoft YaHei", 12))
        lbl.setStyleSheet("color: #98A2B3; background: transparent;")
        pl.addWidget(lbl)
        self._chart_container_layout.addWidget(placeholder)
        self._current_chart_canvas = None
        self._current_chart_figure = None

    def _create_chart_figure(self, chart_key, records, statistics, hidden_count=0, fig_kwargs=None):
        fig_kwargs = fig_kwargs or {}
        base = {'figure_width': fig_kwargs.get('figure_width'), 'figure_height': fig_kwargs.get('figure_height')}
        v = self.visualizer
        fig = None
        try:
            if chart_key == "生产批次平均寿命图":
                if self.batch_lifetime_analyzer.load_data_from_statistics(statistics, include_excluded_models=True):
                    fig = self.batch_lifetime_analyzer.create_batch_lifetime_chart(
                        figsize=(fig_kwargs.get('batch_figure_width', 12.0), fig_kwargs.get('batch_figure_height', 7.0)),
                        excluded_probe_count=hidden_count, show_excluded_note=True)
            elif chart_key == "总使用时间折线图":         fig = v.create_lifetime_chart(statistics, **base)
            elif chart_key == "总使用次数折线图":         fig = v.create_usage_chart(statistics, **base)
            elif chart_key == "管道数量折线图":           fig = v.create_tube_count_chart(statistics, **base)
            elif chart_key == "探头检测速度折线图":       fig = v.create_detection_speed_chart(statistics, **base)
            elif chart_key == "探头类型分布图":           fig = v.create_type_distribution_chart(statistics)
            elif chart_key == "探头型号分布饼图":         fig = v.create_model_distribution_pie_chart(records)
            elif chart_key == "探头型号平均检测速度图":   fig = v.create_model_average_speed_chart(records)
        except Exception as exc:
            logger.error(f"创建图表失败 {chart_key}: {exc}")
        return fig

    # ═══════════════════════════════════════════════════════════
    #  图表悬浮效果（matplotlib 事件，与旧版完全一致）
    # ═══════════════════════════════════════════════════════════
    def _attach_chart_hover(self, fig, chart_key: str):
        """根据图表类型绑定对应的悬浮效果。"""
        try:
            if chart_key in ("探头型号分布饼图", "探头类型分布图"):
                self._attach_pie_hover(fig)
            else:
                unit_map = {
                    "总使用时间折线图": "小时", "生产批次平均寿命图": "根",
                    "总使用次数折线图": "次",
                    "管道数量折线图": "根",
                    "探头检测速度折线图": "管道/小时",
                    "探头型号平均检测速度图": "管道/小时",
                }
                unit = unit_map.get(chart_key, "")
                # 判断是折线图还是柱状图
                has_bars = any(
                    len([p for p in ax.patches if hasattr(p, 'get_height')]) > 0
                    for ax in (fig.axes or [])
                )
                if has_bars:
                    self._attach_bar_hover(fig, unit)
                else:
                    self._attach_line_hover(fig, unit)
        except Exception as exc:
            logger.debug(f"attach hover failed: {exc}")

    def _attach_line_hover(self, fig, unit: str = ""):
        """折线图悬浮：高亮线条 + 显示数值标注。"""
        try:
            axes = fig.axes
            if not axes:
                return
            annotations = []
            hover_markers = {}
            last_state = {"line": None, "index": None}
            line_meta = []

            for ax in axes:
                annot = ax.annotate(
                    "", xy=(0, 0), xytext=(10, 12), textcoords="offset points",
                    ha="left", va="bottom", fontsize=9,
                    bbox=dict(boxstyle="round,pad=0.35", fc="#F7FAFC", ec="#A9B8C9", lw=1.0, alpha=0.96),
                )
                annot.set_visible(False)
                annotations.append(annot)
                marker = ax.scatter([], [], s=62, facecolors="#FFFFFF", edgecolors="#6E8FB3", linewidths=1.4, zorder=6)
                marker.set_visible(False)
                hover_markers[ax] = marker
                for line in ax.lines:
                    xd = list(line.get_xdata())
                    yd = list(line.get_ydata())
                    if len(xd) != len(yd) or not xd:
                        continue
                    line_meta.append({
                        "ax": ax, "annot": annot, "line": line,
                        "x": xd, "y": yd,
                        "orig_lw": line.get_linewidth(),
                        "orig_ms": line.get_markersize(),
                        "label": line.get_label(),
                        "color": line.get_color(),
                    })
            if not line_meta:
                return

            def reset():
                if last_state["line"] is not None:
                    m = last_state["line"]
                    try:
                        m["line"].set_linewidth(m["orig_lw"])
                        m["line"].set_markersize(m["orig_ms"])
                    except Exception:
                        pass
                for a in annotations:
                    a.set_visible(False)
                for mk in hover_markers.values():
                    mk.set_visible(False)
                last_state["line"] = None
                last_state["index"] = None

            def on_move(event):
                try:
                    if event.inaxes not in axes or event.xdata is None:
                        reset(); fig.canvas.draw_idle(); return
                    best = None; best_score = None
                    for meta in line_meta:
                        if meta["ax"] != event.inaxes:
                            continue
                        for idx, (xv, yv) in enumerate(zip(meta["x"], meta["y"])):
                            if xv is None or yv is None:
                                continue
                            score = abs(float(xv) - float(event.xdata)) + abs(float(yv) - float(event.ydata)) * 0.25
                            if best_score is None or score < best_score:
                                best = (meta, idx, xv, yv); best_score = score
                    if best is None:
                        reset(); fig.canvas.draw_idle(); return
                    meta, idx, xv, yv = best
                    x_span = max(meta["x"]) - min(meta["x"]) if len(meta["x"]) > 1 else 1.0
                    y_span = max(meta["y"]) - min(meta["y"]) if len(meta["y"]) > 1 else 1.0
                    if abs(float(xv) - float(event.xdata)) > max(0.45, x_span * 0.06) or \
                       abs(float(yv) - float(event.ydata)) > max(0.9, y_span * 0.10):
                        reset(); fig.canvas.draw_idle(); return
                    reset()
                    meta["line"].set_linewidth(meta["orig_lw"] + 0.35)
                    meta["line"].set_markersize(meta["orig_ms"] + 1.2)
                    last_state["line"] = meta; last_state["index"] = idx
                    hm = hover_markers.get(meta["ax"])
                    if hm is not None:
                        hm.set_offsets([[xv, yv]])
                        hm.set_edgecolors([meta["color"]])
                        hm.set_sizes([78])
                        hm.set_visible(True)
                    x_labels = [t.get_text() for t in meta["ax"].get_xticklabels()]
                    x_label = x_labels[idx] if idx < len(x_labels) and x_labels[idx] else str(xv)
                    if unit == "小时":
                        val_text = f"{yv:.1f}"
                    elif unit in {"次", "个", "根"}:
                        val_text = f"{int(yv)}"
                    else:
                        val_text = f"{yv:.2f}"
                    label = meta["label"] if meta["label"] and not meta["label"].startswith("_") else ""
                    text = f"{label}\n{x_label} | {val_text} {unit}".strip()
                    meta["annot"].xy = (xv, yv)
                    meta["annot"].set_text(text)
                    meta["annot"].set_visible(True)
                    fig.canvas.draw_idle()
                except Exception:
                    pass

            fig.canvas.mpl_connect("motion_notify_event", on_move)
        except Exception as exc:
            logger.debug(f"attach line hover failed: {exc}")

    def _attach_bar_hover(self, fig, unit: str = ""):
        """柱状图悬浮：高亮柱子 + 显示数值标注。"""
        try:
            axes = fig.axes
            if not axes:
                return
            bar_infos = []
            for ax in axes:
                bars = [p for p in ax.patches if hasattr(p, 'get_height')]
                if not bars:
                    continue
                tick_labels = [t.get_text() for t in ax.get_xticklabels()]
                for i, bar in enumerate(bars):
                    label = tick_labels[i] if i < len(tick_labels) else ""
                    bar_infos.append((bar, label, ax))
            if not bar_infos:
                return

            bar_geoms = {}
            for bar, _, _ in bar_infos:
                bar_geoms[id(bar)] = {
                    "x": bar.get_x(), "y": bar.get_y(),
                    "w": bar.get_width(), "h": bar.get_height(),
                }

            ax0 = axes[0]
            annot = ax0.annotate(
                "", xy=(0, 0), xytext=(0, 12), textcoords="offset points",
                ha="center", va="bottom", fontsize=9,
                bbox=dict(boxstyle="round,pad=0.25", fc="#F7FAFC", ec="#A9B8C9", lw=0.8),
            )
            annot.set_visible(False)
            last_bar = {"bar": None}

            def _reset():
                if last_bar["bar"] is not None:
                    b = last_bar["bar"]
                    g = bar_geoms[id(b)]
                    try:
                        b.set_x(g["x"]); b.set_y(g["y"])
                        b.set_width(g["w"]); b.set_height(g["h"])
                        b.set_linewidth(0); b.set_alpha(0.85)
                    except Exception:
                        pass
                    last_bar["bar"] = None

            def on_hover(event):
                if event.inaxes not in axes:
                    _reset(); annot.set_visible(False)
                    try: fig.canvas.draw_idle()
                    except: pass
                    return
                for bar, label, ax in bar_infos:
                    contains, _ = bar.contains(event)
                    if contains:
                        _reset(); last_bar["bar"] = bar
                        g = bar_geoms[id(bar)]
                        scale = 1.03
                        try:
                            bar.set_x(g["x"] - g["w"] * (scale - 1) / 2)
                            bar.set_width(g["w"] * scale)
                            bar.set_y(g["y"])
                            bar.set_height(g["h"] * scale)
                            bar.set_linewidth(1.5)
                            bar.set_edgecolor("#6E8FB3")
                            bar.set_alpha(1.0)
                        except Exception:
                            pass
                        h = bar.get_height()
                        if unit in {"次", "个", "根"}:
                            vs = f"{int(h)}"
                        elif unit == "小时":
                            vs = f"{h:.1f}"
                        else:
                            vs = f"{h:.2f}"
                        text = f"{label} | {vs} {unit}" if label else f"{vs} {unit}"
                        annot.xy = (bar.get_x() + bar.get_width() / 2, h)
                        annot.set_text(text)
                        annot.set_visible(True)
                        try: fig.canvas.draw_idle()
                        except: pass
                        return
                _reset(); annot.set_visible(False)
                try: fig.canvas.draw_idle()
                except: pass

            fig.canvas.mpl_connect("motion_notify_event", on_hover)
        except Exception as exc:
            logger.debug(f"attach bar hover failed: {exc}")

    def _attach_pie_hover(self, fig):
        """饼图悬浮：高亮扇区 + 显示占比标注。"""
        try:
            import numpy as np
            axes = fig.axes
            if not axes:
                return
            pie_ax = axes[0]
            bar_ax = axes[1] if len(axes) > 1 else None

            wedges = [p for p in pie_ax.patches if hasattr(p, 'theta1')]
            if not wedges:
                return
            bars = [p for p in bar_ax.patches if hasattr(p, 'get_height')] if bar_ax else []
            bar_geoms = {id(b): {"x": b.get_x(), "y": b.get_y(), "w": b.get_width(), "h": b.get_height()} for b in bars}

            labels = [t.get_text() for t in bar_ax.get_xticklabels()] if bar_ax else []
            if not labels or len(labels) != len(wedges):
                labels = [w.get_label() or f"类型{i+1}" for i, w in enumerate(wedges)]
            labels = labels[:len(wedges)]
            values = [int(b.get_height()) for b in bars] if bars else [1] * len(wedges)
            values = (values + [0] * len(wedges))[:len(wedges)]
            total = sum(values) or 1

            orig_fc    = [w.get_facecolor() for w in wedges]
            orig_alpha = [w.get_alpha() for w in wedges]
            orig_r     = [getattr(w, 'r', 1.0) for w in wedges]

            pie_annot = pie_ax.annotate(
                "", xy=(0, 0), xytext=(22, 22), textcoords="offset points",
                ha="left", va="bottom", fontsize=12, fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.7", fc="#F7FAFC", ec="#A9B8C9", lw=1.4, alpha=0.94),
                arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.25", lw=1.5, color="#8FA8C4"),
            )
            pie_annot.set_visible(False)
            bar_annot = None
            if bar_ax:
                bar_annot = bar_ax.annotate(
                    "", xy=(0, 0), xytext=(0, 16), textcoords="offset points",
                    ha="center", va="bottom", fontsize=11, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.5", fc="#F7FAFC", ec="#A9B8C9", lw=1.2, alpha=0.95),
                )
                bar_annot.set_visible(False)

            state = {"wedge": None, "bar": None}

            def reset():
                if state["wedge"] is not None:
                    i = state["wedge"]
                    wedges[i].set_facecolor(orig_fc[i])
                    wedges[i].set_alpha(orig_alpha[i] if orig_alpha[i] is not None else 0.8)
                    if hasattr(wedges[i], 'set_radius'):
                        try: wedges[i].set_radius(orig_r[i])
                        except: pass
                state["wedge"] = None
                if state["bar"] is not None and bars:
                    k = state["bar"]
                    g = bar_geoms[id(bars[k])]
                    try:
                        bars[k].set_x(g["x"]); bars[k].set_y(g["y"])
                        bars[k].set_width(g["w"]); bars[k].set_height(g["h"])
                        bars[k].set_linewidth(0); bars[k].set_alpha(0.8)
                    except: pass
                state["bar"] = None

            def on_move(event):
                try:
                    if event.inaxes == pie_ax:
                        for i, w in enumerate(wedges):
                            hit, _ = w.contains(event)
                            if hit:
                                reset(); state["wedge"] = i
                                w.set_alpha(0.95); w.set_linewidth(2); w.set_edgecolor("#8FA8C4")
                                if hasattr(w, 'set_radius'):
                                    try: w.set_radius(orig_r[i] * 1.03)
                                    except: pass
                                theta = (w.theta1 + w.theta2) / 2
                                r = orig_r[i]
                                x = r * 0.7 * np.cos(np.radians(theta))
                                y = r * 0.7 * np.sin(np.radians(theta))
                                pie_annot.xy = (x, y)
                                lbl = labels[i] if i < len(labels) else f"类型{i+1}"
                                val = values[i] if i < len(values) else 0
                                pct = val / total * 100
                                pie_annot.set_text(f"{lbl}\n数量: {val} 个\n占比: {pct:.1f}%")
                                pie_annot.set_visible(True)
                                if bar_annot: bar_annot.set_visible(False)
                                fig.canvas.draw_idle(); return
                        reset(); pie_annot.set_visible(False)
                        if bar_annot: bar_annot.set_visible(False)
                        fig.canvas.draw_idle(); return

                    if bar_ax and event.inaxes == bar_ax:
                        for i, b in enumerate(bars):
                            hit, _ = b.contains(event)
                            if hit:
                                reset(); state["bar"] = i
                                g = bar_geoms[id(b)]
                                scale = 1.03
                                b.set_x(g["x"] - g["w"] * (scale-1)/2)
                                b.set_width(g["w"] * scale)
                                b.set_height(g["h"] * scale)
                                b.set_linewidth(2); b.set_edgecolor("#6E8FB3"); b.set_alpha(1.0)
                                lbl = labels[i] if i < len(labels) else f"类型{i+1}"
                                val = values[i] if i < len(values) else 0
                                pct = val / total * 100
                                bar_annot.xy = (b.get_x() + b.get_width()/2, b.get_height())
                                bar_annot.set_text(f"{lbl}\n数量: {val} 个\n占比: {pct:.1f}%")
                                bar_annot.set_visible(True)
                                pie_annot.set_visible(False)
                                fig.canvas.draw_idle(); return
                        reset(); pie_annot.set_visible(False)
                        if bar_annot: bar_annot.set_visible(False)
                        fig.canvas.draw_idle(); return

                    reset(); pie_annot.set_visible(False)
                    if bar_annot: bar_annot.set_visible(False)
                    fig.canvas.draw_idle()
                except Exception:
                    pass

            fig.canvas.mpl_connect("motion_notify_event", on_move)
            fig.canvas.mpl_connect("axes_leave_event", lambda e: (
                reset(), pie_annot.set_visible(False),
                bar_annot.set_visible(False) if bar_annot else None,
                fig.canvas.draw_idle()
            ))
        except Exception as exc:
            logger.error(f"attach pie hover failed: {exc}")

    def _export_chart(self):
        if not self._current_chart_figure:
            QMessageBox.warning(self, "提示", "当前没有可导出的图表，请先刷新图表")
            return
        _, label, _ = self._get_current_chart_key()
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 添加筛选条件标注
        filter_suffix = ''
        active_filters = self._get_active_filter_items()
        if active_filters:
            filter_suffix = f"_【{','.join(active_filters[:2])}{'...' if len(active_filters) > 2 else ''}】"
        
        group, _, _ = self._get_current_chart_key()
        default_name = f"{group}_{label}{filter_suffix}_{ts}.png"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出图表", default_name,
            "PNG 图片 (*.png);;SVG 矢量图 (*.svg);;PDF 文档 (*.pdf);;所有文件 (*.*)"
        )
        if path:
            try:
                saved_paths = self._save_chart_outputs(self._current_chart_figure, Path(path))
                summary = '\n'.join(str(p) for p in saved_paths[:6])
                if len(saved_paths) > 6:
                    summary += f"\n... 共 {len(saved_paths)} 个文件"
                QMessageBox.information(self, "成功", f"图表导出完成，共生成 {len(saved_paths)} 个文件：\n{summary}")
            except Exception as exc:
                QMessageBox.critical(self, "错误", f"导出图表失败:\n{exc}")

    def _sanitize_export_name(self, text: str) -> str:
        cleaned = re.sub(r'[<>:"/\\|?*]+', '_', str(text or '').strip())
        cleaned = re.sub(r'\s+', '_', cleaned)
        cleaned = re.sub(r'_+', '_', cleaned)
        return cleaned.strip('._ ') or 'chart'

    def _iter_chart_axes_for_export(self, fig) -> list[tuple[int, object, str]]:
        axes = [ax for ax in (getattr(fig, 'axes', []) or []) if getattr(ax, 'has_data', lambda: False)()]
        exported = []
        for idx, ax in enumerate(axes, 1):
            name_parts = []
            for part in ((ax.get_title() or '').strip(), (ax.get_ylabel() or '').strip(), (ax.get_xlabel() or '').strip()):
                if part and part not in name_parts:
                    name_parts.append(part)
            title = "_".join(name_parts[:2]) if name_parts else f"subplot_{idx:02d}"
            exported.append((idx, ax, self._sanitize_export_name(title)))
        return exported
    def _create_batch_lifetime_single_figure(self, spec: dict, chart_title: str):
        import numpy as _np

        batch_labels = list(spec.get('batch_labels', []) or [])
        series_x = list(spec.get('series_x', []) or [])
        series_y = list(spec.get('series_y', []) or [])
        series_key = str(spec.get('series_key', '') or '').strip()

        fig = Figure(figsize=(10.8, 6.2), facecolor='white')
        FigureCanvasAgg(fig)
        ax = fig.add_subplot(111)
        color = self.batch_lifetime_analyzer._get_series_color(series_key)
        x_all = _np.arange(len(batch_labels))
        max_mean = max(series_y) if series_y else 0.0
        upper_margin = max(1.0, max_mean * 0.12)

        ax.plot(
            series_x,
            series_y,
            linewidth=2.1,
            marker='o',
            markersize=5.8,
            color=color,
        )
        ax.scatter(series_x, series_y, color=color, s=38, zorder=3)
        for point_x, point_y in zip(series_x, series_y):
            ax.text(
                point_x,
                point_y + max(0.22, point_y * 0.03),
                f"{point_y:.0f}根",
                ha='center',
                va='bottom',
                fontsize=8.0,
                color=color,
            )

        ax.set_facecolor('white')
        ax.set_ylim(0, max_mean + upper_margin)
        ax.tick_params(axis='both', labelsize=9.3)
        ax.grid(True, axis='y', alpha=0.24, linestyle='--', color='#C7D3DF')
        ax.set_axisbelow(True)
        for spine in ax.spines.values():
            spine.set_color('#D7E0E9')
            spine.set_linewidth(1.0)
        ax.set_ylabel('平均寿命（检测管数）/ 根', fontsize=11, fontweight='bold', labelpad=10)
        ax.set_title(series_key, fontsize=11.0, fontweight='bold', pad=10, loc='center', color='#314456', y=1.01)
        ax.set_xticks(x_all)
        ax.set_xticklabels(batch_labels, rotation=38, ha='right')
        ax.set_xlabel('生产批次（年-月）', fontsize=10.0, fontweight='bold', labelpad=4)
        fig.suptitle(chart_title, fontsize=15, fontweight='bold', y=0.965)
        fig.subplots_adjust(left=0.10, right=0.98, top=0.86, bottom=0.22)
        return fig

    def _create_probe_type_model_export_figure(
        self,
        *,
        probe_type: str,
        model: str,
        items: list[dict],
        chart_title: str,
        ylabel: str,
        value_formatter=None,
    ):
        formatter = value_formatter or (lambda value: f"{value:.1f}")
        fig = Figure(figsize=(10.8, 6.2), facecolor='white')
        FigureCanvasAgg(fig)
        ax = fig.add_subplot(111)

        labels = [str(item.get('probe_sn', '') or '') for item in items]
        values = [float(item.get('value', 0.0) or 0.0) for item in items]
        x = list(range(len(items)))
        color = self.visualizer._get_series_color(probe_type, model)
        linestyle = '-' if len(items) > 1 else 'None'
        probe_type_color = self.visualizer._get_probe_type_base_color(probe_type)

        ax.set_facecolor('#FCFDFE')
        ax.plot(
            x,
            values,
            color=color,
            linewidth=1.9,
            marker='o',
            markersize=4.1,
            markeredgecolor='#1F1F1F',
            markeredgewidth=0.3,
            markerfacecolor=color,
            linestyle=linestyle,
            zorder=3,
        )

        title_y = 1.015
        ax.text(
            0.485,
            title_y,
            probe_type,
            transform=ax.transAxes,
            ha='right',
            va='bottom',
            fontsize=10.0,
            fontweight='bold',
            color=probe_type_color,
        )
        ax.text(
            0.5,
            title_y,
            '  /  ',
            transform=ax.transAxes,
            ha='center',
            va='bottom',
            fontsize=10.0,
            fontweight='bold',
            color='#6B7280',
        )
        ax.text(
            0.515,
            title_y,
            model,
            transform=ax.transAxes,
            ha='left',
            va='bottom',
            fontsize=10.0,
            fontweight='bold',
            color=color,
        )
        ax.set_ylabel(ylabel, fontsize=9.5, color='#243447', labelpad=3)
        ax.set_xticks(x)
        if len(labels) <= 8:
            rotation = 32
            label_fontsize = 8.4
        elif len(labels) <= 16:
            rotation = 40
            label_fontsize = 7.3
        elif len(labels) <= 26:
            rotation = 46
            label_fontsize = 6.9
        else:
            rotation = 50
            label_fontsize = 6.4
        ax.set_xticklabels(labels, rotation=rotation, ha='right', fontsize=label_fontsize)
        ax.tick_params(axis='x', colors='#4C5A67', pad=4)
        ax.tick_params(axis='y', labelsize=9, colors='#4C5A67')
        ax.grid(axis='y', color='#D9E2EC', alpha=0.8, linestyle='--', linewidth=0.8)
        ax.set_axisbelow(True)
        ax.set_xlabel('探头编号', fontsize=9.5, color='#243447', labelpad=4, loc='center')
        max_val = max(values) if values else 0.0
        ax.set_ylim(0, max_val * 1.16 if max_val > 0 else 1)
        if all(float(value).is_integer() for value in values):
            ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax.margins(x=0.05)
        for label in ax.get_xticklabels():
            label.set_clip_on(False)
        for spine in ax.spines.values():
            spine.set_color('#D5DEE8')
            spine.set_linewidth(0.9)

        y_offset = max_val * 0.025 if max_val > 0 else 0.15
        for idx, (x_pos, value) in enumerate(zip(x, values)):
            offset = y_offset if idx % 2 == 0 else y_offset * 1.9
            ax.text(
                x_pos,
                value + offset,
                formatter(value),
                ha='center',
                va='bottom',
                fontsize=7.1,
                color=color,
            )

        fig.suptitle(chart_title, fontsize=15, fontweight='bold', color='#1F2D3D', x=0.5, ha='center', y=0.965)
        fig.subplots_adjust(left=0.09, right=0.985, top=0.84, bottom=0.21)
        return fig

    def _get_chart_display_title(self, group: str, label: str) -> str:
        if group == "按探头编号统计":
            return f"{group}（按探头类型/型号分图） - {label}"
        return f"{group} - {label}"

    @staticmethod
    def _save_figure_file(
        fig,
        target_path: Path,
        *,
        dpi: int = 150,
        tight: bool = False,
        pad_inches: float = 0.08,
    ):
        save_kwargs = {
            'dpi': dpi,
            'facecolor': getattr(fig, 'get_facecolor', lambda: 'white')(),
        }
        if tight:
            save_kwargs['bbox_inches'] = 'tight'
            save_kwargs['pad_inches'] = pad_inches
        fig.savefig(str(target_path), **save_kwargs)

    @staticmethod
    def _build_export_fig_kwargs() -> dict:
        return {
            'figure_width': 10.8,
            'figure_height': 6.2,
            'batch_figure_width': 10.8,
            'batch_figure_height': 6.2,
        }

    def _save_chart_outputs(self, fig, target_path: Path) -> list[Path]:
        target_path = Path(target_path)
        suffix = target_path.suffix.lower() or '.png'
        saved_paths: list[Path] = []

        if not self._ensure_output_path_allowed(target_path):
            return saved_paths

        self._save_figure_file(fig, target_path, tight=False)
        saved_paths.append(target_path)

        if suffix not in {'.png', '.svg', '.pdf'}:
            return saved_paths

        export_mode = getattr(fig, '_pies_export_mode', '')
        if export_mode == 'probe_type_model':
            specs = self.visualizer.get_probe_type_model_export_specs(fig)
            chart_title = str(getattr(fig, '_pies_export_chart_title', '') or '')
            ylabel = str(getattr(fig, '_pies_export_ylabel', '') or '')
            formatter = getattr(fig, '_pies_export_value_formatter', None)
            base_dir = target_path.parent
            stem = self._sanitize_export_name(target_path.stem)
            for idx, spec in enumerate(specs, 1):
                probe_type = str(spec.get('probe_type', '') or '').strip()
                model = str(spec.get('model', '') or '').strip()
                items = [dict(item) for item in (spec.get('items', []) or [])]
                axis_name = self._sanitize_export_name(
                    f"探头类型_{probe_type}_探头型号_{model}_{ylabel}"
                )
                sub_path = base_dir / f"{stem}_子图{idx:02d}_{axis_name}{suffix}"
                sub_fig = self._create_probe_type_model_export_figure(
                    probe_type=probe_type,
                    model=model,
                    items=items,
                    chart_title=chart_title,
                    ylabel=ylabel,
                    value_formatter=formatter,
                )
                try:
                    self._save_figure_file(sub_fig, sub_path, tight=False)
                    saved_paths.append(sub_path)
                finally:
                    plt.close(sub_fig)
            return saved_paths

        if export_mode == 'batch_lifetime':
            specs = self.batch_lifetime_analyzer.get_batch_lifetime_export_specs(fig)
            chart_title = str(getattr(fig, '_pies_export_chart_title', '') or '')
            base_dir = target_path.parent
            stem = self._sanitize_export_name(target_path.stem)
            for idx, spec in enumerate(specs, 1):
                series_key = self._sanitize_export_name(
                    f"批次平均寿命_{str(spec.get('series_key', '') or 'batch_lifetime')}"
                )
                sub_path = base_dir / f"{stem}_子图{idx:02d}_{series_key}{suffix}"
                sub_fig = self._create_batch_lifetime_single_figure(spec, chart_title)
                try:
                    self._save_figure_file(sub_fig, sub_path, tight=False)
                    saved_paths.append(sub_path)
                finally:
                    plt.close(sub_fig)
            return saved_paths

        # 旧图类型暂不自动拆子图，避免把同一整图重复保存成多个“子图”误导用户。
        return saved_paths

    # ═══════════════════════════════════════════════════════════
    #  筛选逻辑
    # ═══════════════════════════════════════════════════════════
    def _filter_value_is_active(self, value) -> bool:
        if value is None or value == '全部':
            return False
        if isinstance(value, (list, tuple, set)):
            return bool([i for i in value if i not in (None, '', '全部')])
        return str(value).strip() != ''

    def _normalize_filter_value(self, value):
        if not self._filter_value_is_active(value):
            return None
        if isinstance(value, set):
            value = sorted(str(i) for i in value if i not in (None, '', '全部'))
        if isinstance(value, (list, tuple)):
            items = [str(i).strip() for i in value if i not in (None, '', '全部')]
            if not items:
                return None
            if len(items) == 1:
                return items[0]
            return tuple(dict.fromkeys(items))
        return str(value).strip()

    def _filter_value_to_list(self, value) -> list:
        n = self._normalize_filter_value(value)
        if n is None:
            return []
        if isinstance(n, tuple):
            return list(n)
        return [n]

    def _format_filter_value_short_text(self, value) -> str:
        items = self._filter_value_to_list(value)
        if not items:
            return '全部'
        if len(items) == 1:
            return items[0]
        if len(items) == 2:
            return '、'.join(items)
        return f'已选{len(items)}项'

    def _get_active_filter_items(self) -> list:
        items = []
        for col, val in (self.summary_filter_values or {}).items():
            if self._filter_value_is_active(val):
                items.append(f"{col}={self._format_filter_value_short_text(val)}")
        if self._global_keyword_filter:
            items.append(f"关键字={self._global_keyword_filter}")
        return items

    def _get_available_filter_values(self, col_name: str) -> list:
        unique = set()
        for r in self.current_records:
            v = self._get_record_value(r, col_name)
            if v:
                unique.add(str(v))
        return sorted(unique)

    def _get_record_value(self, record, col_name: str) -> str:
        mapping = {
            '大修': lambda r: getattr(r, 'outage', '') or '',
            '蒸汽发生器编号': lambda r: getattr(r, 'sg_id', '') or '',
            '数据组': lambda r: r.data_group,
            '操作员': lambda r: r.operator,
            '探头类型': lambda r: getattr(r, 'probe_type_raw', None) or r.probe_type,
            '探头编码': lambda r: r.probe_sn,
            '探头型号': lambda r: r.model,
            '管道数量': lambda r: str(r.tube_number) if r.tube_number is not None else '',
            '累计管道数量': lambda r: str(
                self.current_statistics.get(r.stat_key).unique_tube_count
            ) if self.current_statistics.get(r.stat_key) else '',
            '开始时间': lambda r: r.start_time.strftime('%Y-%m-%d %H:%M:%S') if r.start_time else '',
            '结束时间': lambda r: r.end_time.strftime('%Y-%m-%d %H:%M:%S') if r.end_time else '',
        }
        fn = mapping.get(col_name)
        return fn(record) if fn else ''

    def _summary_record_matches_filters(self, record) -> bool:
        for col in TABLE_HEADER_FILTERABLE_COLUMNS:
            val = self.summary_filter_values.get(col)
            if self._filter_value_is_active(val):
                rv = self._get_record_value(record, col)
                if not self._matches_filter_value(rv, val, col):
                    return False
        return self._record_matches_keyword(record)

    def _matches_filter_value(self, record_value: str, filter_value, col_name: str) -> bool:
        items = self._filter_value_to_list(filter_value)
        if not items:
            return True
        partial_cols = {'探头编码', '探头型号', '数据组', '管道数量', '累计管道数量'}
        if col_name in partial_cols:
            return any(v.lower() in record_value.lower() for v in items)
        return any(record_value == v for v in items)

    def _record_matches_keyword(self, record) -> bool:
        kw = self._global_keyword_filter.strip().lower()
        if not kw:
            return True
        haystack = ' '.join(str(v) for v in [
            getattr(record, 'outage', ''), getattr(record, 'sg_id', ''),
            record.data_group, record.operator,
            getattr(record, 'probe_type_raw', None) or record.probe_type,
            record.probe_sn, record.model,
        ] if v).lower()
        return kw in haystack

    def _set_summary_filter(self, col_name: str, value):
        n = self._normalize_filter_value(value)
        self.summary_filter_values[col_name] = n
        self.filter_values[col_name] = n
        self._invalidate_filtered_cache()
        self._update_summary_table(self.current_statistics)
        self._schedule_chart_refresh()
        self._refresh_overview_metrics()
        self._refresh_warning_button_state()
        self._update_filter_summary_label()

    def _toggle_filter_value(self, col_name: str, value: str):
        current = self.summary_filter_values.get(col_name)
        selected = set(self._filter_value_to_list(current))
        if value in selected:
            selected.remove(value)
        else:
            selected.add(value)
        self._set_summary_filter(col_name, self._normalize_filter_value(sorted(selected)))

    def _clear_all_filters(self):
        for col in TABLE_HEADER_FILTERABLE_COLUMNS:
            self.filter_values[col] = None
            self.summary_filter_values[col] = None
        self._global_keyword_filter = ''
        self._invalidate_filtered_cache()
        self._update_summary_table(self.current_statistics)
        self._schedule_chart_refresh()
        self._refresh_overview_metrics()
        self._refresh_warning_button_state()
        self._update_filter_summary_label()

    def _rebuild_statistics_from_records(self, records: list) -> dict:
        if not records:
            return {}
        try:
            return self._build_lightweight_statistics(records)
        except Exception:
            try:
                tmp = ProbeAnalyzer()
                tmp.records = list(records)
                try:
                    return tmp.analyze(skip_deduplication=True, collect_debug_info=False)
                except TypeError:
                    return tmp.analyze()
            except Exception:
                return {}

    def _build_lightweight_statistics(self, records: list) -> dict:
        grouped: dict[str, dict[str, Any]] = {}
        invalid_time = datetime(1900, 1, 1, 0, 0, 0)

        for record in records:
            stat_key = record.stat_key
            bucket = grouped.get(stat_key)
            if bucket is None:
                probe_type = getattr(record, 'probe_type_raw', None) or getattr(record, 'probe_type', '')
                bucket = {
                    'sample': record,
                    'records': [],
                    'total_duration_minutes': 0.0,
                    'valid_count': 0,
                    'first_use': None,
                    'last_use': None,
                    'probe_type': probe_type,
                }
                grouped[stat_key] = bucket

            bucket['records'].append(record)

            start_time = getattr(record, 'start_time', None)
            end_time = getattr(record, 'end_time', None)
            if start_time is None or end_time is None:
                continue
            if start_time == invalid_time or end_time == invalid_time:
                continue
            if start_time.year < 2000 or end_time.year < 2000:
                continue
            duration_minutes = (end_time - start_time).total_seconds() / 60.0
            if duration_minutes <= 0:
                continue
            if getattr(record, 'tube_number', None) in (None, 0):
                continue

            bucket['total_duration_minutes'] += duration_minutes
            bucket['valid_count'] += 1

            first_use = bucket['first_use']
            if first_use is None or start_time < first_use:
                bucket['first_use'] = start_time

            last_use = bucket['last_use']
            if last_use is None or end_time > last_use:
                bucket['last_use'] = end_time

        statistics: dict[str, ProbeStatistics] = {}
        for stat_key, bucket in grouped.items():
            sample = bucket['sample']
            group_records = bucket['records']
            group_records.sort(
                key=lambda record: (
                    record.start_time if record.start_time is not None else datetime.max,
                    record.end_time if record.end_time is not None else datetime.max,
                    str(getattr(record, 'data_group', '') or ''),
                    str(getattr(record, 'operator', '') or ''),
                ),
            )

            first_use = bucket['first_use']
            last_use = bucket['last_use']
            if first_use is None and group_records:
                first_use = group_records[0].start_time
            if last_use is None and group_records:
                last_use = group_records[-1].end_time

            statistics[stat_key] = ProbeStatistics(
                probe_sn=sample.probe_sn,
                probe_type=bucket['probe_type'],
                model=_normalize_filter_model_name(sample.model),
                stat_key=stat_key,
                total_uses=int(bucket['valid_count'] or len(group_records)),
                total_duration_minutes=float(bucket['total_duration_minutes']),
                first_use_time=first_use,
                last_use_time=last_use,
                records=group_records,
                usage_sessions=[],
                reuse_details=[],
            )

        return statistics

    # ═══════════════════════════════════════════════════════════
    #  历史记录管理（完整版本，兼容旧版 JSON 格式 v2）
    # ═══════════════════════════════════════════════════════════

    # ── 序列化 ──
    def _serialize_probe_record(self, r: ProbeRecord) -> dict:
        return {
            'probe_sn': r.probe_sn, 'probe_type': r.probe_type,
            'probe_type_raw': getattr(r, 'probe_type_raw', ''),
            'model': r.model, 'tube_number': r.tube_number,
            'operator': r.operator, 'data_group': r.data_group,
            'outage': getattr(r, 'outage', '') or '',
            'sg_id':  getattr(r, 'sg_id', '') or '',
            'start_time': r.start_time.isoformat() if r.start_time else '',
            'end_time':   r.end_time.isoformat()   if r.end_time   else '',
            'warning_line_number': getattr(r, 'warning_line_number', '') or '',
        }

    def _deserialize_probe_record(self, item: dict) -> Optional[ProbeRecord]:
        try:
            def _dt(s):
                if not s: return None
                try: return datetime.fromisoformat(s)
                except: return None
            return ProbeRecord(
                probe_sn=str(item.get('probe_sn', '') or ''),
                probe_type=str(item.get('probe_type', '') or ''),
                probe_type_raw=str(item.get('probe_type_raw', '') or ''),
                start_time=_dt(item.get('start_time')),
                end_time=_dt(item.get('end_time')),
                tube_number=int(item.get('tube_number', 0) or 0),
                operator=str(item.get('operator', '') or ''),
                data_group=str(item.get('data_group', '') or ''),
                model=str(item.get('model', '') or ''),
                outage=str(item.get('outage', '') or ''),
                sg_id=str(item.get('sg_id', '') or ''),
            )
        except Exception as exc:
            logger.warning(f"历史记录反序列化失败: {exc}")
            return None

    def _serialize_warning_groups(self, warning_groups: dict) -> list:
        serialized = []
        for _, group in (warning_groups or {}).items():
            item = {
                'probe_sn': group.get('probe_sn', ''),
                'operator': group.get('operator', ''),
                'line_number': group.get('line_number', ''),
                'warning_types': sorted(group.get('warning_types', set()) or set()),
                'details': list(group.get('details', []) or []),
            }
            for field in ('outage', 'sg_id', 'data_group', 'probe_type', 'model',
                          'tube_number', 'start_time', 'end_time'):
                item[field] = str(group.get(field, '') or '')
            serialized.append(item)
        return serialized

    def _deserialize_warning_groups(self, items: list) -> dict:
        restored = {}
        for item in (items or []):
            probe_sn   = str(item.get('probe_sn', '') or '')
            operator   = str(item.get('operator', '') or '')
            line_number = str(item.get('line_number', '') or '')
            key = (probe_sn, operator, line_number)
            restored[key] = {
                'probe_sn': probe_sn, 'operator': operator, 'line_number': line_number,
                'warning_types': set(item.get('warning_types', []) or []),
                'details': list(item.get('details', []) or []),
                'warnings': list(item.get('warning_types', []) or []),
            }
            for field in ('outage', 'sg_id', 'data_group', 'probe_type', 'model',
                          'tube_number', 'start_time', 'end_time'):
                restored[key][field] = str(item.get(field, '') or '')
        return restored

    # ── 历史追加 ──
    def _append_to_history_records(self, records, warning_groups=None, warning_messages=None, error_records=None):
        merged, added = self._merge_unique_records(self.history_records, records)
        self.history_records    = merged
        self._history_statistics_dirty = True
        self._invalidate_filtered_cache()
        if self._current_scope == "history":
            self._ensure_history_statistics()
        # 合并警告
        if warning_messages:
            seen = set(self.history_warning_messages)
            for m in warning_messages:
                if m not in seen:
                    seen.add(m)
                    self.history_warning_messages.append(m)
        if warning_groups:
            for k, v in warning_groups.items():
                if k not in self.history_warning_by_probe:
                    self.history_warning_by_probe[k] = dict(v)
                else:
                    existing = self.history_warning_by_probe[k]
                    existing.setdefault('warnings', [])
                    for w in (v.get('warnings') or []):
                        if w not in existing['warnings']:
                            existing['warnings'].append(w)

        if error_records:
            seen_errors = {
                (
                    self._normalize_warning_context_value(e.get('探头编号', '')),
                    self._normalize_warning_context_value(e.get('操作员', '')),
                    self._normalize_warning_context_value(e.get('行号', '')),
                    self._normalize_warning_context_value(e.get('错误类型', '') or e.get('错误信息', '')),
                )
                for e in (self.history_error_records or [])
                if isinstance(e, dict)
            }
            for err in error_records:
                if not isinstance(err, dict):
                    continue
                err_key = (
                    self._normalize_warning_context_value(err.get('探头编号', '')),
                    self._normalize_warning_context_value(err.get('操作员', '')),
                    self._normalize_warning_context_value(err.get('行号', '')),
                    self._normalize_warning_context_value(err.get('错误类型', '') or err.get('错误信息', '')),
                )
                if err_key in seen_errors:
                    continue
                seen_errors.add(err_key)
                self.history_error_records.append(self._json_safe(dict(err)))

        # 记录本次导入摘要
        session_dedup = self.session_deduplication_info or {}
        self.history_import_summaries.append({
            'imported_at': datetime.now().isoformat(timespec='seconds'),
            'original_count': session_dedup.get('original_count', len(records)),
            'unique_count':   session_dedup.get('unique_count', len(records)),
            'removed_count':  session_dedup.get('removed_count', 0),
            'removed_records': [],
            'added_to_history_count': added,
        })
        self._rebuild_history_dedup_info()
        self._save_history_store()
        return added

    def _merge_unique_records(self, existing: list, new_records: list) -> Tuple[list, int]:
        def _key(r):
            def _n(v):
                if v is None: return ''
                if hasattr(v, 'isoformat'):
                    try: return v.isoformat(sep=' ', timespec='seconds')
                    except: return v.isoformat()
                return str(v).strip()
            return (
                _n(getattr(r, 'outage', '')), _n(getattr(r, 'sg_id', '')),
                _n(r.data_group), _n(r.operator),
                _n(getattr(r, 'probe_type_raw', None) or r.probe_type),
                _n(r.probe_sn), _n(r.model),
                r.tube_number if r.tube_number is not None else '',
                _n(r.start_time), _n(r.end_time),
            )
        seen = {_key(r) for r in existing}
        added = 0; result = list(existing)
        for r in new_records:
            k = _key(r)
            if k not in seen:
                seen.add(k); result.append(r); added += 1
        return result, added

    def _rebuild_history_dedup_info(self):
        total_orig = sum(s.get('original_count', 0) for s in self.history_import_summaries)
        removed    = max(0, total_orig - len(self.history_records))
        self.history_deduplication_info = {
            'original_count': total_orig,
            'unique_count':   len(self.history_records),
            'removed_count':  removed,
            'removed_records': self.history_deduplication_info.get('removed_records', []),
        }

    # ── 持久化（兼容旧版 v2 格式）──
    def _save_history_store(self):
        try:
            session_dedup = self.session_deduplication_info or {}
            history_dedup = self.history_deduplication_info or {}
            history_summaries = []
            for s in (self.history_import_summaries or []):
                history_summaries.append({
                    'imported_at': s.get('imported_at', ''),
                    'original_count': s.get('original_count', 0),
                    'unique_count':   s.get('unique_count', 0),
                    'removed_count':  s.get('removed_count', 0),
                    'added_to_history_count': s.get('added_to_history_count', 0),
                    'removed_records': [
                        self._serialize_probe_record(r)
                        for r in (s.get('removed_records', []) or [])
                        if hasattr(r, 'probe_sn')
                    ],
                })
            payload = {
                'version': 2,
                'updated_at': datetime.now().isoformat(timespec='seconds'),
                'records': [self._serialize_probe_record(r) for r in self.history_records],
                'error_records': self._json_safe(list(self.history_error_records or [])),
                'warning_messages': list(self.history_warning_messages or []),
                'warning_groups': self._serialize_warning_groups(self.history_warning_by_probe),
                'current_records': [self._serialize_probe_record(r) for r in self.session_records],
                'current_error_records': self._json_safe(list(self.session_error_records or [])),
                'current_warning_messages': list(self._warning_messages or []),
                'current_warning_groups': self._serialize_warning_groups(self._warning_by_probe),
                'session_deduplication_info': {
                    'original_count': session_dedup.get('original_count', 0),
                    'unique_count':   session_dedup.get('unique_count', 0),
                    'removed_count':  session_dedup.get('removed_count', 0),
                    'removed_records': [
                        self._serialize_probe_record(r)
                        for r in (session_dedup.get('removed_records', []) or [])
                        if hasattr(r, 'probe_sn')
                    ],
                },
                'history_deduplication_info': {
                    'original_count': history_dedup.get('original_count', 0),
                    'unique_count':   history_dedup.get('unique_count', 0),
                    'removed_count':  history_dedup.get('removed_count', 0),
                    'removed_records': [
                        self._serialize_probe_record(r)
                        for r in (history_dedup.get('removed_records', []) or [])
                        if hasattr(r, 'probe_sn')
                    ],
                },
                'history_import_summaries': history_summaries,
            }
            self.history_store_path.parent.mkdir(parents=True, exist_ok=True)
            self.history_store_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )
        except Exception as exc:
            logger.warning(f"保存历史记录失败: {exc}")

    def _load_history_store(self):
        """加载历史记录，兼容旧版 v1/v2 JSON 格式。"""
        load_path = self.history_store_path
        if not load_path.exists() and LEGACY_HISTORY_STORE_PATH.exists():
            load_path = LEGACY_HISTORY_STORE_PATH
        if not load_path.exists():
            return
        try:
            payload = json.loads(load_path.read_text(encoding='utf-8'))
            version = payload.get('version', 1)

            # 加载历史记录
            loaded_history = [self._deserialize_probe_record(d) for d in payload.get('records', [])]
            self.history_records = [r for r in loaded_history if r]
            self.history_statistics = {}
            self._history_statistics_dirty = bool(self.history_records)
            self.history_error_records = list(payload.get('error_records', []) or [])

            # 加载当前批次记录（v2 格式）
            loaded_session = [self._deserialize_probe_record(d) for d in payload.get('current_records', [])]
            self.session_records = [r for r in loaded_session if r]
            self.session_statistics = self._rebuild_statistics_from_records(self.session_records)

            # 加载警告
            self.history_warning_messages = list(payload.get('warning_messages', []) or [])
            self.history_warning_by_probe = self._deserialize_warning_groups(payload.get('warning_groups', []))
            self.session_error_records = list(payload.get('current_error_records', []) or [])
            self._warning_messages = list(payload.get('current_warning_messages', []) or [])
            self._warning_by_probe = self._deserialize_warning_groups(payload.get('current_warning_groups', []))

            # 加载导入摘要
            self.history_import_summaries = []
            for item in (payload.get('history_import_summaries', []) or []):
                removed_records = [
                    self._deserialize_probe_record(rd)
                    for rd in (item.get('removed_records', []) or [])
                ]
                self.history_import_summaries.append({
                    'imported_at': item.get('imported_at', ''),
                    'original_count': int(item.get('original_count', 0) or 0),
                    'unique_count':   int(item.get('unique_count', 0) or 0),
                    'removed_count':  int(item.get('removed_count', 0) or 0),
                    'added_to_history_count': int(item.get('added_to_history_count', 0) or 0),
                    'removed_records': [r for r in removed_records if r],
                })

            # 加载去重信息
            if version >= 2:
                hd = payload.get('history_deduplication_info', {}) or {}
                hd_removed = [self._deserialize_probe_record(d) for d in (hd.get('removed_records', []) or [])]
                self.history_deduplication_info = {
                    'original_count': hd.get('original_count', len(self.history_records)),
                    'unique_count':   hd.get('unique_count', len(self.history_records)),
                    'removed_count':  hd.get('removed_count', 0),
                    'removed_records': [r for r in hd_removed if r],
                }
                sd = payload.get('session_deduplication_info', {}) or {}
                sd_removed = [self._deserialize_probe_record(d) for d in (sd.get('removed_records', []) or [])]
                self.session_deduplication_info = {
                    'original_count': sd.get('original_count', len(self.session_records)),
                    'unique_count':   sd.get('unique_count', len(self.session_records)),
                    'removed_count':  sd.get('removed_count', 0),
                    'removed_records': [r for r in sd_removed if r],
                }
            else:
                self.history_deduplication_info = {
                    'original_count': len(self.history_records),
                    'unique_count':   len(self.history_records),
                    'removed_count':  0, 'removed_records': [],
                }
                self.session_deduplication_info = {
                    'original_count': len(self.session_records),
                    'unique_count':   len(self.session_records),
                    'removed_count':  0, 'removed_records': [],
                }

        except Exception as exc:
            logger.warning(f"加载历史记录失败: {exc}")
            self.session_records = []; self.session_statistics = {}
            self.history_records = []; self.history_statistics = {}
            self._history_statistics_dirty = False
            self._warning_messages = []; self._warning_by_probe = {}
            self.history_warning_messages = []; self.history_warning_by_probe = {}
            self.session_error_records = []; self.history_error_records = []
            self.history_import_summaries = []

    def _clear_history_records(self):
        if not (self.session_records or self.history_records or self.current_records):
            QMessageBox.information(self, "提示", "当前没有可清空的数据")
            return
        if QMessageBox.question(
            self, "确认",
            "确定要清空当前批次和本地累计数据吗？此操作不会影响原始 Excel 文件。",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return
        self.session_records = []; self.session_statistics = {}
        self.current_records = []; self.current_statistics = {}
        self._warning_messages.clear(); self._warning_by_probe.clear()
        self.session_error_records = []
        self.history_records = []; self.history_statistics = {}
        self._history_statistics_dirty = False
        self.history_warning_messages = []; self.history_warning_by_probe = {}
        self.history_error_records = []
        self.history_import_summaries = []
        self.history_deduplication_info = {}
        self.session_deduplication_info = {}
        self._invalidate_filtered_cache()
        self._save_history_store()
        self._current_scope = "current"
        self._apply_active_dataset(refresh_ui=True)
        self._refresh_warning_button_state()
        self._set_status("数据已清空")

    # ═══════════════════════════════════════════════════════════
    #  警告
    # ═══════════════════════════════════════════════════════════
    def _get_scope_warning_snapshot(self):
        if self._current_scope == 'history' and self.history_warning_by_probe:
            groups = {k: dict(v) for k, v in (self.history_warning_by_probe or {}).items()}
            self._enrich_warning_groups(groups, include_error_records=False)
            groups = self._filter_display_warning_groups(groups)
            return self._current_scope_label(), groups, list(self.history_warning_messages or [])
        groups = {k: dict(v) for k, v in (self._warning_by_probe or {}).items()}
        self._enrich_warning_groups(groups, include_error_records=True)
        groups = self._filter_display_warning_groups(groups)
        return self._current_scope_label(), groups, list(self._warning_messages or [])

    def _get_filtered_warning_snapshot(self):
        cache_key = ('warnings',) + self._make_filter_cache_key()
        cached = self._filtered_warning_cache.get(cache_key)
        if cached is not None:
            return cached

        scope, groups, messages = self._get_scope_warning_snapshot()
        active = self._get_active_filter_items()
        if not active:
            result = (scope, groups, messages)
            self._filtered_warning_cache[cache_key] = result
            return result

        visible_records, _ = self._get_filtered_records_and_stats()
        filtered = {}
        for group_key, group in (groups or {}).items():
            if self._warning_group_matches_active_filters(group):
                filtered[group_key] = group
                continue
            if any(self._warning_group_matches_record(group, record) for record in visible_records):
                filtered[group_key] = group

        filtered_messages = []
        if filtered:
            for group in filtered.values():
                for detail in group.get('details', []) or []:
                    msg = detail.get('message') or detail.get('raw_text') or ''
                    if msg:
                        filtered_messages.append(msg)
        result = (scope, filtered, filtered_messages)
        self._filtered_warning_cache[cache_key] = result
        return result

    def _warning_merge_key(self, probe_sn: str, operator: str, line_number: str = ''):
        return (
            self._normalize_warning_context_value(probe_sn) or '未知探头',
            self._normalize_warning_context_value(operator),
            self._normalize_warning_context_value(line_number),
        )

    def _find_warning_group_key(self, context: dict, warning_type: str = ''):
        for key, group in (self._warning_by_probe or {}).items():
            if self._warning_groups_compatible(group, context, warning_type):
                return key
        return None

    def _warning_groups_compatible(self, group: dict, context: dict, warning_type: str = '') -> bool:
        if not group or not context:
            return False
        group_probe = self._normalize_warning_context_value(group.get('probe_sn', ''))
        ctx_probe = self._normalize_warning_context_value(context.get('probe_sn', ''))
        if group_probe and group_probe != '未知探头' and ctx_probe and group_probe != ctx_probe:
            return False
        group_line = self._normalize_warning_context_value(group.get('line_number', ''))
        ctx_line = self._normalize_warning_context_value(context.get('line_number', ''))
        if group_line or ctx_line:
            return bool(group_line and ctx_line and group_line == ctx_line)
        for field in ('operator', 'data_group', 'outage', 'sg_id'):
            group_value = self._normalize_warning_context_value(group.get(field, ''))
            ctx_value = self._normalize_warning_context_value(context.get(field, ''))
            if group_value and ctx_value and group_value != ctx_value:
                return False
        if any(
            self._normalize_warning_context_value(group.get(field, ''))
            and self._normalize_warning_context_value(context.get(field, ''))
            for field in ('operator', 'data_group', 'outage', 'sg_id', 'line_number')
        ):
            return True
        existing_types = set(group.get('warning_types', set()) or set())
        return self._warning_type_matches(warning_type, existing_types)

    def _warning_group_matches_active_filters(self, group: dict) -> bool:
        active = [
            (col, val)
            for col, val in (self.summary_filter_values or {}).items()
            if self._filter_value_is_active(val)
        ]
        keyword = getattr(self, '_global_keyword_filter', '').strip().lower()
        if not active:
            if not keyword:
                return True

        column_to_field = {
            '大修': 'outage',
            '蒸汽发生器编号': 'sg_id',
            '数据组': 'data_group',
            '操作员': 'operator',
            '探头类型': 'probe_type',
            '探头编码': 'probe_sn',
            '探头型号': 'model',
            '管道数量': 'tube_number',
            '开始时间': 'start_time',
            '结束时间': 'end_time',
        }
        for column, filter_value in active:
            field = column_to_field.get(column)
            if not field:
                continue
            group_value = self._normalize_warning_context_value(group.get(field, ''))
            if not group_value or not self._matches_filter_value(group_value, filter_value, column):
                return False

        if keyword:
            values = [
                group.get('probe_sn', ''), group.get('model', ''), group.get('probe_type', ''),
                group.get('outage', ''), group.get('sg_id', ''), group.get('data_group', ''),
                group.get('operator', ''), group.get('line_number', ''),
            ]
            if keyword not in ' '.join(str(value) for value in values if value).lower():
                return False
        return True

    def _warning_group_matches_record(self, group: dict, record: ProbeRecord) -> bool:
        if not group or not record:
            return False
        probe = self._normalize_warning_context_value(group.get('probe_sn', ''))
        record_probe = self._normalize_warning_context_value(getattr(record, 'probe_sn', ''))
        if probe and probe != '未知探头' and probe != record_probe:
            return False

        context_fields = (
            ('operator', getattr(record, 'operator', '')),
            ('outage', getattr(record, 'outage', '')),
            ('sg_id', getattr(record, 'sg_id', '')),
            ('data_group', getattr(record, 'data_group', '')),
            ('probe_type', getattr(record, 'probe_type_raw', None) or getattr(record, 'probe_type', '')),
            ('model', getattr(record, 'model', '')),
            ('tube_number', getattr(record, 'tube_number', '')),
        )
        matched_context = 0
        for field, record_value in context_fields:
            group_value = self._normalize_warning_context_value(group.get(field, ''))
            if not group_value:
                continue
            if self._normalize_warning_context_value(record_value) != group_value:
                return False
            matched_context += 1

        start_value = self._normalize_warning_context_value(group.get('start_time', ''))
        end_value = self._normalize_warning_context_value(group.get('end_time', ''))
        if start_value and self._normalize_warning_context_value(getattr(record, 'start_time', '')) != start_value:
            return False
        if end_value and self._normalize_warning_context_value(getattr(record, 'end_time', '')) != end_value:
            return False
        return matched_context > 0 or bool(probe and probe == record_probe)

    def _record_warning_key(self, record: ProbeRecord):
        return (
            self._normalize_warning_context_value(getattr(record, 'probe_sn', '')),
            self._normalize_warning_context_value(getattr(record, 'operator', '')),
            self._normalize_warning_context_value(getattr(record, 'outage', '')),
            self._normalize_warning_context_value(getattr(record, 'sg_id', '')),
            self._normalize_warning_context_value(getattr(record, 'data_group', '')),
            self._normalize_warning_context_value(getattr(record, 'start_time', '')),
            self._normalize_warning_context_value(getattr(record, 'end_time', '')),
        )

    def _warning_key_from_context(self, context: dict):
        if not context:
            return None
        return (
            self._normalize_warning_context_value(context.get('probe_sn', '')),
            self._normalize_warning_context_value(context.get('operator', '')),
            self._normalize_warning_context_value(context.get('outage', '')),
            self._normalize_warning_context_value(context.get('sg_id', '')),
            self._normalize_warning_context_value(context.get('data_group', '')),
            self._normalize_warning_context_value(context.get('start_time', '')),
            self._normalize_warning_context_value(context.get('end_time', '')),
        )

    def _refresh_warning_button_state(self):
        """更新警告按钮状态：有警告时橙色实心，无警告时禁用"""
        _, groups, messages = self._get_filtered_warning_snapshot()
        warn_count = self._count_warning_items(groups, messages)
        has = warn_count > 0
        if has:
            self._btn_warnings.setEnabled(True)
            self._btn_warnings.setText(f"异常提醒({warn_count})")
            _apply_btn_style(self._btn_warnings, "btn_warning_active")
        else:
            self._btn_warnings.setEnabled(False)
            self._btn_warnings.setText("异常提醒")
            _apply_btn_style(self._btn_warnings, "btn_warning")

    def _count_warning_items(self, grouped_warnings: dict | None, warning_messages: list | None) -> int:
        if grouped_warnings:
            return len(self._build_warning_detail_rows(grouped_warnings))
        return len(warning_messages or [])

    def _display_warning_types(self, warning_group: dict) -> list:
        warnings_list = sorted(set((warning_group or {}).get('warning_types', set()) or (warning_group or {}).get('warnings', []) or []))
        if "开始时间和结束时间均为空" in warnings_list:
            warnings_list = [w for w in warnings_list if w not in ("开始时间为空", "结束时间为空")]
        if len(warnings_list) > 1 and "其他警告" in warnings_list:
            warnings_list = [w for w in warnings_list if w != "其他警告"]
        if "管道数量异常" in warnings_list:
            details = (warning_group or {}).get('details', []) or []
            tube_zero = any("管道数量=0" in str(d.get('raw_text', '')) or "管道数量为0" in str(d.get('raw_text', '')) for d in details if isinstance(d, dict))
            warnings_list = ["管道数量为0" if w == "管道数量异常" and tube_zero else w for w in warnings_list]
        return warnings_list

    def _filter_display_warning_groups(self, groups: dict) -> dict:
        filtered = {}
        for key, group in (groups or {}).items():
            if not self._is_real_warning_group(group):
                continue
            filtered[key] = group
        return filtered

    def _build_warning_detail_rows(self, warning_groups: dict) -> list[tuple[object, dict, str]]:
        rows = []
        seen = set()
        for key, group in sorted(
            (warning_groups or {}).items(),
            key=lambda x: (
                str((x[1] or {}).get('probe_sn', '')),
                str((x[1] or {}).get('line_number', '')),
                str((x[1] or {}).get('data_group', '')),
            ),
        ):
            if not isinstance(group, dict):
                continue
            self._force_complete_warning_group(group)
            warning_types = self._display_warning_types(group) or ['其他警告']
            for warning_type in warning_types:
                dedupe_key = (
                    self._normalize_warning_context_value(group.get('probe_sn', '')),
                    self._normalize_warning_context_value(group.get('operator', '')),
                    self._normalize_warning_context_value(group.get('outage', '')),
                    self._normalize_warning_context_value(group.get('sg_id', '')),
                    self._normalize_warning_context_value(group.get('data_group', '')),
                    self._normalize_warning_context_value(group.get('line_number', '')),
                    self._normalize_warning_context_value(group.get('tube_number', '')),
                    self._normalize_warning_context_value(group.get('start_time', '')),
                    self._normalize_warning_context_value(group.get('end_time', '')),
                    self._normalize_warning_context_value(warning_type),
                )
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                rows.append((key, group, warning_type))
        return rows

    def _is_real_warning_group(self, group: dict) -> bool:
        if not isinstance(group, dict):
            return False
        probe = self._normalize_warning_context_value(group.get('probe_sn', ''))
        warning_types = set(group.get('warning_types', set()) or group.get('warnings', []) or [])
        context_fields = ('line_number', 'data_group', 'model', 'outage', 'sg_id', 'probe_type')
        has_context = any(self._normalize_warning_context_value(group.get(field, '')) for field in context_fields)
        if probe in {'', '未知探头', '000000'} and not has_context:
            return False
        if warning_types == {'其他警告'} and not has_context:
            return False
        return True

    def _show_warning_details(self):
        """完整还原旧版的警告详情对话框（表格形式）。"""
        scope, active_by_probe, active_messages = self._get_filtered_warning_snapshot()

        if not active_by_probe and not active_messages:
            QMessageBox.information(self, "提示", "当前没有任何警告信息")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("警告详情")
        dlg.resize(1380, 720)
        self._configure_dialog_window(dlg, min_size=(1120, 600))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel("异常数据提醒")
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #C43D32; background: transparent;")
        lay.addWidget(t)

        warning_detail_rows = self._build_warning_detail_rows(active_by_probe) if active_by_probe else []
        if active_by_probe:
            unique_probes = len({
                self._normalize_warning_context_value(row_group.get('probe_sn', ''))
                for _, row_group, _ in warning_detail_rows
            } or {
                self._normalize_warning_context_value(group.get('probe_sn', ''))
                for group in active_by_probe.values()
            })
            total_w = len(warning_detail_rows)
            note_text = f"{scope}下共 {total_w} 条警告，涉及 {unique_probes} 个探头、{len(active_by_probe)} 个场景。"
        else:
            note_text = f"{scope}下共有 {len(active_messages)} 条原始警告信息。"
        note = QLabel(note_text)
        note.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
        note.setWordWrap(True)
        lay.addWidget(note)

        # 内容区
        if active_by_probe:
            cols = ('警告类型', '探头编号', '探头型号', '大修编号', 'SG编号',
                    '数据组', '操作员', '行号', '管道数量', '开始时间', '结束时间', '探头类型', '说明')
            col_widths = {
                '警告类型': 260, '探头编号': 150, '探头型号': 190, '大修编号': 105,
                'SG编号': 100, '数据组': 145, '操作员': 105, '行号': 80,
                '管道数量': 105, '开始时间': 175, '结束时间': 175, '探头类型': 120, '说明': 520,
            }
            tbl = QTableWidget(len(warning_detail_rows), len(cols))
            tbl.setHorizontalHeaderLabels(cols)
            tbl.setAlternatingRowColors(True)
            tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
            tbl.setFocusPolicy(Qt.NoFocus)
            tbl.setShowGrid(False)
            tbl.setFrameShape(QFrame.NoFrame)
            tbl.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
            tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            tbl.setWordWrap(False)
            tbl.setStyleSheet("""
                QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
                QHeaderView::section {
                    background: #FAFBFC;
                    color: #475467;
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid #E5E7EB;
                    font-weight: bold;
                }
                QTableWidget::item { padding: 6px; border: none; }
                QTableWidget::item:selected { background: #E9F2FF; color: #1F2937; }
            """)
            tbl.verticalHeader().setVisible(False)
            header = tbl.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)
            header.setStretchLastSection(False)
            header.setMinimumSectionSize(70)
            header.setDefaultSectionSize(120)
            header.setSectionsMovable(False)
            for i, col in enumerate(cols):
                tbl.setColumnWidth(i, col_widths.get(col, 120))
                header_item = tbl.horizontalHeaderItem(i)
                if header_item:
                    header_item.setToolTip(col)

            for row_idx, (key, probe_info, warning_desc) in enumerate(warning_detail_rows):
                probe_sn = probe_info.get('probe_sn', str(key) if isinstance(key, str) else str(key[0]))
                note_str = self._build_warning_note_text(probe_info, warning_desc)

                row_vals = [
                    self._warning_display_value(warning_desc),
                    self._warning_display_value(probe_sn),
                    self._warning_display_value(probe_info.get('model', '')),
                    self._warning_display_value(probe_info.get('outage', '')),
                    self._warning_display_value(probe_info.get('sg_id', '')),
                    self._warning_display_value(probe_info.get('data_group', '')),
                    self._warning_display_value(probe_info.get('operator', '')),
                    self._warning_display_value(probe_info.get('line_number', '')),
                    self._warning_display_value(probe_info.get('tube_number', '')),
                    self._warning_display_value(probe_info.get('start_time', '')),
                    self._warning_display_value(probe_info.get('end_time', '')),
                    self._warning_display_value(probe_info.get('probe_type', '')),
                    self._warning_display_value(note_str),
                ]
                bg = QColor("#F9FBFD") if row_idx % 2 == 0 else QColor("#FFFFFF")
                for col_idx, val in enumerate(row_vals):
                    item = QTableWidgetItem(str(val) if val else '\\')
                    item.setTextAlignment(Qt.AlignCenter if col_idx != len(cols) - 1 else Qt.AlignLeft | Qt.AlignVCenter)
                    item.setBackground(bg)
                    tbl.setItem(row_idx, col_idx, item)
            lay.addWidget(tbl, 1)
        else:
            editor = QTextEdit()
            editor.setReadOnly(True)
            editor.setFont(QFont("Microsoft YaHei", 10))
            editor.setPlainText('\n'.join(f"{i+1}. {m}" for i, m in enumerate(active_messages)))
            lay.addWidget(editor, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_close = _btn("关闭", "btn_primary")
        btn_close.setMinimumHeight(34)
        btn_close.setMinimumWidth(92)
        btn_close.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)
        self._present_dialog(dlg, key="warning_details")

    # ═══════════════════════════════════════════════════════════
    #  导出
    # ═══════════════════════════════════════════════════════════
    def _export_summary_table(self):
        import pandas as pd
        if not self.current_statistics:
            QMessageBox.warning(self, "提示", "没有可导出的数据，请先处理文件")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 添加筛选条件标注
        filter_suffix = ''
        active_filters = self._get_active_filter_items()
        if active_filters:
            filter_suffix = f"_【{';'.join(active_filters[:2])}{'...' if len(active_filters) > 2 else ''}】"
        
        path, _ = QFileDialog.getSaveFileName(
            self, "导出探头使用信息表格", f"探头使用信息表格{filter_suffix}_{ts}.xlsx",
            "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if not path:
            return
        if not self._ensure_output_path_allowed(Path(path)):
            return
        prog = ProgressDialog("导出表格进度", self)
        prog.show_and_paint("正在准备导出表格...", 0)
        try:
            prog.set_progress("正在整理导出数据...", 10)
            filtered, filtered_stats = self._get_filtered_records_and_stats() if self._get_active_filter_items() else (self.current_records, self.current_statistics)
            rows = self._build_summary_export_rows(filtered, filtered_stats)
            df = pd.DataFrame(rows)
            prog.set_progress("正在写入 Excel 文件...", 45)
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='探头使用信息')
                ws = writer.sheets['探头使用信息']
                from openpyxl.styles import Font, PatternFill, Alignment
                hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                hfont = Font(bold=True, color="FFFFFF", size=11)
                for cell in ws[1]:
                    cell.fill = hf; cell.font = hfont
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in ws.columns:
                    ml = max((len(str(c.value or '')) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 36)
                
                # 添加筛选条件说明
                if active_filters:
                    info_row = ws.max_row + 2
                    ws[f'A{info_row}'] = "筛选条件："
                    ws[f'A{info_row}'].font = Font(bold=True, size=10)
                    for idx, f in enumerate(active_filters):
                        ws[f'A{info_row + idx + 1}'] = f"• {f}"
                        ws[f'A{info_row + idx + 1}'].font = Font(size=9, color="666666")
            prog.set_progress("导出完成", 100)
            prog.accept()
            QMessageBox.information(self, "成功", f"已成功导出 {len(rows)} 条记录到:\n{path}")
            self._set_status(f"表格已导出 - {len(rows)} 条记录")
        except Exception as exc:
            try:
                prog.accept()
            except Exception:
                pass
            logger.error(f"导出表格失败: {exc}")
            QMessageBox.critical(self, "错误", f"导出表格失败:\n{exc}")

    def _export_error_records(self):
        if self._current_scope == 'history':
            QMessageBox.information(self, "提示", "历史累计模式不提供错误记录导出，请切换到当前批次后导出。")
            return
        error_records = list(self.session_error_records or [])
        if not error_records:
            QMessageBox.information(self, "提示", "没有错误记录需要导出")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path, _ = QFileDialog.getSaveFileName(
            self, "导出错误记录", f"error_records_{ts}.xlsx",
            "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if path:
            if not self._ensure_output_path_allowed(Path(path)):
                return
            prog = ProgressDialog("导出错误记录进度", self)
            prog.show_and_paint("正在准备导出错误记录...", 0)
            try:
                prog.set_progress("正在导出错误记录...", 35)
                import pandas as pd
                pd.DataFrame(error_records).to_excel(path, index=False)
                prog.set_progress("导出完成", 100)
                prog.accept()
                QMessageBox.information(self, "成功", f"错误记录已导出到:\n{path}")
            except Exception as exc:
                try:
                    prog.accept()
                except Exception:
                    pass
                QMessageBox.critical(self, "错误", f"导出失败:\n{exc}")

    def _save_all(self):
        import pandas as pd
        if not self.current_statistics or not self.current_records:
            QMessageBox.warning(self, "提示", "没有可保存的数据，请先处理文件")
            return
        save_dir = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if not save_dir:
            return
        if not self._ensure_output_path_allowed(Path(save_dir), title="禁止保存到数据源目录"):
            return

        prog = ProgressDialog("一键保存进度", self)
        prog.show_and_paint("正在准备一键保存...", 0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = Path(save_dir) / f"探头数据导出_{ts}"
        base_dir.mkdir(parents=True, exist_ok=True)
        saved = []
        rows = []

        # 保存表格
        try:
            prog.set_progress("正在保存表格数据...", 10)
            filtered, filtered_stats = self._get_filtered_records_and_stats() if self._get_active_filter_items() else (self.current_records, self.current_statistics)
            rows = self._build_summary_export_rows(filtered, filtered_stats)
            if rows:
                ep = base_dir / f"探头使用信息表格_{ts}.xlsx"
                pd.DataFrame(rows).to_excel(str(ep), index=False)
                saved.append(str(ep))
        except Exception as exc:
            logger.error(f"一键保存表格失败: {exc}")

        # 保存所有图表
        chart_count = 0
        all_charts = [(g, lbl, key) for g, opts in CHART_GROUPS.items() for lbl, key in opts]
        records, statistics, hidden = self._get_chart_source_data()
        export_fig_kwargs = self._build_export_fig_kwargs()
        for i, (group, label, chart_key) in enumerate(all_charts):
            pct = 20 + (i / max(len(all_charts), 1)) * 75
            prog.set_progress(f"正在保存图表: {label}...", pct)
            try:
                fig = self._create_chart_figure(chart_key, records, statistics, hidden, export_fig_kwargs)
                if fig:
                    safe_name = re.sub(r'[<>:"/\\|?*]', '_', f"{group}_{label}")
                    cp = base_dir / f"{safe_name}_{ts}.png"
                    exported_paths = self._save_chart_outputs(fig, cp)
                    saved.extend(str(p) for p in exported_paths)
                    chart_count += len(exported_paths)
                    plt.close(fig)
            except Exception as exc:
                logger.error(f"保存图表 {chart_key} 失败: {exc}")

        prog.set_progress("正在整理保存结果...", 98)
        prog.accept()
        QMessageBox.information(self, "成功",
            f"一键保存完成！\n保存目录: {base_dir}\n\n共保存:\n- 表格: {1 if rows else 0} 个\n- 图表: {chart_count} 个")
        self._set_status(f"一键保存完成 - {len(saved)} 个文件")

    # ═══════════════════════════════════════════════════════════
    #  对话框：历史详情
    # ═══════════════════════════════════════════════════════════
    def _show_history_details(self):
        if not self.history_records:
            QMessageBox.information(self, "历史详情", "当前还没有累计的数据。")
            return
        self._ensure_history_statistics()
        dlg = QDialog(self)
        dlg.setWindowTitle("历史记录详情")
        dlg.resize(920, 580)
        self._configure_dialog_window(dlg, min_size=(760, 460))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel("本地累计数据概览")
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #163F72; background: transparent;")
        lay.addWidget(t)
        s = QLabel("历史数据仅保存在本机，用于跨批次分析，不会回写原始文件。")
        s.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
        s.setWordWrap(True)
        lay.addWidget(s)

        hist_count  = len(self.history_records)
        hist_probes = len(self.history_statistics)
        orig_count  = self.history_deduplication_info.get('original_count', hist_count)
        removed     = max(0, orig_count - hist_count)
        batch_count = len(self.history_import_summaries)

        sf = QTableWidget(1, 6)
        sf.setHorizontalHeaderLabels(("历史记录", "历史探头", "去重前", "累计去重", "导入次数", "自动累计"))
        sf.verticalHeader().setVisible(False)
        sf.setEditTriggers(QAbstractItemView.NoEditTriggers)
        sf.setSelectionMode(QAbstractItemView.NoSelection)
        sf.setFocusPolicy(Qt.NoFocus)
        sf.setShowGrid(False)
        sf.setFrameShape(QFrame.NoFrame)
        sf.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        sf.setFixedHeight(112)
        sf.setStyleSheet("""
            QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
            QHeaderView::section {
                background: #FAFBFC;
                color: #475467;
                padding: 8px;
                border: none;
                border-bottom: 1px solid #E5E7EB;
                font-weight: bold;
            }
            QTableWidget::item { padding: 12px; font-size: 18px; font-weight: bold; color: #163F72; }
        """)
        for col, value in enumerate((
            f"{hist_count} 条", f"{hist_probes} 个", f"{orig_count} 条",
            f"{removed} 条", f"{batch_count} 次", "已开启" if self._history_enabled else "已关闭",
        )):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            sf.setItem(0, col, item)
        lay.addWidget(sf)

        editor = QTextEdit()
        editor.setReadOnly(True)
        editor.setFont(QFont("Microsoft YaHei", 10))
        editor.setStyleSheet("QTextEdit { background: #FFFFFF; border: 1px solid #E5E7EB; padding: 8px; }")
        top_stats = sorted(self.history_statistics.values(), key=lambda s: s.total_duration_minutes, reverse=True)[:12]
        lines = ["历史累计样本预览\n"]
        if removed > 0:
            lines.append(f"累计导入原始记录 {orig_count} 条，当前历史仓保留 {hist_count} 条有效记录，共去重 {removed} 条。\n")
        for i, stat in enumerate(top_stats, 1):
            lines.append(f"{i}. {stat.probe_sn} | {stat.probe_type} | {stat.model} | 总寿命 {stat.total_duration_minutes/60:.2f}h | 使用 {stat.total_uses} 次")
        if self.history_import_summaries:
            lines.append("\n最近导入摘要")
            for i, s in enumerate(reversed(self.history_import_summaries[-5:]), 1):
                lines.append(f"{i}. {s.get('imported_at','')} | 原始 {s.get('original_count',0)} 条 | 历史新增 {s.get('added_to_history_count',0)} 条")
        editor.setPlainText('\n'.join(lines))
        lay.addWidget(editor, 1)

        btn_row = QHBoxLayout()
        bs = _btn("切换到历史累计", "btn_secondary")
        bs.clicked.connect(lambda: [self._set_data_scope('history'), dlg.accept()])
        bs.setMinimumHeight(34)
        bc = _btn("关闭", "btn_primary"); bc.clicked.connect(dlg.accept)
        bc.setMinimumHeight(34)
        bc.setMinimumWidth(92)
        btn_row.addWidget(bs); btn_row.addStretch(1); btn_row.addWidget(bc)
        lay.addLayout(btn_row)
        self._present_dialog(dlg, key="history_details")

    # ═══════════════════════════════════════════════════════════
    #  对话框：软件说明
    # ═══════════════════════════════════════════════════════════
    def _show_software_notes(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("软件说明")
        dlg.resize(880, 580)
        self._configure_dialog_window(dlg, min_size=(720, 440))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel("统计与展示说明")
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #163F72; background: transparent;")
        lay.addWidget(t)

        s = QLabel("以下规则已应用到文件读取、寿命统计、异常判断和界面展示。")
        s.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
        s.setWordWrap(True)
        lay.addWidget(s)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea { border: none; background: transparent; }
            QScrollBar:vertical { width: 8px; }
            QScrollBar::handle:vertical { background: #D1D5DB; border-radius: 4px; }
            QScrollBar::handle:vertical:hover { background: #9CA3AF; }
        """)
        
        body = QFrame()
        body.setStyleSheet("background: transparent; border: none;")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(0, 4, 6, 4)
        bl.setSpacing(4)
        
        for i, note in enumerate(SOFTWARE_NOTES, 1):
            row_wrap = QFrame()
            row_wrap.setStyleSheet("background: #FFFFFF; border: 1px solid #EEF2F6;")
            row = QHBoxLayout(row_wrap)
            row.setContentsMargins(14, 12, 14, 12)
            row.setSpacing(14)
            
            num = QLabel(f"{i}")
            num.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            num.setStyleSheet("color: #245EA9; background: #EEF4FB; border: none;")
            num.setFixedSize(28, 28)
            num.setAlignment(Qt.AlignCenter)
            row.addWidget(num)
            
            lbl = QLabel(note)
            lbl.setFont(QFont("Microsoft YaHei", 10))
            lbl.setStyleSheet("color: #1F2937; background: transparent; line-height: 1.5;")
            lbl.setWordWrap(True)
            row.addWidget(lbl, 1)
            bl.addWidget(row_wrap)
        
        bl.addStretch(1)
        scroll_area.setWidget(body)
        lay.addWidget(scroll_area, 1)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        bc = _btn("关闭", "btn_primary")
        bc.setMinimumHeight(34)
        bc.setMinimumWidth(92)
        bc.clicked.connect(dlg.accept)
        btn_layout.addWidget(bc)
        lay.addLayout(btn_layout)
        
        self._present_dialog(dlg, key="software_notes")

    # ═══════════════════════════════════════════════════════════
    #  对话框：数据安全提示
    # ═══════════════════════════════════════════════════════════
    def _show_data_safety_warning(self, force: bool = False):
        if not force and self._safety_shown:
            return
        self._safety_shown = True
        dlg = QDialog(self)
        dlg.setWindowTitle("数据安全提示")
        dlg.resize(760, 440)
        self._configure_dialog_window(dlg, min_size=(560, 320))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel("数据安全提示")
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #C4321F; background: transparent;")
        lay.addWidget(t)

        s = QLabel("支持读取网络服务器或共享目录中的数据，程序只读访问，不会修改源文件。")
        s.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
        s.setWordWrap(True)
        lay.addWidget(s)

        body = QTableWidget(len(SAFETY_REMINDERS), 2)
        body.setHorizontalHeaderLabels(("序号", "说明"))
        body.setEditTriggers(QAbstractItemView.NoEditTriggers)
        body.setSelectionMode(QAbstractItemView.NoSelection)
        body.setFocusPolicy(Qt.NoFocus)
        body.setShowGrid(False)
        body.setFrameShape(QFrame.NoFrame)
        body.verticalHeader().setVisible(False)
        body.horizontalHeader().setStretchLastSection(True)
        body.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        body.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        body.setColumnWidth(0, 56)
        body.setStyleSheet("""
            QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
            QHeaderView::section {
                background: #FAFBFC;
                color: #475467;
                padding: 8px;
                border: none;
                border-bottom: 1px solid #E5E7EB;
                font-weight: bold;
            }
            QTableWidget::item { padding: 8px 10px; border-bottom: 1px solid #EEF2F6; }
        """)
        for idx, reminder in enumerate(SAFETY_REMINDERS, 1):
            num_item = QTableWidgetItem(str(idx))
            num_item.setTextAlignment(Qt.AlignCenter)
            num_item.setForeground(QColor("#D92D20"))
            text_item = QTableWidgetItem(reminder)
            text_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            text_item.setForeground(QColor("#7A271A"))
            body.setItem(idx - 1, 0, num_item)
            body.setItem(idx - 1, 1, text_item)
        lay.addWidget(body, 1)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        bc = _btn("我已了解", "btn_primary")
        bc.setMinimumHeight(34)
        bc.setMinimumWidth(96)
        bc.clicked.connect(dlg.accept)
        btn_layout.addWidget(bc)
        lay.addLayout(btn_layout)
        self._present_dialog(dlg, key="data_safety_warning")

    # ═══════════════════════════════════════════════════════════
    #  对话框：去重信息
    # ═══════════════════════════════════════════════════════════
    def _show_deduplication_info(self):
        if self._current_scope == 'history' and self.history_deduplication_info:
            ded = self.history_deduplication_info
            scope_title = "历史累计去重信息"
        elif self.session_deduplication_info:
            ded = self.session_deduplication_info
            scope_title = "当前批次去重信息"
        else:
            QMessageBox.information(self, "提示", "暂无去重数据。\n\n请先导入并分析 Excel 文件。")
            return

        original = ded.get('original_count', 0)
        removed  = ded.get('removed_count', 0)
        unique   = len(self.history_records if self._current_scope == 'history' else self.session_records)
        removed_records = ded.get('removed_records', [])

        dlg = QDialog(self)
        dlg.setWindowTitle(scope_title)
        dlg.resize(1240, 730)
        self._configure_dialog_window(dlg, min_size=(900, 550))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel("数据去重统计")
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #0E5C8B; background: transparent;")
        lay.addWidget(t)

        desc = QLabel("系统按整行字段完全一致的规则自动去重。以下为当前数据范围内的统计结果。")
        desc.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
        desc.setWordWrap(True)
        lay.addWidget(desc)

        summary = QTableWidget(1, 3)
        summary.setHorizontalHeaderLabels(("原始记录数", "去重后数量", "移除重复数"))
        summary.verticalHeader().setVisible(False)
        summary.setEditTriggers(QAbstractItemView.NoEditTriggers)
        summary.setSelectionMode(QAbstractItemView.NoSelection)
        summary.setFocusPolicy(Qt.NoFocus)
        summary.setShowGrid(False)
        summary.setFrameShape(QFrame.NoFrame)
        summary.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        summary.setFixedHeight(116)
        summary.setStyleSheet("""
            QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
            QHeaderView::section {
                background: #FAFBFC;
                color: #475467;
                padding: 8px;
                border: none;
                border-bottom: 1px solid #E5E7EB;
                font-weight: bold;
            }
            QTableWidget::item { padding: 12px; font-size: 24px; font-weight: bold; }
        """)
        for col, (value, color) in enumerate(((str(original), "#EB6E00"), (str(unique), "#0E5C8B"), (str(removed), "#2F8F46"))):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            item.setForeground(QColor(color))
            summary.setItem(0, col, item)
        lay.addWidget(summary)

        detail_title = QLabel("被移除的重复记录")
        detail_title.setFont(QFont("Microsoft YaHei", 13, QFont.Bold))
        detail_title.setStyleSheet("color: #1F2937; background: transparent;")
        lay.addWidget(detail_title)

        if removed > 0 and removed_records:
            cols = ('编号', '探头编号', '探头型号', '大修', 'SG编号', '数据组', '操作员', '管道数', '开始时间', '结束时间')
            tbl = QTableWidget(len(removed_records), len(cols))
            tbl.setHorizontalHeaderLabels(cols)
            tbl.setAlternatingRowColors(True)
            tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
            tbl.setFocusPolicy(Qt.NoFocus)
            tbl.setShowGrid(False)
            tbl.setFrameShape(QFrame.NoFrame)
            tbl.setStyleSheet("""
                QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
                QHeaderView::section { 
                    background: #F3F4F6; 
                    color: #1F2937; 
                    padding: 8px; 
                    border: none; 
                    border-bottom: 2px solid #D1D5DB;
                    font-weight: bold;
                    font-size: 12px;
                }
                QTableWidget::item { padding: 6px; border: none; }
                QTableWidget::item:selected { background: #DBE5F7; }
            """)
            header = tbl.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)
            header.setStretchLastSection(False)
            header.setMinimumSectionSize(80)
            header.setDefaultSectionSize(130)
            tbl.verticalHeader().setVisible(False)
            for i, r in enumerate(removed_records):
                for j, v in enumerate([
                    str(i+1), getattr(r, 'probe_sn', '') or '—', getattr(r, 'model', '') or '—',
                    getattr(r, 'outage', '') or '—', getattr(r, 'sg_id', '') or '—',
                    getattr(r, 'data_group', '') or '—', getattr(r, 'operator', '') or '—',
                    str(getattr(r, 'tube_number', '') or '—'),
                    str(getattr(r, 'start_time', '') or '—'), str(getattr(r, 'end_time', '') or '—'),
                ]):
                    item = QTableWidgetItem(v)
                    item.setTextAlignment(Qt.AlignCenter)
                    bg_color = QColor("#F9FAFB") if i % 2 == 0 else QColor("#FFFFFF")
                    item.setBackground(bg_color)
                    tbl.setItem(i, j, item)
            lay.addWidget(tbl, 1)
        else:
            no_data_frame = QFrame()
            no_data_frame.setStyleSheet("background: #FAFCFA; border: 1px dashed #B7D9BF;")
            no_data_layout = QVBoxLayout(no_data_frame)
            no_data_layout.setContentsMargins(20, 20, 20, 20)
            no_data = QLabel("当前范围内没有检测到重复记录")
            no_data.setFont(QFont("Microsoft YaHei", 13, QFont.Bold))
            no_data.setStyleSheet("color: #2F8F46; background: transparent;")
            no_data.setAlignment(Qt.AlignCenter)
            no_data_layout.addWidget(no_data)
            lay.addWidget(no_data_frame, 1)

        tip = QLabel("去重规则：只有两条记录所有字段内容完全相同，才会判定为重复并自动移除。")
        tip.setStyleSheet("color: #667085; font-size: 11px; background: transparent;")
        tip.setWordWrap(True)
        lay.addWidget(tip)

        # 按钮区
        bc = _btn("关闭", "btn_primary")
        bc.setMinimumHeight(34)
        bc.setMinimumWidth(92)
        bc.clicked.connect(dlg.accept)
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(bc)
        lay.addLayout(btn_layout)
        
        self._present_dialog(dlg, key=f"dedup_info:{self._current_scope}")

    # ═══════════════════════════════════════════════════════════
    #  对话框：最长连续使用段详情
    # ═══════════════════════════════════════════════════════════
    # ─── 通用详情表格对话框 ───
    def _show_detail_table_dialog(self, title: str, cols: tuple, rows: list,
                                   description: str = '', export_filename: str = ''):
        import pandas as pd
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(1400, 800)
        self._configure_dialog_window(dlg, min_size=(1000, 600))
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 18)
        lay.setSpacing(14)

        t = QLabel(title)
        t.setFont(QFont("Microsoft YaHei", 17, QFont.Bold))
        t.setStyleSheet("color: #174A7E; background: transparent;")
        lay.addWidget(t)
        if description:
            d = QLabel(description)
            d.setStyleSheet("color: #667085; font-size: 12px; background: transparent;")
            d.setWordWrap(True)
            lay.addWidget(d)

        tbl = QTableWidget(len(rows), len(cols))
        tbl.setHorizontalHeaderLabels(cols)
        tbl.setAlternatingRowColors(True)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl.setFocusPolicy(Qt.NoFocus)
        tbl.setShowGrid(False)
        tbl.setFrameShape(QFrame.NoFrame)
        tbl.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        tbl.setStyleSheet("""
            QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; }
            QHeaderView::section { 
                background: #FAFBFC; 
                color: #475467; 
                padding: 8px; 
                border: none; 
                border-bottom: 1px solid #E5E7EB;
                font-weight: bold;
                font-size: 12px;
            }
            QTableWidget::item { padding: 8px; border: none; }
            QTableWidget::item:selected { background: #E9F2FF; color: #1F2937; }
        """)
        header = tbl.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(80)
        header.setDefaultSectionSize(130)
        tbl.verticalHeader().setVisible(False)
        
        for i, row in enumerate(rows):
            for j, col in enumerate(cols):
                val = str(row.get(col, ''))
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter if j != len(cols) - 1 else Qt.AlignLeft | Qt.AlignVCenter)
                # 最长段高亮
                if str(row.get('是否最长段', '')).startswith('★'):
                    item.setBackground(QColor("#FFF8E1"))
                    item.setForeground(QColor("#7B5800"))
                    item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
                else:
                    bg_color = QColor("#F8FAFC") if i % 2 == 1 else QColor("#FFFFFF")
                    item.setBackground(bg_color)
                tbl.setItem(i, j, item)
        lay.addWidget(tbl, 1)

        footer = QHBoxLayout()
        footer.setContentsMargins(0, 0, 0, 0)
        footer.setSpacing(12)
        footer.addWidget(QLabel(f"共 {len(rows)} 条记录",
            styleSheet="color: #6B7280; font-size: 12px; font-weight: 500;"))
        footer.addStretch(1)

        def _export():
            # 添加筛选条件标注
            filter_suffix = ''
            active_filters = self._get_active_filter_items()
            if active_filters:
                filter_suffix = f"_【{';'.join(active_filters[:2])}{'...' if len(active_filters) > 2 else ''}】"
            
            path, _ = QFileDialog.getSaveFileName(
                dlg, f"导出{title}",
                f"{export_filename}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*.*)"
            )
            if not path:
                return
            prog = ProgressDialog(f"导出{title}进度", dlg)
            prog.show_and_paint(f"正在准备导出{title}...", 0)
            try:
                prog.set_progress("正在整理导出数据...", 15)
                df = pd.DataFrame(rows, columns=list(cols))
                if path.lower().endswith('.csv'):
                    prog.set_progress("正在写入 CSV 文件...", 55)
                    df.to_csv(path, index=False, encoding='utf-8-sig')
                else:
                    prog.set_progress("正在写入 Excel 文件...", 55)
                    with pd.ExcelWriter(path, engine='openpyxl') as writer:
                        sheet_name = (export_filename[:31] or 'Sheet1')
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
                        ws = writer.sheets[sheet_name]
                        
                        # 美化表头
                        from openpyxl.styles import Font, PatternFill, Alignment
                        hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        hfont = Font(bold=True, color="FFFFFF", size=11)
                        for cell in ws[1]:
                            cell.fill = hf
                            cell.font = hfont
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 居中对齐
                        for row in ws.iter_rows(min_row=2):
                            for cell in row:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 自动列宽
                        for col in ws.columns:
                            ml = max((len(str(c.value or '')) for c in col), default=0)
                            ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 40)
                        
                        # 添加筛选条件说明
                        if active_filters:
                            info_row = ws.max_row + 2
                            ws[f'A{info_row}'] = "筛选条件："
                            ws[f'A{info_row}'].font = Font(bold=True, size=10)
                            for idx, f in enumerate(active_filters):
                                ws[f'A{info_row + idx + 1}'] = f"• {f}"
                                ws[f'A{info_row + idx + 1}'].font = Font(size=9, color="666666")
                prog.set_progress("导出完成", 100)
                prog.accept()
                QMessageBox.information(dlg, "成功", f"已导出到:\n{path}")
            except Exception as exc:
                try:
                    prog.accept()
                except Exception:
                    pass
                QMessageBox.critical(dlg, "错误", f"导出失败:\n{exc}")

        be = _btn("导出详情", "btn_blue")
        be.setMinimumHeight(34)
        be.setMinimumWidth(96)
        be.clicked.connect(_export)
        bc = _btn("关闭", "btn_primary")
        bc.setMinimumHeight(34)
        bc.setMinimumWidth(92)
        bc.clicked.connect(dlg.accept)
        footer.addWidget(be)
        footer.addWidget(bc)
        lay.addLayout(footer)
        
        self._present_dialog(dlg, key=f"detail_table:{title}")

    # ═══════════════════════════════════════════════════════════
    #  警告捕获（logging handler，与旧版完全一致）
    # ═══════════════════════════════════════════════════════════
    def _create_warning_handler(self):
        """创建 logging handler，捕获 extractor/analyzer 的 WARNING 级别日志。"""
        parent = self

        class _GUIWarningHandler(logging.Handler):
            def __init__(self):
                super().__init__(level=logging.WARNING)
                self.setFormatter(logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
                ))

            def emit(self, record):
                try:
                    msg = self.format(record)
                except Exception:
                    msg = record.getMessage()
                try:
                    parent._on_warning_captured(record, msg)
                except Exception:
                    pass

        return _GUIWarningHandler()

    def _on_warning_captured(self, record, formatted_msg: str):
        """处理捕获到的警告日志。"""
        self._capture_warning_message(record.getMessage(), formatted_msg)

    def _capture_warning_message(self, text: str, formatted_msg: str):
        """主线程内合并后台收集到的警告消息。"""
        text = str(text or '')
        formatted_msg = str(formatted_msg or text)
        if formatted_msg not in self._warning_messages:
            self._warning_messages.append(formatted_msg)

        context = self._extract_warning_context_from_text(text)
        warning_type, _ = self._classify_warning_type(text, context)
        probe_sn = context.get('probe_sn') or '未知探头'
        operator = context.get('operator') or ''
        line_number = context.get('line_number') or ''
        if not line_number:
            return
        key = self._find_warning_group_key(context, warning_type) or self._warning_merge_key(probe_sn, operator, line_number)

        if key not in self._warning_by_probe:
            self._warning_by_probe[key] = {
                'probe_sn': probe_sn, 'warnings': [], 'details': [],
                'warning_types': set(),
                'outage': context.get('outage', ''), 'sg_id': context.get('sg_id', ''),
                'data_group': context.get('data_group', ''), 'operator': operator,
                'probe_type': context.get('probe_type', ''), 'model': context.get('model', ''),
                'tube_number': context.get('tube_number', ''),
                'start_time': context.get('start_time', ''), 'end_time': context.get('end_time', ''),
                'line_number': line_number,
            }
        probe_info = self._warning_by_probe[key]
        self._apply_warning_context(probe_info, context)
        probe_info.setdefault('warning_types', set()).add(warning_type)
        if warning_type not in probe_info['warnings']:
            probe_info['warnings'].append(warning_type)
        probe_info['details'].append({'raw_text': text, 'message': formatted_msg, 'type': warning_type})

    def _install_warning_handler(self):
        """在分析开始前安装 logging handler。"""
        self._warning_handler = self._create_warning_handler()
        for name in ('pies_pyqt.core.extractor', 'pies_pyqt.core.analyzer',
                     'pies.extractor', 'pies.analyzer'):
            logging.getLogger(name).addHandler(self._warning_handler)

    def _uninstall_warning_handler(self):
        """在分析完成后移除 logging handler。"""
        if hasattr(self, '_warning_handler'):
            for name in ('pies_pyqt.core.extractor', 'pies_pyqt.core.analyzer',
                         'pies.extractor', 'pies.analyzer'):
                try:
                    logging.getLogger(name).removeHandler(self._warning_handler)
                except Exception:
                    pass

    # ── 辅助方法（与旧版对应）──
    def _format_filter_value_text(self, value) -> str:
        n = self._normalize_filter_value(value)
        if n is None:
            return ''
        if isinstance(n, tuple):
            return '、'.join(n)
        return str(n)

    def _is_table_header_filterable(self, col_name: str) -> bool:
        return col_name in TABLE_HEADER_FILTERABLE_COLUMNS

    def _menu_selected_prefix(self, selected: bool) -> str:
        return '√ ' if selected else '  '

    def _truncate_filter_label(self, value, max_length: int = 10) -> str:
        text = str(value or '').strip()
        if len(text) <= max_length:
            return text
        return f"{text[:max_length - 3]}..."

    def _sanitize_filename_part(self, value: str) -> str:
        safe = re.sub(r'[<>:"/\\|*?]+', '_', (value or '').strip())
        safe = re.sub(r'\s+', '_', safe)
        return safe

    @staticmethod
    def _sanitize_excel_sheet_name(sheet_name: str) -> str:
        safe = re.sub(r'[\[\]\*\?/\\:]+', '_', str(sheet_name or '').strip())
        safe = safe.strip("'")
        return (safe[:31] or 'Sheet1')

    def _json_safe(self, value):
        if isinstance(value, dict):
            return {str(k): self._json_safe(v) for k, v in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [self._json_safe(v) for v in value]
        if hasattr(value, 'isoformat'):
            try:
                return value.isoformat(sep=' ') if value.__class__.__name__ == 'Timestamp' else value.isoformat()
            except Exception:
                pass
        if hasattr(value, 'item'):
            try:
                return self._json_safe(value.item())
            except Exception:
                pass
        return value

    @staticmethod
    def _normalize_warning_context_value(value) -> str:
        """将各种类型的值规范化为字符串，用于警告上下文比较。"""
        if value is None:
            return ''
        if hasattr(value, 'strftime'):
            try:
                return value.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                pass
        text = str(value).strip()
        if text.lower() in {'none', 'nan', 'nat', ''}:
            return ''
        if (
            text in {"'", '"', "'}", '"}', '{}', '[]'}
            or "ID':" in text
            or re.fullmatch(r"['\"\{\}\[\],:]+", text)
        ):
            return ''
        return text

    @classmethod
    def _parse_warning_datetime_value(cls, value) -> Optional[datetime]:
        text = cls._normalize_warning_context_value(value)
        if not text:
            return None
        normalized = text.replace('/', '-').replace('T', ' ')
        normalized = re.sub(r'\.\d+', '', normalized)
        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            pass
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return datetime.strptime(normalized, fmt)
            except ValueError:
                continue
        return None

    def _warning_display_value(self, value, placeholder: str = '\\') -> str:
        text = self._normalize_warning_context_value(value)
        return text if text else placeholder

    def _apply_warning_context(self, target: dict, context: dict | None, overwrite: bool = False):
        if not target or not context:
            return
        for field in ('probe_sn', 'operator', 'line_number', 'outage', 'sg_id',
                      'data_group', 'probe_type', 'model', 'tube_number', 'start_time', 'end_time'):
            incoming = self._normalize_warning_context_value(context.get(field, ''))
            if not incoming:
                continue
            existing = self._normalize_warning_context_value(target.get(field, ''))
            if field == 'probe_sn' and (not existing or existing == '未知探头'):
                target[field] = incoming
            elif field == 'line_number' and existing and not self._warning_context_value_is_bad(existing):
                continue
            elif overwrite or not existing or self._warning_context_value_is_bad(existing):
                target[field] = incoming

    @staticmethod
    def _warning_context_value_is_bad(value: str) -> bool:
        text = str(value or '').strip()
        return (
            not text
            or text in {'\\', '待确认'}
            or text in {"'", '"', "'}", '"}', '{}', '[]'}
            or "ID':" in text
            or text.lower() in {'none', 'nan', 'nat'}
            or bool(re.fullmatch(r"['\"\{\}\[\],:]+", text))
        )

    def _extract_warning_context_from_text(self, text: str) -> dict:
        context = {field: '' for field in (
            'probe_sn', 'operator', 'line_number', 'outage', 'sg_id',
            'data_group', 'probe_type', 'model', 'tube_number', 'start_time', 'end_time'
        )}
        if not text:
            return context

        def _extract(patterns):
            for pattern in patterns:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    return self._normalize_warning_context_value(m.group(1))
            return ''

        context['probe_sn'] = _extract([
            r'(?:探头编号|探头)[=：:]\s*([^\s,，\)]+)',
            r'probe[=：:]\s*([^\s,，\)]+)',
            r'([A-Z]{2,}\d+[A-Z]*\d*|\d{6,})',
        ])
        if context['probe_sn'].isdigit() and len(context['probe_sn']) < 6:
            context['probe_sn'] = ''
        context['operator'] = _extract([r'操作员[=：:\s]\s*([^\s,，\)]+)'])
        context['line_number'] = _extract([r'\(行\s*(\d+)\)', r'源数据行\s*(\d+)', r'行\s*(\d+)'])
        context['outage'] = _extract([r'(?:大修编号|大修)[=：:\s]\s*([^\s,，\)]+)'])
        context['sg_id'] = _extract([r'SG\s*ID[=：:\s]\s*([^\s,，\)]+)', r'SG[=：:\s]\s*([^\s,，\)]+)'])
        context['data_group'] = _extract([r'数据组[=：:\s]\s*([^\s,，\)]+)'])
        context['probe_type'] = _extract([r'探头类型[=：:\s]\s*([^\s,，\)]+)'])
        context['model'] = _extract([r'探头型号[=：:\s]\s*([^\s,，\)]+)'])
        context['tube_number'] = _extract([r'管道数量[=：:\s]\s*([^\s,，\)]+)'])
        context['start_time'] = _extract([
            r'原始开始时间[=：:\s]\s*([^,，]+)', r'原始开始[=：:\s]\s*([^,，]+)',
            r'开始时间[=：:\s]\s*([^,，]+)',
        ])
        context['end_time'] = _extract([
            r'原始结束时间[=：:\s]\s*([^,，]+)', r'原始结束[=：:\s]\s*([^,，]+)',
            r'结束时间[=：:\s]\s*([^,，]+)',
        ])
        return context

    def _warning_type_matches(self, candidate_type: str, warning_types: set) -> bool:
        if not warning_types:
            return True
        candidate = self._normalize_warning_context_value(candidate_type)
        if not candidate:
            return True
        aliases = set()
        for warning_type in warning_types:
            aliases.add(warning_type)
            if warning_type == '记录时间异常':
                aliases.update({'开始时间和结束时间相同', '时间差小于1分钟', '结束时间早于开始时间'})
            if warning_type in {'开始时间和结束时间相同', '时间差小于1分钟', '结束时间早于开始时间'}:
                aliases.add('记录时间异常')
        return candidate in aliases

    def _lookup_warning_error_context(self, warning_group: dict) -> dict:
        records = (
            getattr(self, 'history_error_records', []) or []
            if self._current_scope == 'history'
            else getattr(self, 'session_error_records', []) or []
        )
        if not records:
            return {}
        line_number = self._normalize_warning_context_value(warning_group.get('line_number', ''))
        probe_sn = self._normalize_warning_context_value(warning_group.get('probe_sn', ''))
        operator = self._normalize_warning_context_value(warning_group.get('operator', ''))
        data_group = self._normalize_warning_context_value(warning_group.get('data_group', ''))
        warning_types = set(warning_group.get('warning_types', set()) or set())
        candidates = []
        for record in records:
            record_context = self._warning_context_from_error_record(record)
            record_line = self._normalize_warning_context_value(record_context.get('line_number', ''))
            record_probe = self._normalize_warning_context_value(record_context.get('probe_sn', ''))
            record_operator = self._normalize_warning_context_value(record_context.get('operator', ''))
            record_group = self._normalize_warning_context_value(record_context.get('data_group', ''))
            record_type = self._normalize_warning_context_value(record_context.get('warning_type', ''))
            if line_number and record_line and record_line != line_number:
                continue
            if probe_sn and probe_sn != '未知探头' and record_probe and record_probe != probe_sn:
                continue
            if operator and record_operator and record_operator != operator:
                continue
            if data_group and record_group and record_group != data_group:
                continue
            if not self._warning_type_matches(record_type, warning_types):
                continue
            score = 0
            if line_number and record_line == line_number: score += 8
            if probe_sn and record_probe == probe_sn: score += 6
            if operator and record_operator == operator: score += 3
            if data_group and record_group == data_group: score += 4
            if warning_types and record_type in warning_types: score += 5
            candidates.append((score, record_context))
        if not candidates:
            return {}
        candidates.sort(key=lambda x: x[0], reverse=True)
        return dict(candidates[0][1])

    def _warning_context_from_error_record(self, record: dict) -> dict:
        if not isinstance(record, dict):
            return {}
        return {
            'probe_sn': record.get('探头编号', '') or record.get('probe_sn', '') or record.get('Probe SN', ''),
            'operator': record.get('操作员', '') or record.get('operator', '') or record.get('Operator', ''),
            'line_number': record.get('行号', '') or record.get('line_number', ''),
            'outage': record.get('大修编号', '') or record.get('outage', '') or record.get('Outage', ''),
            'sg_id': record.get('SG ID', '') or record.get('SG_ID', '') or record.get('sg_id', ''),
            'data_group': record.get('数据组', '') or record.get('data_group', '') or record.get('Data Group', ''),
            'probe_type': record.get('探头类型', '') or record.get('probe_type', '') or record.get('Probe Type', ''),
            'model': record.get('探头型号', '') or record.get('model', '') or record.get('Model', ''),
            'tube_number': record.get('管道数量', '') or record.get('tube_number', '') or record.get('Tube Number', ''),
            'start_time': record.get('开始时间', '') or record.get('start_time', '') or record.get('Start Time', ''),
            'end_time': record.get('结束时间', '') or record.get('end_time', '') or record.get('End Time', ''),
            'warning_type': record.get('错误类型', '') or record.get('错误信息', '') or record.get('warning_type', ''),
        }

    def _warning_context_from_detail(self, detail: dict) -> dict:
        raw = ''
        if isinstance(detail, dict):
            raw = detail.get('raw_text', '') or detail.get('message', '')
        if not raw:
            return {}
        if isinstance(raw, dict):
            return self._warning_context_from_error_record(raw)
        text = str(raw)
        stripped = text.strip()
        if stripped.startswith('{') and stripped.endswith('}'):
            try:
                parsed = ast.literal_eval(stripped)
                if isinstance(parsed, dict):
                    return self._warning_context_from_error_record(parsed)
            except Exception:
                pass
        return self._extract_warning_context_from_text(text)

    def _force_complete_warning_group(self, group: dict):
        if not isinstance(group, dict):
            return
        for detail in group.get('details', []) or []:
            self._apply_warning_context(group, self._warning_context_from_detail(detail), overwrite=True)
        self._apply_warning_context(group, self._lookup_warning_error_context(group), overwrite=True)
        self._apply_warning_context(group, self._lookup_warning_record_context(group), overwrite=True)

    def _lookup_warning_record_context(self, warning_group: dict) -> dict:
        probe = self._normalize_warning_context_value(warning_group.get('probe_sn', ''))
        operator = self._normalize_warning_context_value(warning_group.get('operator', ''))
        data_group = self._normalize_warning_context_value(warning_group.get('data_group', ''))
        candidates = []
        for record in (self.current_records or self.session_records or self.history_records or []):
            record_probe = self._normalize_warning_context_value(getattr(record, 'probe_sn', ''))
            record_operator = self._normalize_warning_context_value(getattr(record, 'operator', ''))
            record_group = self._normalize_warning_context_value(getattr(record, 'data_group', ''))
            if probe and probe != '未知探头' and record_probe != probe:
                continue
            if operator and record_operator and record_operator != operator:
                continue
            if data_group and record_group and record_group != data_group:
                continue
            score = 0
            if probe and record_probe == probe:
                score += 8
            if operator and record_operator == operator:
                score += 4
            if data_group and record_group == data_group:
                score += 5
            if getattr(record, 'model', ''):
                score += 2
            if getattr(record, 'outage', ''):
                score += 2
            if getattr(record, 'sg_id', ''):
                score += 2
            candidates.append((score, record))
        if not candidates:
            return {}
        candidates.sort(key=lambda item: item[0], reverse=True)
        record = candidates[0][1]
        return {
            'probe_sn': getattr(record, 'probe_sn', ''), 'operator': getattr(record, 'operator', ''),
            'outage': getattr(record, 'outage', ''), 'sg_id': getattr(record, 'sg_id', ''),
            'data_group': getattr(record, 'data_group', ''),
            'probe_type': getattr(record, 'probe_type_raw', None) or getattr(record, 'probe_type', ''),
            'model': getattr(record, 'model', ''), 'tube_number': getattr(record, 'tube_number', ''),
            'start_time': getattr(record, 'start_time', ''), 'end_time': getattr(record, 'end_time', ''),
        }

    def _enrich_warning_groups(self, warning_groups: dict, include_error_records: bool = False):
        for group in (warning_groups or {}).values():
            if not isinstance(group, dict):
                continue
            group.setdefault('warning_types', set())
            if not isinstance(group['warning_types'], set):
                group['warning_types'] = set(group.get('warning_types') or [])
            group.setdefault('warnings', [])
            group.setdefault('details', [])
            for field in ('probe_sn', 'operator', 'line_number', 'outage', 'sg_id', 'data_group',
                          'probe_type', 'model', 'tube_number', 'start_time', 'end_time'):
                group.setdefault(field, '')
            recalculated = set()
            for detail in group.get('details', []) or []:
                raw = detail.get('raw_text', '') or detail.get('message', '')
                context = self._warning_context_from_detail(detail)
                warning_type, _ = self._classify_warning_type(raw)
                existing_type = self._normalize_warning_context_value(detail.get('type', ''))
                if warning_type == '其他警告' and existing_type:
                    warning_type = existing_type
                detail['type'] = warning_type
                if warning_type:
                    recalculated.add(warning_type)
                self._apply_warning_context(group, context, overwrite=True)
            if recalculated:
                group['warning_types'].update(recalculated)
            if include_error_records:
                self._apply_warning_context(group, self._lookup_warning_error_context(group), overwrite=True)
            self._apply_warning_context(group, self._lookup_warning_record_context(group), overwrite=True)

    def _build_warning_note_text(self, warning_group: dict, warning_desc: str) -> str:
        """构建警告说明文本（与旧版完全一致）。"""
        details_count = len(warning_group.get('details', []) or [])
        outage_display     = self._warning_display_value(warning_group.get('outage', ''))
        sg_display         = self._warning_display_value(warning_group.get('sg_id', ''))
        data_group_display = self._warning_display_value(warning_group.get('data_group', ''))
        line_display       = self._warning_display_value(warning_group.get('line_number', ''))
        start_time_display = self._normalize_warning_context_value(warning_group.get('start_time', ''))
        end_time_display   = self._normalize_warning_context_value(warning_group.get('end_time', ''))
        note_parts = [
            f'大修 {outage_display}', f'SG {sg_display}',
            f'数据组 {data_group_display}', f'源数据行 {line_display}',
        ]
        if '管道数量为0' in warning_desc and self._normalize_warning_context_value(warning_group.get('tube_number', '')) == '0':
            note_parts.append('存在管道数量为0的记录')
        if start_time_display and end_time_display and start_time_display == end_time_display:
            note_parts.append(f'时间点 {start_time_display}')
        else:
            note_parts.append(f"时间范围 {start_time_display or '\\'} ~ {end_time_display or '\\'}")
        note_parts.append(f'关联警告 {details_count} 条')
        required_fields = ('outage', 'sg_id', 'data_group', 'line_number', 'start_time', 'end_time')
        if any(not self._normalize_warning_context_value(warning_group.get(f, '')) for f in required_fields):
            note_parts.append('上下文待确认')
        return '，'.join(note_parts)

    def _classify_warning_type(self, text: str, extracted_context: dict | None = None) -> tuple:
        if '开始时间和结束时间均为空' in text:
            return '开始时间和结束时间均为空', '开始时间和结束时间均为空'
        if '结束时间早于开始时间' in text:
            return '结束时间早于开始时间', '结束时间早于开始时间'
        if '开始时间和结束时间相同' in text:
            return '开始时间和结束时间相同', '开始时间和结束时间相同'
        if '时间差小于1分钟' in text:
            return '时间差小于1分钟', '时间差小于1分钟'
        if '开始时间为空' in text:
            return '开始时间为空', '开始时间为空'
        if '结束时间为空' in text:
            return '结束时间为空', '结束时间为空'
        if '管道数量为0' in text:
            return '管道数量为0', '管道数量为0'
        if '管道数量异常' in text:
            return '管道数量异常', '管道数量异常'
        return '其他警告', '其他警告'

    def _generate_filter_suffix(self) -> str:
        active = {}
        for col, val in (self.summary_filter_values or {}).items():
            if self._filter_value_is_active(val):
                active[col] = val
        kw = self._global_keyword_filter.strip()
        if not active and not kw:
            return ''
        parts = []
        for col in ('操作员', '探头类型', '探头型号', '大修', '蒸汽发生器编号'):
            if col in active:
                short = self._format_filter_value_short_text(active[col])
                label_map = {'操作员': '操作员', '探头类型': '类型', '探头型号': '型号', '大修': '大修', '蒸汽发生器编号': 'SG'}
                parts.append(f"{label_map.get(col, col)}{short}")
        for col, val in active.items():
            if col not in ('操作员', '探头类型', '探头型号', '大修', '蒸汽发生器编号'):
                parts.append(f"{col}{self._format_filter_value_short_text(val)}")
        if kw:
            parts.append(f"关键字{kw[:12]}")
        suffix = '_'.join(parts)
        suffix = re.sub(r'[<>:"/\\|?*]+', '_', suffix)
        return suffix[:50] if len(suffix) > 50 else suffix

    def _style_exported_worksheet(self, worksheet, header_fill, header_font):
        from openpyxl.styles import Alignment
        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in worksheet.columns:
            ml = max((len(str(c.value or '')) for c in col), default=0)
            worksheet.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 36)

    # ═══════════════════════════════════════════════════════════
    #  窗口关闭
    # ═══════════════════════════════════════════════════════════
    def closeEvent(self, event):
        if self._chart_refresh_timer:
            self._chart_refresh_timer.stop()
        if self._resize_timer:
            self._resize_timer.stop()
        self._save_history_store()
        # 关闭所有 matplotlib 图形，避免内存泄漏
        try:
            plt.close('all')
        except Exception:
            pass
        try:
            self._pool.waitForDone(1500)
        except Exception:
            logger.debug("线程池关闭等待失败", exc_info=True)
        event.accept()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._pending_chart_refresh and self._current_view != "table":
            self._schedule_chart_render_ready_checks()
